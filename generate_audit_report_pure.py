#!/usr/bin/env python3
"""
BrightView Branch Audit Report Generator — Pure Python
=======================================================
Reads a completed audit Excel workbook, identifies all Fail tests,
populates pre-written observation templates, and injects everything
directly into the PPTX report.

Usage:
    python generate_audit_report_pure.py YourBranch.xlsx
    python generate_audit_report_pure.py YourBranch.xlsx --period "March 2026" --month "May 2026"

Files (default folder: ~/Downloads):
    Input Excel   ~/Downloads/YourBranch.xlsx
    Template PPTX ~/Downloads/template.pptx   <- save OC example here
    Output PPTX   ~/Downloads/<Branch>_Audit_Report.pptx

Requirements:
    pip install openpyxl
"""

import sys, os, re, shutil, zipfile, argparse, tempfile
from pathlib import Path
from datetime import datetime, date
import openpyxl

DOWNLOADS = Path.home() / "Downloads"

# ── Rating colors ──────────────────────────────────────────────────────────
# Colors matched exactly to slide 14 definitions slide
RATING_COLORS = {
    "Exceeds Expectations": "4472C4",   # blue
    "Satisfactory":         "92D050",   # green (from slide 14)
    "Needs Improvement":    "FFC000",   # orange/gold
    "Unsatisfactory":       "FF0000",   # red
}

# ── Overall rating descriptions ──────────────────────────────────────────────
RATING_DESCRIPTIONS = {
    "Exceeds Expectations": (
        "The branch demonstrated strong adherence to the Company’s policies and procedures "
        "across all risk areas, with few or no exceptions identified. The branch is commended "
        "for its commitment to operational excellence."),
    "Satisfactory": (
        "The branch generally adheres to the Company’s policies and procedures. A limited "
        "number of exceptions were identified; however, the overall control environment is "
        "effective. Action plans should be implemented to address the noted observations."),
    "Needs Improvement": (
        "Several areas of non-compliance were identified, indicating gaps in adherence to the "
        "Company’s policies and procedures. Prompt attention and corrective action are "
        "required to strengthen the control environment."),
    "Unsatisfactory": (
        "Multiple areas of non-compliance were identified across risk areas, indicating a "
        "significant deviation from the Company’s policies and procedures. Given the volume "
        "of exceptions found, immediate attention is required to help the branch comply with "
        "the Company’s operational policies and procedures."),
}

# ── Template slide map ─────────────────────────────────────────────────────
# slide index (1-based) -> (risk_area, obs_slots, existing_test_names)
# This matches the OC template structure exactly
SLIDE_MAP = {
    2:  ("Financial",    3),
    3:  ("Financial",    2),
    4:  ("Financial",    2),
    5:  ("Operational",  3),
    6:  ("Compliance",   3),
    7:  ("Compliance",   2),
    8:  ("Safety / HR",  3),
    9:  ("Safety / HR",  2),
}
# Total slots per area
AREA_SLOTS = {"Financial": 7, "Operational": 3, "Compliance": 5, "Safety / HR": 5}
RISK_AREA_ORDER = ["Financial", "Operational", "Compliance", "Safety / HR"]

# ─────────────────────────────────────────────────────────────────────────────
# OBSERVATION TEMPLATES
# ─────────────────────────────────────────────────────────────────────────────

def _pct(v):
    try: return f"{float(v)*100:.0f}%"
    except: return str(v)

def _te_obs(d, p):
    parts = []
    ns = d.get('not_submitted', 0)
    if ns:
        parts.append(f"Based on review of the Concur Expense Reports, noted {ns} expense "
            f"report(s) were in the \u201cNot Submitted\u201d status. Expenses that are not "
            f"submitted timely are not uploaded to E1 via the automatic posting from Concur at "
            f"month-end, creating difficulty assessing job/branch expenses.")
    kw = d.get('keyword_fails', 0)
    if kw:
        parts.append(f"{kw} transaction(s) were identified with insufficient business rationale "
            f"or non-allowable purchases.")
    return " ".join(parts) or ("Based on review of the Concur Expense Reports, exceptions to "
        "the T\u0026E policy were identified.")

def _etc_obs(d, p):
    weeks = d.get('fail_weeks_detail', [])
    if weeks:
        strs = [f"the week ending {w['date']} was {w['pct']}%" for w in weeks]
        return (f"ETC usage for {'; '.join(strs)}, below the required 95% threshold. "
            f"Inquiry with branch personnel identified "
            f"{d.get('rationale','lack of training and compliance')} as the root cause.")
    return ("ETC usage for the sampled weeks fell below the required 95% threshold. "
        "Inquiry with branch personnel identified lack of training as the root cause.")

def _incident_obs(d, p):
    failing = d.get('failing_types', [])
    if failing:
        strs = [f"{t['type']} ({t['pct']}%)" for t in failing]
        return (f"It was noted the branch had a reporting lag exceeding 25% for "
            f"{', '.join(strs)} claims into Riskonnect. In accordance with "
            f"\u201cBEHSP-054 Accident Reporting Policy,\u201d all incidents must be "
            f"reported within 24 hours of the event.")
    return ("It was noted the branch had a reporting lag exceeding the 25% threshold for one "
        "or more incident types in Riskonnect. In accordance with \u201cBEHSP-054 Accident "
        "Reporting Policy,\u201d all incidents must be reported within 24 hours of the event.")

def _poster_obs(d, p):
    missing = d.get('missing_posters', [])
    legacy  = d.get('has_legacy', False)
    parts   = []
    if missing:
        parts.append(f"Auditor observed that {len(missing)} required poster(s) were not "
            f"present: {', '.join(missing)}.")
    if legacy:
        parts.append("A legacy poster (Valleycrest/Brickman) was observed at the branch. "
            "Retaining legacy materials may result in inconsistent messaging with current "
            "strategic branding.")
    return " ".join(parts) or ("Auditor observed that one or more required posters were not "
        "present or were out of date at the branch.")

def _dvir_obs(d, p):
    nsd = d.get('not_same_day', 0)
    nse = d.get('not_same_employee', 0)
    tot = d.get('sample_size', 3)
    parts = []
    if nsd:
        parts.append(f"Based on the review of the DVIR report, {nsd} of {tot} inspections "
            f"sampled had a post-check date that was not on the same day as the pre-check date.")
    if nse:
        parts.append(f"{nse} of {tot} inspection(s) had the pre-check and post-check completed "
            f"by different employees.")
    return " ".join(parts) or ("Based on the review of the DVIR report, exceptions were "
        "identified in the inspections sampled.")

TEMPLATES = {
    "TCPV Review": {
        "objective": "To verify the completeness and accuracy of Accounts Receivable, Net Service Revenue, and WIP.",
        "risk": "Potential financial misalignment of billing, revenue, and contract amounts fail to reconcile",
        "period_type": "MTD",
        "obs_fn": lambda d,p: (f"Based on review of the TCPV report for {p}, {d.get('fail_count','[X]')} of "
            f"{d.get('sample_size',7)} jobs selected had a variance between the contract value and E1. "
            f"The adjustment was not made timely as of the audit."),
        "action_plan": ("Operational Accountant will work with the branch leadership to review status of jobs and "
            "variances between E1 billings, BVE1 service plans, and contracts as part of the month-end close "
            "process, document decisions within the month-end close workbook, and make timely adjustments to "
            "the E1 billing schedule or BVE1 service plan, as necessary."),
    },
    "Lost Jobs": {
        "objective": "To verify that Billings, Service Plans and CRM jobs are cancelled timely to prevent incorrect A/R, Revenue, WIP, and net new reporting.",
        "risk": "Delayed cancellation of Billings, Service Plans may lead to inaccurate Accounts Receivable, Revenue, and Work in Progress figures, impacting financial reporting. Inaccurate CRM data may impact strategic business decision-making",
        "period_type": "Trailing 3 Months",
        "obs_fn": lambda d,p: (f"{d.get('fail_count','[X]')} of {d.get('sample_size','[X]')} jobs sampled "
            f"{'was' if d.get('fail_count',2)==1 else 'were'} not marked as lost in CRM timely "
            f"(e.g., greater than 30 days subsequent to cancellation)."),
        "action_plan": ("Branch leadership will work with Account Managers and Branch Administrators to review "
            "lost jobs and ensure service plans are cancelled, jobs are marked as lost in E1, billings are "
            "suspended, a termination date is entered in E1, and job is marked as lost in CRM."),
    },
    "Changes in PO\u2019s": {
        "objective": "To verify the accuracy of purchase order estimates.",
        "risk": "Inaccurate purchase order (PO) estimates can create challenges in budgeting and forecasting, impact financial statement accuracy, and add an administrative burden to the voucher process.",
        "period_type": "MTD",
        "obs_fn": lambda d,p: (
            f"Based on review of the \u201cChange in PO Value\u201d report for {p}, observed "
            f"{d.get('variance_count','[X]')} POs that were received at amounts that significantly "
            f"varied from the estimate (i.e., greater than $500)."
            + (f" Additionally, {d.get('bypass_count')} PO(s) were entered for less than $1,500 "
               f"but receipted above $1,500, bypassing the appropriate approval workflow."
               if d.get('bypass_count',0)>0 else "")),
        "action_plan": ("Branch leadership will retrain Branch Administrators and Account Managers on the Purchase "
            "Order process. They will reiterate that reasonable estimates should be provided for the value of "
            "the Purchase Order. Purchase Orders needing approval in E1 should be rejected by the Branch "
            "leadership if the Purchase Order value is not appropriate."),
    },
    "T&E Expense Review": {
        "objective": "To verify purchases are reasonable according to policy and supported with receipts (if required).",
        "risk": "Risk of misuse of company funds, unauthorized expenses, and increased susceptibility to fraudulent activities.",
        "period_type": "Trailing 3 Months",
        "obs_fn": _te_obs,
        "action_plan": ("Branch leadership will recommunicate expected adherence to the protocols outlined in the "
            "T&E policy, including appropriate supporting evidence and timely submission. Approvers are required "
            "to thoroughly review each expense and compare it to the attached receipt. Branch Leadership should "
            "review and monitor the weekly Concur compliance reports provided by the Corporate Credit Card team "
            "to identify reports not submitted."),
    },
    "A/R Aging Review": {
        "objective": "To review the Accounts Receivable (A/R) report for collection notes, mitigating the risk of mismanagement and inefficiencies in the collection process.",
        "risk": "Risk of mismanagement and inefficiencies in the collection process, potentially impacting cash flow and financial reporting accuracy",
        "period_type": "MTD",
        "obs_fn": lambda d,p: (f"Based on review of the {p} Accounts Receivable Aging Report, "
            f"{d.get('fail_count','[X]')} invoice(s) over 120 days past due lacked adequate status "
            f"notes in E1 for collection efforts."),
        "action_plan": ("Branch Leadership and Branch Administrator will ensure that collection notes are captured "
            "in E1 in accordance with the guidance issued titled \u201cManaging Your A/R Aging Report-Entering "
            "Collection Notes.\u201d If text limits are reached, previous notes should be summarized so most "
            "current notes are visible."),
    },
    "Ancillary Billing": {
        "objective": "To verify that ancillary billings are recorded timely and consistently throughout the month and that accruals are supported and billed promptly to ensure accurate Accounts Receivable and Revenue.",
        "risk": "Ancillary billings concentrated in the final week of the month or accrued without timely billing may indicate improper billing practices, resulting in inaccurate Accounts Receivable, Revenue, and may result in billings and revenue being recorded in the wrong period.",
        "period_type": "MTD",
        "obs_fn": lambda d,p: (f"Based on a review of the weekly ancillary billing report, noted "
            f"{d.get('week4_pct','[X]%')} of the billings were in the final week of the month. "
            f"Upon inquiry, branch noted {d.get('rationale','volume was a result of untimely ancillary billings')}."),
        "action_plan": ("Branch leadership will re-communicate expectations to Account Managers regarding the timely "
            "notification to Branch Administrators when work is complete and ready for billing. Accruals will be "
            "minimized, reviewed monthly, and reversed and billed within the first week of the subsequent month "
            "to ensure revenue is recognized in the appropriate reporting period."),
    },
    "Sales Taxability Review": {
        "objective": "To validate the accuracy of the taxable v. non-taxable flag in E1, mitigating the risk of inaccurate sales tax calculations.",
        "risk": "Risk of mismanagement and inefficiencies in the collection process, potentially impacting cash flow and financial reporting accuracy",
        "period_type": "MTD",
        "obs_fn": lambda d,p: (f"Based on review of the PO Receipts report, {d.get('fail_count','[X]')} of "
            f"{d.get('sample_size',5)} PO lines sampled had an incorrect taxability assessment in E1."),
        "action_plan": ("Branch Administrator should engage with the Corporate Tax Team to reassess "
            "taxability on the purchases identified."),
    },
    "Time & Pay Review": {
        "objective": "To verify a formal review was performed over the completeness and accuracy of hours worked for branch employees",
        "risk": "Inaccurate recording or incomplete reporting of hours worked for branch employees may result in financial errors, compliance issues, and create challenges in budgeting, forecasting, and managing operations.",
        "period_type": "2 Weeks",
        "obs_fn": lambda d,p: (f"Based on inquiry and review of the Time and Pay Reports, there was no evidence "
            f"of timely review by the Branch Manager for {d.get('fail_weeks','the sampled week(s)')}."),
        "action_plan": ("The branch leadership will physically evidence review (i.e., physical sign-off or email) of "
            "the weekly \u201cTime & Pay\u201d report to demonstrate the review of payroll and job costing "
            "reasonableness prior to payroll submission."),
    },
    '"Custom Activity Detail" Review': {
        "objective": "To verify a formal review was performed over the completeness and accuracy of hours and appropriateness of the job those are recorded to for each employee.",
        "risk": "Failure to ensure timely approval and accuracy of adjustments, coupled with improper allocation of costs to respective jobs, may result in inaccurate financial reporting and ineffective job management.",
        "period_type": "2 Weeks",
        "obs_fn": lambda d,p: (f"Based on inquiry and review of the Custom Activity Detail Reports, there was no "
            f"evidence of review by the Operations Manager for {d.get('fail_weeks','the sampled week(s)')}."),
        "action_plan": ("Branch leadership will retrain Field Operations to ensure they are evidencing their review "
            "(via sign-off or email), obtaining evidence of supervisor and employee approval for hour changes."),
    },
    "QSA Review": {
        "objective": "To assess the implementation of key Management actions aimed to mitigate the operational risk associated with not achieving target contract renewal rates, client relationships, and overall organic growth.",
        "risk": "Failure to achieve target contract renewal rates poses a risk, potentially impacting revenue stability, client relationships, and overall organic growth.",
        "period_type": "MTD",
        "obs_fn": lambda d,p: (f"Observed that only {d.get('qsa_pct','[X]%')} of QSA\u2019s for jobs greater than "
            f"$100K were completed during the period under audit."
            + (f" Per inquiry with the branch, {d.get('rationale')}." if d.get('rationale') else "")),
        "action_plan": ("Branch leadership will retrain Field Operations to reinforce routine performance of QSAs "
            "in accordance with established procedures."),
    },
    "ETC Usage": {
        "objective": "To validate compliance with the Electronic Time Card (ETC) process.",
        "risk": "Resistance to technology adoption may lead to inefficiencies, reduced visibility, and potential mismanagement of timesheets, misallocation of time to jobs, or incomplete/inaccurate payroll expenses.",
        "period_type": "2 Weeks",
        "obs_fn": _etc_obs,
        "action_plan": ("Regional Operations Leader will conduct retraining in field, including communicating "
            "importance of utilizing ETC to branch employees."),
    },
    "Timely Estimate Renewal": {
        "objective": "To determine the % of estimates that are outdated",
        "risk": "Outdated costs can lead to lower profit margins, inaccurate budget and forecasting, and a competitive disadvantage",
        "period_type": "Point in Time",
        "obs_fn": lambda d,p: (f"After reviewing the aging of active job estimates, it was found that "
            f"{d.get('overdue_count','[X]')} job estimates (approximately {d.get('overdue_pct','[X]%')} of "
            f"active jobs) had not been updated in the past year."),
        "action_plan": ("Branch management will actively track upcoming job renewals, ensuring estimated costs and "
            "services are being reviewed and refined on an annual basis."),
    },
    "Fixed Assets Existence": {
        "objective": "To validate the existence of fixed assets.",
        "risk": "PP&E does not exist or has been incorrectly capitalized",
        "period_type": "Point in Time",
        "obs_fn": lambda d,p: (f"{d.get('fail_count','[X]')} of the {d.get('sample_size',3)} sampled assets "
            f"could not be identified or located by the branch during the audit."
            + (f" Of note, {d.get('notes')}." if d.get('notes') else "")),
        "action_plan": ("ROL will assist branch in performing periodic inventory counts of their fixed assets to "
            "identify potentially unprocessed additions, transfers or disposed assets."),
    },
    "Fixed Assets Barcode": {
        "objective": "To validate the correct barcode is properly adhered to the asset.",
        "risk": "Creates inefficiencies in the Fixed Asset Inventory, resulting in manual processes to match asset to the fixed asset register.",
        "period_type": "Point in Time",
        "obs_fn": lambda d,p: (f"{d.get('fail_count','[X]')} of the {d.get('sample_size',3)} assets selected "
            f"for testing did not have a barcode attached to the asset."),
        "action_plan": ("As part of ongoing asset inventory maintenance, the ROL should help the branch apply "
            "barcodes for any assets that are missing barcodes and securely adhered to the assets."),
    },
    "Fixed Assets Completeness": {
        "objective": "To validate fixed assets are appropriately capitalized.",
        "risk": "PP&E exists but either not appropriately capitalized or improperly disposed of.",
        "period_type": "On-Site",
        "obs_fn": lambda d,p: (f"During the yard walk, {d.get('fail_count','[X]')} of "
            f"{d.get('sample_size',3)} assets selected did not have a barcode properly adhered. "
            f"The asset could not be traced back to the fixed asset register via barcode scan."),
        "action_plan": ("Branch management will perform a full inventory count and ensure all assets in the yard "
            "are properly tagged with barcodes. Any assets not reflected in the fixed asset register should be "
            "promptly added or disposals initiated in LEADR."),
    },
    "IT Access Removal": {
        "objective": "To identify untimely removal of terminated employees from the IT system, enhancing overall data security and access control measures within the organization",
        "risk": "Failure to promptly remove system access for terminated employees may pose a security risk and lead to unauthorized access, potentially resulting in legal, financial, or reputational consequences",
        "period_type": "Trailing 3 Months",
        "obs_fn": lambda d,p: (f"System access for {d.get('violation_count','[X]')} terminated employee(s) was "
            f"not removed within one day from their termination date."
            + (" Of note, the employees did not have access to E1 or BVE1."
               if d.get('no_e1_access', True) else "")),
        "action_plan": ("Branch Management will retrain all employees on the IT Termination policy and the "
            "importance of communicating terminations immediately to the Branch Administrator for timely processing."),
    },
    "Personnel File Review": {
        "objective": "Ensure comprehensive compliance with personnel file documentation, focusing on the retention and completeness of records and exclusion of sensitive data.",
        "risk": "Incomplete employee files or inclusion of sensitive data may lead to regulatory issues, especially regarding protections around voluntary disclosure, arbitration, non-competes, and non-solicits",
        "period_type": "Point in Time",
        "obs_fn": lambda d,p: (f"Upon review of {d.get('sample_size','[X]')} personnel files, auditor observed "
            f"that {d.get('fail_count','[X]')} {d.get('missing_doc','required document(s)')} could not be located."),
        "action_plan": ("Branch Administrator will work with Human Resources (HRC or HRBP) to review all branch "
            "employee files and ensure all necessary paperwork is electronically stored in a confidential folder "
            "with limited access."),
    },
    "Vehicle Registration": {
        "objective": "To identify vehicles with expired registrations.",
        "risk": "Failure to maintain up-to-date vehicle registration can lead to legal penalties, insurance issues, operational disruptions, reputation damage, and increased liability.",
        "period_type": "Point in Time",
        "obs_fn": lambda d,p: (f"Based on review of the Vehicle Registration report, "
            f"{d.get('expired_count','[X]')} vehicle(s) had expired registrations as of the testing date."
            + (f" Per inquiry with the branch, {d.get('rationale')}." if d.get('rationale') else "")),
        "action_plan": ("Branch management will work with the relevant vehicle owners to initiate the disposal "
            "process and ensure that any vehicles remaining in active use have current registrations."),
    },
    "Timely Incident Lag Reporting": {
        "objective": "To verify that accidents, incidents, injuries are reported timely to the Crew Leader/Supervisor or Branch Management and an accident report is completed within 24 hours of the incident.",
        "risk": "Untimely incident reporting may impact employee safety, disrupt operations, and/or result in financial implications linked to increased costs associated with insurance claims.",
        "period_type": "Trailing 6 Months",
        "obs_fn": _incident_obs,
        "action_plan": ("Branch management will retrain branch personnel to ensure compliance with the "
            "\u201cBEHSP-054 Accident Reporting Policy\u201d to report accidents, incidents, or injuries "
            "within 24 hours into Riskonnect."),
    },
    "Safety Training Compliance": {
        "objective": "To verify branch compliance with required onboarding safety trainings and OSHA training requirements.",
        "risk": "Risk relative to well-being of employees and overall stability of the organization (e.g., ability to do business, costs, and reputational risk).",
        "period_type": "Point in Time",
        "obs_fn": lambda d,p: (f"Per review of the BrightPath Safety Training compliance dashboard, only "
            f"{d.get('trained_pct','[X]%')} of branch team members had completed their key safety trainings "
            f"as of the end of {p}."),
        "action_plan": ("Branch leadership, including the Safety Leader, should ensure that 100% of team members "
            "have completed required trainings within 90 days of hire."),
    },
    "Awareness (Poster) Compliance": {
        "objective": "To validate that the branch displays and maintains required safety, HR, and labor law posters visibly on its walls, ensuring compliance with regulations and organizational policies.",
        "risk": "Risk relative to well-being of employees and overall stability of the organization (e.g., ability to do business, liability, and reputational risk).",
        "period_type": "Point in Time",
        "obs_fn": _poster_obs,
        "action_plan": ("Branch Manager will retrain Branch Administrator on internal and regulatory poster "
            "requirements and Branch Administrator will ensure posters and signs are up-to-date and visible."),
    },
    "PPE and Safety Observation": {
        "objective": "To assess adherence to the established safety protocol, confirming the proper utilization of safety vests as a protective measure within the designated area.",
        "risk": "Risk relative to well-being of employees and overall stability of the organization (e.g., ability to do business, costs, and reputational risk).",
        "period_type": "Point in Time",
        "obs_fn": lambda d,p: (f"During the yard walk, {d.get('fail_detail','team members were observed not wearing required PPE')}."),
        "action_plan": ("Branch Leadership will retrain all field employees on PPE requirements and conduct "
            "periodic yard walk observations to ensure ongoing compliance."),
    },
    "Electronic DVIR Compliance": {
        "objective": "To ensure completeness, accuracy, and timely submission of electronic DVIR, mitigating the risk of incomplete inspections and enhancing safety protocols.",
        "risk": "Incomplete electronic DVIR may lead to incomplete inspections, heightened safety risks, and potential resistance to technology adoption.",
        "period_type": "MTD",
        "obs_fn": _dvir_obs,
        "action_plan": ("Branch Leadership will retrain field employees on the DVIR policy and requirements for "
            "pre-check, gate check, and post-check, including the requirement to complete checks on the same "
            "day and by the same employee."),
    },
    "Driver Compliance": {
        "objective": "To ensure completeness and accuracy of the branch\u2019s active driver listing and to verify that all required driver documentation is properly recorded in the system.",
        "risk": "Incomplete or inaccurate active driver listings and missing required documentation in FirstAdvantage may result in regulatory non-compliance, increased safety risk, and operational or financial impacts.",
        "period_type": "Point in Time",
        "obs_fn": lambda d,p: (
            f"Based on review of FirstAdvantage, {d.get('compliant','[X]')} of "
            f"{d.get('total','[X]')} drivers ({d.get('compliant_pct','[X]%')}) were compliant. "
            f"{d.get('non_compliant','[X]')} driver(s) did not have required documentation to support driver eligibility."
            + (f" Additionally, {d.get('leadr_missing')} drivers were not included within LEADR."
               if d.get('leadr_missing',0)>0 else "")),
        "action_plan": ("Branch leadership will reinforce shared accountability for safety and the importance of "
            "accurate driver records and will oversee processes to ensure timely follow-up on non-compliance, "
            "routine monitoring of First Advantage, and removal of non-compliant drivers from driving duties "
            "until requirements are satisfied."),
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL DATA EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def xl_date(val):
    if isinstance(val, (datetime, date)): return val.strftime("%m/%d/%Y")
    try:
        from datetime import timedelta
        d = date(1899,12,30) + timedelta(days=int(float(val)))
        return d.strftime("%m/%d/%Y")
    except: return str(val)

def pct_str(v):
    try: return f"{float(v)*100:.0f}%"
    except: return str(v)

def extract_data(wb, tab):
    if tab not in wb.sheetnames: return {}
    ws = wb[tab]
    rows = list(ws.iter_rows(values_only=True))
    d = {}

    if tab == "3.2":
        for i,row in enumerate(rows):
            label = str(row[1]).strip() if len(row)>1 and row[1] else ""
            if label == "# Variances > $500" and i+1<len(rows):
                dr = rows[i+1]
                try:
                    d['variance_count'] = int(float(dr[1])) if len(dr)>1 and dr[1] else 0
                    d['total_pos']      = int(float(dr[2])) if len(dr)>2 and dr[2] else 0
                except: pass
            if "# of PO's not entered accurately" in label and i+1<len(rows):
                dr = rows[i+1]
                try:
                    d['bypass_count']    = int(float(dr[1])) if len(dr)>1 and dr[1] else 0
                    d['bypass_variance'] = int(float(dr[2])) if len(dr)>2 and dr[2] else 0
                except: pass
        d['sample_size'] = "All"

    elif tab == "4":
        for i,row in enumerate(rows):
            row_str = " ".join(str(c) for c in row if c is not None)
            if "TS04" in row_str and "# of Transactions" in row_str and i+1<len(rows):
                dr = rows[i+1]
                vals = [c for c in dr if c is not None]
                for v in reversed(vals):
                    try:
                        n = int(float(v))
                        if 0<n<100: d['not_submitted']=n; break
                    except: pass
                break

    elif tab == "9":
        for i,row in enumerate(rows):
            row_str = " ".join(str(c) for c in row if c is not None)
            if ("Input %" in row_str or "As of Date" in row_str) and i+1<len(rows):
                for cell in rows[i+1]:
                    try:
                        v=float(cell)
                        if 0<v<=1: d['qsa_pct']=pct_str(v); break
                    except: pass
            if ("If % is less than" in row_str or "inquire with Branch" in row_str.lower()) and i+1<len(rows):
                rat = " ".join(str(c) for c in rows[i+1] if c is not None).strip()
                if rat and "step" not in rat.lower(): d['rationale']=rat[:100]

    elif tab == "10":
        fail_weeks=[]
        for row in rows:
            row_str = " ".join(str(c) for c in row if c is not None)
            if "Fail" in row_str:
                date_v=pct_v=None
                for cell in row:
                    if cell is None or str(cell) in ["Fail","Pass","Do Not Edit"]: continue
                    try:
                        v=float(cell)
                        if v>1000: date_v=xl_date(v)
                        elif 0<v<=2: pct_v=f"{v*100:.0f}"
                    except: pass
                if pct_v: fail_weeks.append({'date':date_v or "sampled week",'pct':pct_v})
        d['fail_weeks_detail']=fail_weeks

    elif tab == "17":
        for i,row in enumerate(rows):
            row_str = " ".join(str(c) for c in row if c is not None)
            if "# of Expired" in row_str:
                for j,cell in enumerate(row):
                    if j==0: continue
                    try:
                        v=int(float(cell))
                        if v>=0: d['expired_count']=v; break
                    except: pass
            if "disposal" in row_str.lower():
                d['rationale']="vehicles are on the disposal list and registrations will not be renewed"

    elif tab == "18":
        for row in rows:
            if not row or row[0] is None: continue
            label=str(row[0]).strip()
            if label in ['Auto','Injury','Property']:
                try:
                    lag=float(row[1]) if row[1] else 0
                    if lag>0.25:
                        d.setdefault('failing_types',[]).append({'type':label,'pct':f"{lag*100:.0f}"})
                except: pass

    elif tab == "19":
        for i,row in enumerate(rows):
            row_str = " ".join(str(c) for c in row if c is not None)
            if ("Total Employees" in row_str or "% Trained" in row_str) and i+1<len(rows):
                for cell in rows[i+1]:
                    try:
                        v=float(cell)
                        if 0<v<1: d['trained_pct']=pct_str(v); break
                    except: pass

    elif tab == "20":
        missing=[]; legacy=False
        for row in rows:
            if not row or row[0] is None: continue
            poster=str(row[0]).strip()
            visible=str(row[1]).strip().lower() if len(row)>1 and row[1] else ""
            row_str=" ".join(str(c) for c in row if c is not None)
            is_fail="Fail" in row_str or visible=="no"
            if is_fail:
                if "OSHA 300A" in poster or "out of timeframe" in str(row).lower(): continue
                if "Legacy" in poster and visible=="yes": legacy=True
                elif "Legacy" not in poster and "OSHA" not in poster: missing.append(poster[:60])
        d['missing_posters']=missing; d['has_legacy']=legacy

    elif tab == "22":
        nsd=0; nse=0
        for row in rows:
            row_str=" ".join(str(c) for c in row if c is not None)
            if "Fail" in row_str and "Test" not in row_str and "Results" not in row_str: nsd+=1
            if len(row)>1 and str(row[1]).strip()=="No" and "Same Employee" not in row_str: nse+=1
        d['not_same_day']=nsd; d['not_same_employee']=nse; d['sample_size']=3

    elif tab == "23":
        for i,row in enumerate(rows):
            if not row: continue
            row_str=" ".join(str(c) for c in row if c is not None)
            if "Total Employees" in row_str and "Compliant Employees" in row_str and i+1<len(rows):
                vals=[c for c in rows[i+1] if c is not None]
                try:
                    total=int(float(vals[0])); compliant=int(float(vals[1]))
                    d['total']=total; d['compliant']=compliant
                    d['non_compliant']=max(0,total-compliant)
                    if total>0: d['compliant_pct']=pct_str(compliant/total)
                    break
                except: pass

    elif tab == "15":
        for row in rows:
            row_str=" ".join(str(c) for c in row if c is not None)
            if "# of Separation Violations" in row_str:
                for cell in row:
                    try:
                        v=int(float(cell))
                        if v>=0: d['violation_count']=v; break
                    except: pass
        d['no_e1_access']=True

    return d


def read_finish_tab(wb):
    ws=wb["FINISH"]; scores={}; overall=None
    for row in ws.iter_rows(values_only=True):
        if not row or not row[1]: continue
        label=str(row[1]).strip()
        if label in ["Financial","Operational","Compliance","Safety / HR"]:
            try:
                sr=row[5]; rt=str(row[6]).strip() if len(row)>6 and row[6] else ""
                if sr is not None:
                    scores[label]={"score_decimal":float(sr),"score_pct":f"{float(sr)*100:.0f}%","rating":rt}
            except: pass
        if "Calculated Overall Audit Rating" in label:
            rc=row[6] if len(row)>6 else None
            if rc and str(rc).strip() in ["Satisfactory","Exceeds Expectations","Needs Improvement","Unsatisfactory"]:
                overall=str(rc).strip()
    return scores, overall


def get_branch_info(wb, filepath):
    fn = Path(filepath).stem
    # Strip common junk suffixes from filenames like "35210_Homestead_-_Claude_to_produce..."
    fn_clean = re.sub(
        r'[-_](Claude|to_produce|produce|powerpoint|from_this|from|this|Template|'
        r'Audit|Draft|DRAFT|Example|for_Claude)[\w_-]*$', '', fn, flags=re.I)
    fn_clean = fn_clean.replace('-', ' ').replace('_', ' ').strip()

    # Try leading branch number (e.g. "35210 BVLS Homestead" or "35210 Homestead")
    m = re.match(r'(\d{4,6})[\s_]+(.+)', fn_clean)
    if m:
        num  = m.group(1)
        name = m.group(2).strip()
        # If name already has BVLS prefix keep it, otherwise add it
        if 'BVLS' not in name.upper():
            return num, f"{num} BVLS {name}"
        return num, f"{num} {name}"

    # No leading number — use cleaned name, prompt user to rename file or use --branch
    return "XXXXX", fn_clean


def scan_for_exceptions(wb):
    TAB_MAP={
        "1":   ("TCPV Review",               "Financial"),
        "2":   ("Lost Jobs",                 "Financial"),
        "3.2": ("Changes in PO\u2019s",      "Financial"),
        "4":   ("T&E Expense Review",        "Financial"),
        "5.1": ("A/R Aging Review",          "Financial"),
        "5.2": ("Ancillary Billing",         "Financial"),
        "6":   ("Sales Taxability Review",   "Financial"),
        "7":   ("Time & Pay Review",         "Operational"),
        "8":   ('"Custom Activity Detail" Review',"Operational"),
        "9":   ("QSA Review",                "Operational"),
        "10":  ("ETC Usage",                 "Operational"),
        "11":  ("Timely Estimate Renewal",   "Operational"),
        "12":  ("Fixed Assets Existence",    "Compliance"),
        "13":  ("Fixed Assets Barcode",      "Compliance"),
        "14":  ("Fixed Assets Completeness", "Compliance"),
        "15":  ("IT Access Removal",         "Compliance"),
        "16":  ("Personnel File Review",     "Compliance"),
        "17":  ("Vehicle Registration",      "Compliance"),
        "18":  ("Timely Incident Lag Reporting","Safety / HR"),
        "19":  ("Safety Training Compliance",   "Safety / HR"),
        "20":  ("Awareness (Poster) Compliance","Safety / HR"),
        "21":  ("PPE and Safety Observation",   "Safety / HR"),
        "22":  ("Electronic DVIR Compliance",   "Safety / HR"),
        "23":  ("Driver Compliance",            "Safety / HR"),
    }
    exceptions={}
    for tab,(test_name,risk_area) in TAB_MAP.items():
        if tab not in wb.sheetnames: continue
        ws=wb[tab]
        for row in ws.iter_rows(values_only=True):
            if not row: continue
            if any(str(c).strip()=="Fail" for c in row if c is not None):
                data=extract_data(wb,tab)
                tmpl=TEMPLATES.get(test_name,{})
                exceptions[test_name]={
                    "risk_area":risk_area,"tab":tab,"data":data,
                    "objective":tmpl.get("objective",""),
                    "risk":tmpl.get("risk",""),
                    "period_type":tmpl.get("period_type","MTD"),
                    "observation":"","action_plan":tmpl.get("action_plan",""),
                    "financial_impact":"$-",
                }
                break
    return exceptions


# ─────────────────────────────────────────────────────────────────────────────
# XML HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def xml_esc(text):
    if not text: return ""
    for old,new in [("&","&amp;"),("<","&lt;"),(">","&gt;"),
                    ("\u2018","&#x2018;"),("\u2019","&#x2019;"),
                    ("\u201c","&#x201C;"),("\u201d","&#x201D;"),
                    ("\u2013","&#x2013;"),("\u2014","&#x2014;")]:
        text=text.replace(old,new)
    return text


def make_bullet_txbody(text):
    safe = xml_esc(str(text)) if text else ""
    return (
        '<a:txBody>\n'
        '                    <a:bodyPr/>\n'
        '                    <a:lstStyle/>\n'
        '                    <a:p>\n'
        + (
        '                      <a:pPr marL="171450" indent="-171450" algn="l" defTabSz="457200" '
        'rtl="0" eaLnBrk="1" latinLnBrk="0" hangingPunct="1">\n'
        '                        <a:buFont typeface="Arial" panose="020B0604020202020204" '
        'pitchFamily="34" charset="0"/>\n'
        '                        <a:buChar char="\u2022"/>\n'
        '                      </a:pPr>\n'
        '                      <a:r>\n'
        '                        <a:rPr lang="en-US" sz="850" kern="1200" dirty="0">\n'
        '                          <a:solidFill><a:schemeClr val="tx1"/></a:solidFill>\n'
        '                          <a:latin typeface="+mn-lt"/>\n'
        '                          <a:ea typeface="+mn-ea"/>\n'
        '                          <a:cs typeface="+mn-cs"/>\n'
        '                        </a:rPr>\n'
        f'                        <a:t>{safe}</a:t>\n'
        '                      </a:r>\n'
        if safe else
        '                      <a:endParaRPr lang="en-US" dirty="0"/>\n'
        )
        + '                    </a:p>\n'
        '                  </a:txBody>'
    )


def replace_cell_content(content, label_text, slot_index, new_txbody):
    """Replace txBody of the content cell that follows the Nth label cell."""
    pattern = f'<a:t>{re.escape(label_text)}</a:t>'
    positions = [m.start() for m in re.finditer(pattern, content)]
    if slot_index >= len(positions):
        return content
    pos = positions[slot_index]
    label_cell_end = content.find('</a:tc>', pos) + 7
    content_cell_start = content.find('<a:tc', label_cell_end)
    if content_cell_start == -1: return content
    txbody_start = content.find('<a:txBody>', content_cell_start)
    txbody_end   = content.find('</a:txBody>', txbody_start) + 11
    if txbody_start == -1 or txbody_end < 11: return content
    return content[:txbody_start] + new_txbody + content[txbody_end:]


def replace_text_cell(content, old_text, new_text, occurrence=1):
    """Replace a specific <a:t> text value, Nth occurrence."""
    old_tag = f'<a:t>{old_text}</a:t>'
    new_tag = f'<a:t>{xml_esc(new_text)}</a:t>'
    count = 0
    pos = 0
    while True:
        idx = content.find(old_tag, pos)
        if idx == -1: break
        count += 1
        if count == occurrence:
            return content[:idx] + new_tag + content[idx+len(old_tag):]
        pos = idx + 1
    return content


def update_rating_cell(content, old_rating, new_rating, new_color):
    """Replace rating text and its background color."""
    idx = content.find(f'<a:t>{old_rating}</a:t>')
    if idx == -1: return content
    # Replace text
    content = content.replace(f'<a:t>{old_rating}</a:t>',
                               f'<a:t>{new_rating}</a:t>', 1)
    # Find the tcPr fill for this cell and update color
    tc_start = content.rfind('<a:tc>', 0, idx)
    tc_end   = content.find('</a:tc>', idx) + 7
    cell     = content[tc_start:tc_end]
    # Replace srgbClr in tcPr
    new_cell = re.sub(
        r'(<a:tcPr>.*?<a:solidFill>.*?<a:srgbClr val=")[A-F0-9]{6}(")',
        f'\\g<1>{new_color}\\2', cell, flags=re.DOTALL, count=1)
    if new_cell == cell:
        # No existing srgbClr — add one (replace <a:noFill/> or append before </a:tcPr>)
        new_cell = cell.replace(
            '<a:noFill/>',
            f'<a:solidFill><a:srgbClr val="{new_color}"/></a:solidFill>', 1)
    return content[:tc_start] + new_cell + content[tc_end:]


# ─────────────────────────────────────────────────────────────────────────────
# COVER SLIDE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

COVER_SUMMARY_BULLETS = {
    "Financial": [
        ("TCPV Review",               "Non-compliance with TCPV review to make corrections to billing schedules and service plans timely."),
        ("Lost Jobs",                 "Failure to mark job as lost in CRM timely."),
        ("Changes in PO\u2019s",      "Imprecise estimating of POs."),
        ("T&E Expense Review",        "Non-compliance with T&E policy due to untimely report submissions."),
        ("A/R Aging Review",          "Non-compliance with AR policy to update E1 notes for invoices aged over 60 days."),
        ("Ancillary Billing",         "Failure to bill ancillary work timely."),
        ("Sales Taxability Review",   "Incorrect sales taxability assessment."),
    ],
    "Operational": [
        ("Time & Pay Review",                   "Failure of approval of the Time and Pay report by the Branch Manager."),
        ('"Custom Activity Detail" Review',     "Non-compliance with Operations Manager Training guide to document employee and supervisor approvals for changes in hours."),
        ("QSA Review",                          "Non-compliance with QSA policies to complete required QSAs."),
        ("ETC Usage",                           "Non-compliance with the Electronic Time Card (ETC) policy."),
        ("Timely Estimate Renewal",             "Failure to refresh overdue Estimates."),
    ],
    "Compliance": [
        ("Fixed Assets Existence",    "Non-compliance with Fixed Asset Existence policy."),
        ("Fixed Assets Barcode",      "Non-compliance with \u2018Fixed Asset Disposal\u2019 and \u2018Barcode\u2019 Policy."),
        ("Fixed Assets Completeness", "Non-compliance with Fixed Asset Completeness policy."),
        ("IT Access Removal",         "Failure to adhere to IT Account Disable policy."),
        ("Personnel File Review",     "Non-compliance with Personnel File Policy."),
        ("Vehicle Registration",      "Non-Compliance with the Company Vehicle Policy to renew expired registrations."),
    ],
    "Safety / HR": [
        ("Timely Incident Lag Reporting",  "Non-compliance with BEHSP-054 Accident Reporting Policy to report incidents promptly in Riskonnect."),
        ("Safety Training Compliance",     "Non-compliance with BrightPath Safety Policies."),
        ("Awareness (Poster) Compliance",  "Non-compliance with posting of mandatory internal and regulatory posters at the branch."),
        ("PPE and Safety Observation",     "Non-compliance with PPE and Safety requirements."),
        ("Electronic DVIR Compliance",     "Non-compliance with DVIR requirements."),
        ("Driver Compliance",              "Non-compliance with required driver qualification documentation."),
    ],
}

def update_cover_summary(content, exceptions, scores, overall_rating):
    """Update the cover slide summary bullets for each risk area."""
    # Each risk area has a cell in the summary table containing bullet points
    # We'll replace the content of each area's observation cell

    # Summary table label map - the table uses 'Safety/HR' not 'Safety / HR'
    summary_label_map = {
        "Financial":   "Financial",
        "Operational": "Operational",
        "Compliance":  "Compliance",
        "Safety / HR": "Safety/HR",
    }

    for area in RISK_AREA_ORDER:
        area_exceptions = [k for k,v in exceptions.items() if v['risk_area']==area]
        if not area_exceptions:
            continue

        # Build bullet text from only the exceptions that fired
        bullet_lines = []
        for test_name, bullet_text in COVER_SUMMARY_BULLETS.get(area, []):
            if test_name in area_exceptions:
                bullet_lines.append(bullet_text)

        if not bullet_lines:
            continue

        # Find the area cell in the SUMMARY table using the correct label
        summary_label = summary_label_map.get(area, area)
        area_pos = content.find(f'<a:t>{summary_label}</a:t>')
        if area_pos == -1:
            area_pos = content.find(f'<a:t>{area}</a:t>')
        if area_pos == -1:
            continue

        # Find the next gridSpan="2" cell after area label - that's the observations cell
        search_from = content.find('</a:tc>', area_pos) + 7
        obs_cell_start = content.find('<a:tc', search_from)
        if obs_cell_start == -1:
            continue

        # Build multi-bullet txBody
        bullet_paras = ""
        for line in bullet_lines:
            safe = xml_esc(line)
            bullet_paras += (
                '                    <a:p>\n'
                '                      <a:pPr marL="171450" indent="-171450">\n'
                '                        <a:buFont typeface="Arial"/>\n'
                '                        <a:buChar char="\u2022"/>\n'
                '                      </a:pPr>\n'
                '                      <a:r>\n'
                '                        <a:rPr lang="en-US" sz="700" dirty="0"/>\n'
                f'                        <a:t>{safe}</a:t>\n'
                '                      </a:r>\n'
                '                    </a:p>\n'
            )

        new_txbody = (
            '<a:txBody>\n'
            '                    <a:bodyPr/>\n'
            '                    <a:lstStyle/>\n'
            + bullet_paras +
            '                  </a:txBody>'
        )

        # Replace txBody of the observations cell
        txbody_start = content.find('<a:txBody>', obs_cell_start)
        txbody_end   = content.find('</a:txBody>', txbody_start) + 11
        if txbody_start != -1 and txbody_end > 11:
            content = content[:txbody_start] + new_txbody + content[txbody_end:]

    return content


# ─────────────────────────────────────────────────────────────────────────────
# PPTX BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def safe_rep(content, old, new, n=1):
    return content.replace(old, new, n) if old in content else content


def build_pptx(template_path, output_path, branch_display, scores,
               overall_rating, findings, period_under_review,
               audit_month, auditor_names, audit_approach="On-Site"):

    work_dir = Path(tempfile.mkdtemp())
    with zipfile.ZipFile(template_path,'r') as z:
        z.extractall(work_dir)

    slides_dir = work_dir/"ppt"/"slides"
    slides = sorted(slides_dir.glob("slide*.xml"),
                    key=lambda p: int(re.search(r'\d+',p.name).group()))
    print(f"  Template: {len(slides)} slides")

    # Replacement strings
    old_names = ["3208 BVLS Orange County Central ","40060 Denver East ","40060 BVLS Denver East ",
                 "35210 BVLS Homestead ","XXXXX [Branch Name] ","XXXXX [Branch Name]"]
    old_dates = ["Branch Audit \u2013 Report \u2013 February 2026",
                 "Branch Audit \u2013 Report \u2013 May 2026",
                 "Branch Audit \u2013 Report \u2013 December 2025",
                 "Branch Audit \u2013 Report \u2013 [Month] 2025",
                 "Branch Audit \u2013 Report \u2013 [Month] 2026"]
    new_hdr  = f"{branch_display} "
    new_date = f"Branch Audit \u2013 Report \u2013 {audit_month}"
    yr = audit_month.split()[-1] if audit_month else "2026"
    mo = audit_month.split()[0]  if audit_month else "May"

    # Score replacement map - old values that may appear in template
    area_old_scores = {
        "Financial":   ["68%","81%","89%","90%","61%","75%"],
        "Operational": ["76%","85%","80%","67%","90%","55%"],
        "Compliance":  ["88%","100%","96%","70%","55%"],
        "Safety / HR": ["41%","90%","72%","48%"],
    }

    # Rating text that may appear in template
    area_old_ratings = {
        "Financial":   ["Needs Improvement","Satisfactory","Exceeds Expectations","Unsatisfactory"],
        "Operational": ["Needs Improvement","Satisfactory","Exceeds Expectations","Unsatisfactory"],
        "Compliance":  ["Unsatisfactory","Needs Improvement","Satisfactory","Exceeds Expectations"],
        "Safety / HR": ["Unsatisfactory","Needs Improvement","Satisfactory","Exceeds Expectations"],
    }

    # Group findings by area in template slot order
    area_findings = {a: [] for a in RISK_AREA_ORDER}
    for name, data in findings.items():
        area_findings[data['risk_area']].append((name, data))

    # Build a flat list of (slide_num, slot_within_slide) for each area
    # so we know exactly which slide/slot each finding goes in
    area_slot_map = {}  # area -> list of (slide_num, slot_within_slide)
    for area in RISK_AREA_ORDER:
        slots = []
        for slide_num, (slide_area, num_slots) in SLIDE_MAP.items():
            if slide_area == area:
                for s in range(num_slots):
                    slots.append((slide_num, s))
        area_slot_map[area] = slots

    # ── Edit every slide ──────────────────────────────────────────────────────
    for slide_idx, sp in enumerate(slides, start=1):
        with open(sp, encoding='utf-8') as f:
            content = f.read()

        # Universal replacements
        for old in old_names: content = content.replace(old, new_hdr)
        for old in old_dates: content = content.replace(old, new_date)
        content = content.replace('<a:t>[Month]</a:t>', f'<a:t>{mo}</a:t>')
        content = content.replace('<a:t> 2025</a:t>', f'<a:t> {yr}</a:t>')
        content = content.replace('<a:t> 2026</a:t>', f'<a:t> {yr}</a:t>')

        # Auditor
        content = safe_rep(content,'<a:t>Yumee Song</a:t>',f'<a:t>{xml_esc(auditor_names)}</a:t>')
        content = safe_rep(content,'<a:t>Theresa Eichner</a:t>',f'<a:t>{xml_esc(auditor_names)}</a:t>')
        content = safe_rep(content,'<a:t>[Auditor Name]</a:t>',f'<a:t>{xml_esc(auditor_names)}</a:t>')

        # Period under review (all month/year variants that may appear in template)
        for old_p in ['February 2026','March 2026','December 2025','January 2026',
                      'April 2026','June 2026','July 2026','August 2026']:
            content = safe_rep(content, f'<a:t>{old_p}</a:t>', f'<a:t>{xml_esc(period_under_review)}</a:t>')

        # Audit Approach — cover slide has "Remote" or "On-Site" hardcoded
        # Default to On-Site; user can override via --approach flag
        content = safe_rep(content, '<a:t>Remote</a:t>', f'<a:t>{xml_esc(audit_approach)}</a:t>')

        # ── Cover slide (slide 1) ─────────────────────────────────────────────
        if slide_idx == 1:
            # Update overall score
            if scores:
                op = int(sum(s['score_decimal'] for s in scores.values())/len(scores)*100)
                content = safe_rep(content, '<a:t>XX%</a:t>', f'<a:t>{op}%</a:t>')

            # Fix Period under Review - may be split across two runs: "December " + "2025"
            period_parts = period_under_review.split()
            if len(period_parts) == 2:
                mo_part, yr_part = period_parts[0], period_parts[1]
                # Replace split runs
                for old_mo in ['January ','February ','March ','April ','May ','June ',
                                'July ','August ','September ','October ','November ','December ']:
                    if f'<a:t>{old_mo}</a:t>' in content:
                        content = safe_rep(content, f'<a:t>{old_mo}</a:t>',
                                           f'<a:t>{mo_part} </a:t>')
                for old_yr in ['2024','2025','2026','2027']:
                    tag = f'<a:t>{old_yr}</a:t>'
                    if tag in content:
                        content = safe_rep(content, tag, f'<a:t>{yr_part}</a:t>')
                        break

            # Update each area score AND rating by finding each area's row
            # in the results table and updating its score and rating cells directly
            idx_results = content.find('Audit Results by Risk Area')
            if idx_results == -1:
                idx_results = content.find('<a:t>Area</a:t>')
            
            area_order_in_table = ["Financial", "Operational", "Compliance", "Safety / HR"]
            for area in area_order_in_table:
                if area not in scores: continue
                data = scores[area]
                new_pct    = data["score_pct"]
                new_rating = data["rating"]
                new_color  = RATING_COLORS.get(new_rating, "70AD47")
                
                # Find this area's label in the results table
                area_idx = content.find(f'<a:t>{area}</a:t>', idx_results)
                if area_idx == -1: continue
                
                # Score cell = next <a:tc> after area cell
                tc1_end    = content.find('</a:tc>', area_idx) + 7
                sc_start   = content.find('<a:tc', tc1_end)
                sc_end     = content.find('</a:tc>', sc_start) + 7
                
                # Rating cell = next <a:tc> after score cell
                rt_start   = content.find('<a:tc', sc_end)
                rt_end     = content.find('</a:tc>', rt_start) + 7
                
                # Update score cell text
                score_cell = content[sc_start:sc_end]
                # Replace any existing <a:t> text with new score
                new_score_cell = re.sub(r'<a:t>[^<]*</a:t>',
                    f'<a:t>{new_pct}</a:t>', score_cell, count=1)
                content = content[:sc_start] + new_score_cell + content[sc_end:]
                
                # Re-find rating cell after score update
                rt_start = content.find('<a:tc', sc_start + len(new_score_cell))
                rt_end   = content.find('</a:tc>', rt_start) + 7
                
                # Update rating cell - text and color
                rating_cell = content[rt_start:rt_end]
                
                # Replace text
                new_rating_cell = re.sub(r'<a:t>[^<]*</a:t>',
                    f'<a:t>{new_rating}</a:t>', rating_cell, count=1)
                
                # Update background color - replace LAST solidFill in tcPr
                txb_end = new_rating_cell.find('</a:txBody>') + 11
                tcpr_part = new_rating_cell[txb_end:]
                fill_xml = f'<a:solidFill><a:srgbClr val="{new_color}"/></a:solidFill>'
                
                # Replace the last solidFill occurrence in tcPr (background, not borders)
                fills = list(re.finditer(
                    r'<a:solidFill>(?:(?!</a:solidFill>).)*</a:solidFill>',
                    tcpr_part, re.DOTALL))
                if fills:
                    last = fills[-1]
                    tcpr_part = tcpr_part[:last.start()] + fill_xml + tcpr_part[last.end():]
                else:
                    tcpr_part = tcpr_part.replace('</a:tcPr>',
                        f'  {fill_xml}\n                  </a:tcPr>', 1)
                
                new_rating_cell = new_rating_cell[:txb_end] + tcpr_part
                content = content[:rt_start] + new_rating_cell + content[rt_end:]

            # Update overall rating label
            for old_r in ["Unsatisfactory","Needs Improvement","Satisfactory","Exceeds Expectations"]:
                if overall_rating and f'<a:t>{old_r}</a:t>' in content:
                    content = safe_rep(content, f'<a:t>{old_r}</a:t>',
                                       f'<a:t>{overall_rating}</a:t>')
                    break

            # Update overall rating description text
            if overall_rating and overall_rating in RATING_DESCRIPTIONS:
                new_desc = RATING_DESCRIPTIONS[overall_rating]
                for old_desc in RATING_DESCRIPTIONS.values():
                    if old_desc != new_desc and old_desc[:40] in content:
                        # Replace first 40 chars as anchor then replace full cell
                        idx_desc = content.find(old_desc[:40])
                        if idx_desc != -1:
                            # Find the containing a:t tag and replace its text
                            t_start = content.rfind('<a:t>', 0, idx_desc)
                            t_end   = content.find('</a:t>', idx_desc) + 6
                            if t_start != -1:
                                content = (content[:t_start] + f'<a:t>{xml_esc(new_desc)}</a:t>'
                                          + content[t_end:])
                        break

            # Update summary observation bullets
            content = update_cover_summary(content, findings, scores, overall_rating)

        # ── Observation slides ────────────────────────────────────────────────
        elif slide_idx in SLIDE_MAP:
            slide_area, num_slots = SLIDE_MAP[slide_idx]
            slot_list  = area_slot_map[slide_area]
            area_finds = area_findings[slide_area]

            # Update area score and rating on this slide
            area_score_data = scores.get(slide_area, {})
            if area_score_data:
                for old_pct in area_old_scores.get(slide_area, []):
                    if f'<a:t>{old_pct}</a:t>' in content:
                        content = safe_rep(content, f'<a:t>{old_pct}</a:t>',
                                           f'<a:t>{area_score_data["score_pct"]}</a:t>')
                        break
                # Update rating text and color
                for old_rating in area_old_ratings.get(slide_area, []):
                    if f'<a:t>{old_rating}</a:t>' in content:
                        new_rating = area_score_data['rating']
                        new_color  = RATING_COLORS.get(new_rating, "70AD47")
                        content = update_rating_cell(content, old_rating, new_rating, new_color)
                        break

            # Update observation count
            total_area_exceptions = len(area_finds)
            for old_count in ["0","1","2","3","4","5","6","7"]:
                tag = f'<a:t>{old_count}</a:t>'
                if tag in content[:2000]:  # only in header area
                    content = safe_rep(content, tag,
                                       f'<a:t>{total_area_exceptions}</a:t>')
                    break

            # Figure out which findings go in this slide's slots
            # Each slide's slots are a contiguous subset of the area's slot list
            slide_slot_start = sum(
                SLIDE_MAP[sn][1] for sn in sorted(SLIDE_MAP.keys())
                if SLIDE_MAP[sn][0] == slide_area and sn < slide_idx
            )

            for local_slot in range(num_slots):
                global_slot = slide_slot_start + local_slot
                if global_slot < len(area_finds):
                    test_name, fdata = area_finds[global_slot]
                    obs_text = fdata['observation']
                    ap_text  = fdata['action_plan']

                    # Inject observation
                    content = replace_cell_content(
                        content, 'Observation(s):', local_slot, make_bullet_txbody(obs_text))
                    # Inject action plan
                    content = replace_cell_content(
                        content, 'Action Plan:', local_slot, make_bullet_txbody(ap_text))
                else:
                    # Clear unused slot - obs and action plan
                    content = replace_cell_content(
                        content, 'Observation(s):', local_slot, make_bullet_txbody(""))
                    content = replace_cell_content(
                        content, 'Action Plan:', local_slot, make_bullet_txbody(""))


        # ── Audit program slides (10-13) ──────────────────────────────────────
        elif slide_idx in [10,11,12,13]:
            # Update overall score if present
            if scores:
                op = int(sum(s['score_decimal'] for s in scores.values())/len(scores)*100)
                content = safe_rep(content, '<a:t>XX%</a:t>', f'<a:t>{op}%</a:t>')

        with open(sp, 'w', encoding='utf-8') as f:
            f.write(content)

    # Repack
    output_path = Path(output_path)
    with zipfile.ZipFile(output_path,'w',zipfile.ZIP_DEFLATED) as zout:
        for fp in work_dir.rglob('*'):
            if fp.is_file():
                zout.write(fp, fp.relative_to(work_dir))
    shutil.rmtree(work_dir)
    print(f"  Saved: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# PATH HELPERS & MAIN
# ─────────────────────────────────────────────────────────────────────────────

def resolve_path(p, default_dir, extensions=None):
    p = Path(p)
    if p.is_absolute() or "/" in str(p) or "\\" in str(p):
        return p
    for directory in [default_dir, Path.cwd()]:
        candidate = directory/p
        if candidate.exists(): return candidate
        if extensions:
            for ext in extensions:
                c2 = directory/(str(p)+ext)
                if c2.exists(): return c2
    return default_dir/p


def main():
    parser = argparse.ArgumentParser(
        description="Generate BrightView Branch Audit Report (Pure Python)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
FILES  (default folder: ~/Downloads)
  Drop audit Excel here:   ~/Downloads/YourBranch.xlsx
  Save template here:      ~/Downloads/template.pptx
  Output saved here:       ~/Downloads/<BranchName>_Audit_Report.pptx

EXAMPLES
  python generate_audit_report_pure.py Homestead.xlsx
  python generate_audit_report_pure.py Homestead.xlsx --period "March 2026" --month "May 2026"
  python generate_audit_report_pure.py Homestead.xlsx --auditors "Jane Smith"
        """
    )
    parser.add_argument("excel")
    parser.add_argument("--template", default="template.pptx")
    parser.add_argument("--period",   help="Period under review, e.g. 'March 2026'")
    parser.add_argument("--month",    help="Report month, e.g. 'May 2026'")
    parser.add_argument("--auditors", default="[Auditor Name]")
    parser.add_argument("--branch",   help="Branch display name override, e.g. '35210 BVLS Homestead'")
    parser.add_argument("--approach", default="On-Site", help="Audit approach: On-Site or Remote (default: On-Site)")
    parser.add_argument("--output")
    args = parser.parse_args()

    excel_path    = resolve_path(args.excel,    DOWNLOADS, [".xlsx",".xls"])
    template_path = resolve_path(args.template, DOWNLOADS, [".pptx"])

    if not excel_path.exists():
        print(f"ERROR: Excel file not found.\n  Looked in: {DOWNLOADS}\n  File: {args.excel}")
        sys.exit(1)
    if not template_path.exists():
        print(f"ERROR: Template PPTX not found.\n  Looked in: {DOWNLOADS}\n  File: {args.template}")
        sys.exit(1)

    period      = args.period or input("Period under review (e.g. 'March 2026'): ").strip()
    audit_month = args.month  or input("Month of audit report (e.g. 'May 2026'): ").strip()

    print(f"\n{'='*60}")
    print("STEP 1: Reading Excel...")
    print('='*60)
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    branch_num, branch_display = get_branch_info(wb, excel_path)
    if args.branch:
        branch_display = args.branch
        m = re.match(r'(\d{4,6})', args.branch)
        if m: branch_num = m.group(1)
    scores, overall = read_finish_tab(wb)
    print(f"  Branch : {branch_display}")
    print(f"  Overall: {overall or 'Not found'}")
    for area,data in scores.items():
        print(f"  {area:15s}: {data['score_pct']:5s} — {data['rating']}")

    print("\nSTEP 2: Scanning for exceptions...")
    exceptions = scan_for_exceptions(wb)
    print(f"  Found {len(exceptions)} exceptions:")
    for name,data in exceptions.items():
        print(f"    [{data['risk_area']}] {name}")

    print("\nSTEP 3: Drafting observations...")
    for test_name, exc_data in exceptions.items():
        tmpl   = TEMPLATES.get(test_name,{})
        obs_fn = tmpl.get("obs_fn")
        if obs_fn:
            exc_data["observation"] = obs_fn(exc_data["data"], period)
        print(f"  ✓ {test_name}")
        print(f"    {exc_data['observation'][:100]}...")

    safe_name = re.sub(r'[^\w\-]','_', branch_display)
    if args.output:
        out_p = Path(args.output)
        output = str(out_p) if (out_p.is_absolute() or "/" in str(out_p) or "\\" in str(out_p)) \
                 else str(DOWNLOADS/out_p)
    else:
        output = str(DOWNLOADS/f"{safe_name}_Audit_Report.pptx")

    print(f"\nSTEP 4: Building PPTX...")
    build_pptx(template_path=template_path, output_path=output,
               branch_display=branch_display, scores=scores,
               overall_rating=overall, findings=exceptions,
               period_under_review=period, audit_month=audit_month,
               auditor_names=args.auditors,
               audit_approach=args.approach)

    print(f"\n{'='*60}")
    print(f"✓  Report generated: {output}")
    print(f"   Branch    : {branch_display}")
    print(f"   Period    : {period}")
    print(f"   Overall   : {overall or 'See FINISH tab'}")
    print(f"   Exceptions: {len(exceptions)}")
    print('='*60)
    print("\nNOTE: Branch overview slide still needs team names and")
    print("      financial data filled in manually.")


if __name__ == "__main__":
    main()
