"""The joint: the agent's conclusions -> a CY workpaper that mirrors PY.

Both halves of this repo already existed and never met. agent/loop.py and
run_control.py decide what is true; agent/ooxml/ reproduces the prior-year
workpaper. This is the single call that runs one into the other, and the
only place that knows both exist.

TWO RENDERERS, ON PURPOSE
-------------------------
The PY-mirroring renderer needs the prior-year workpaper as a real .xlsx
with the tabs it writes into. That is the normal case and the good one:
the output inherits last year's layout instead of inventing one.

But a PY workpaper is not always an xlsx -- plenty are scanned or exported
to PDF -- and a template can be shaped differently enough that writing
into it would produce something wrong. In those cases this falls back to
agent/workpaper.py, the self-contained builder that owes nothing to a
template. The fallback output does NOT look like last year's; that is the
whole point of preferring the other one, and the caller is told which it
got rather than left to guess from the file.

The rule the fallback exists to serve: never refuse to produce a workpaper
because the layout could not be mirrored. A run that reached real
conclusions has already spent real money and real evidence review, and a
worse-looking file that carries them is worth far more than an error.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from agent.ooxml import ooxml_utils as ox
from agent.ooxml.build_workpaper import build_workpaper as build_ooxml_workpaper
from agent.ooxml.models import PopulationRow, WorkpaperRequest
from agent.render_bridge import build_workpaper_request, summary_rows
from agent.schemas import ConclusionOutput, SamplePopulationManifest

# Column letters, matching the bridge's own extract-order convention.
_LETTERS = [chr(ord("A") + i) for i in range(26)]


@dataclass
class WorkpaperOutcome:
    """What was produced, and everything the caller must be told about it.

    `warnings` is not decoration. It carries the two things that are
    invisible in the finished file and matter most: evidence the bridge
    could not attach, and tickmarks whose anchor could not be located on
    the page. A legend entry pointing at a box that was never drawn sends
    a reviewer hunting for a mark that does not exist, so a caller that
    drops these is shipping a quietly wrong workpaper.
    """

    path: Path
    mirrored_py: bool
    warnings: list[str] = field(default_factory=list)
    unplaced_callouts: list[tuple[str, str, str, str]] = field(default_factory=list)

    @property
    def summary_line(self) -> str:
        how = (
            "mirroring the prior-year workpaper's layout"
            if self.mirrored_py
            else "using the standalone layout (the PY file could not be used as a template)"
        )
        note = ""
        if self.unplaced_callouts:
            note = f"; {len(self.unplaced_callouts)} tickmark(s) could not be placed"
        return f"Wrote {self.path.name} {how}{note}"


def _population_rows(
    population_path: Path | None, sheet_name: str | None
) -> list[PopulationRow]:
    """The population tab, as rows keyed by template column letter.

    Returns [] when there is nothing to read, and the OOXML builder treats
    an empty list as "leave the template's own population alone" rather
    than clearing it -- deleting the population removes the basis for the
    sample, and a real output shipped that way once.
    """
    if not population_path or not sheet_name or not population_path.exists():
        return []
    import openpyxl

    wb = openpyxl.load_workbook(population_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows: list[PopulationRow] = []
        for raw in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in raw):
                continue
            rows.append(
                PopulationRow(
                    values={
                        _LETTERS[i]: v
                        for i, v in enumerate(raw)
                        if i < len(_LETTERS) and v is not None
                    }
                )
            )
        return rows
    finally:
        wb.close()


def _can_mirror(py_path: Path, sample_sheet: str, population_sheet: str) -> tuple[bool, str]:
    """Whether the PY file can be written into, and why not when it can't.

    Checked by actually unpacking and inspecting it rather than by
    extension. A file can be named .xlsx and still be the wrong shape, and
    finding that out here -- before any conclusion has been rendered -- is
    much better than a half-written workbook.
    """
    if py_path.suffix.lower() not in (".xlsx", ".xlsm"):
        return False, f"the PY workpaper is a {py_path.suffix or 'file with no extension'}, not a workbook"
    import tempfile

    with tempfile.TemporaryDirectory() as probe:
        try:
            ox.extract_template(str(py_path), probe)
            ox.check_template(
                probe,
                required_sheets=[sample_sheet, population_sheet],
                max_style_index=14,  # build_workpaper's highest STYLE_* index
            )
        except ox.TemplateMismatch as exc:
            return False, str(exc)
        except Exception as exc:  # noqa: BLE001 -- a corrupt or exotic file is a fallback, not a crash
            return False, f"the PY workpaper could not be opened as a template: {exc}"
    return True, ""


def build_cy_workpaper(
    spec: dict[str, Any],
    results: dict[str, dict],
    py_testing_filename: str,
    out_dir: Path,
    *,
    support_dir: Path | None = None,
    sample_manifests: dict[str, SamplePopulationManifest] | None = None,
    population_workbook: str | None = None,
    population_sheet: str | None = None,
    sample_sheet_name: str = "Sample",
    population_sheet_name: str = "Population",
) -> WorkpaperOutcome:
    """Render one control's results, mirroring PY where that is possible."""
    support = Path(support_dir or out_dir)
    py_path = support / py_testing_filename
    out_dir.mkdir(parents=True, exist_ok=True)

    conclusions: list[tuple[str, ConclusionOutput]] = [
        (step_id, r["conclusion"]) for step_id, r in results.items() if "conclusion" in r
    ]
    if not conclusions:
        # Nothing concluded: fall straight through to the legacy builder,
        # which knows how to render a control whose steps all failed.
        return _legacy(spec, results, py_testing_filename, out_dir, support, population_workbook,
                       ["no test step reached a conclusion, so there was nothing to mirror"])

    can, why = _can_mirror(py_path, sample_sheet_name, population_sheet_name)
    if not can:
        return _legacy(spec, results, py_testing_filename, out_dir, support, population_workbook,
                       [f"could not mirror the prior-year layout: {why}"])

    # One Sample tab holds every sampled item across every test step, which
    # is how the prior-year workpapers are built -- the sample listing is
    # per control, not per step.
    warnings: list[str] = []
    sample_items = []
    step_labels = []
    for step_id, conclusion in conclusions:
        manifest = (sample_manifests or {}).get(step_id)
        request, step_warnings = build_workpaper_request(
            conclusion,
            manifest,
            control_id=spec["control_id"],
            template_path=py_path,
            output_path=out_dir / f"{spec['control_id']}_CY_Testing_wp.xlsx",
            support_dir=support,
        )
        sample_items.extend(request.sample_items)
        warnings.extend(f"[{step_id}] {w}" for w in step_warnings)
        text = next(
            (s.get("test_step_text", "") for s in spec.get("test_steps", []) if s.get("test_step_id") == step_id),
            "",
        )
        step_labels.append(f"{step_id}: {text}" if text else step_id)

    output_path = out_dir / f"{spec['control_id']}_CY_Testing_wp.xlsx"
    request = WorkpaperRequest(
        template_path=str(py_path),
        output_path=str(output_path),
        sample_items=sample_items,
        population_rows=_population_rows(
            support / population_workbook if population_workbook else None, population_sheet
        ),
    )

    unplaced: list[tuple[str, str, str, str]] = []
    build_ooxml_workpaper(
        request,
        sample_sheet=sample_sheet_name,
        population_sheet=population_sheet_name,
        on_unplaced=unplaced.extend,
        summary_rows=summary_rows(list(zip(step_labels, [c for _s, c in conclusions]))),
    )
    return WorkpaperOutcome(output_path, mirrored_py=True, warnings=warnings, unplaced_callouts=unplaced)


def _legacy(
    spec: dict[str, Any],
    results: dict[str, dict],
    py_testing_filename: str,
    out_dir: Path,
    support: Path,
    population_workbook: str | None,
    warnings: list[str],
) -> WorkpaperOutcome:
    """The standalone builder, for when PY cannot be mirrored."""
    from agent.workpaper import build_workpaper as build_legacy_workpaper

    path = build_legacy_workpaper(
        spec,
        results,
        py_testing_filename,
        out_dir,
        support_dir=support,
        source_workbook=population_workbook,
    )
    return WorkpaperOutcome(Path(path), mirrored_py=False, warnings=warnings)
