"""Workpaper generation tests: build from real ConclusionOutput objects and
read the produced files back (openpyxl / pdfplumber) rather than trusting
that "no exception" means a usable document.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pdfplumber
import pytest

from agent.loop import AuditLogEntry
from agent.schemas import ConclusionOutput, EvidenceCitation, ModelMetadata, SampleCoverage
from agent.workpaper import build_workpaper, workpaper_path_for


@pytest.fixture
def spec() -> dict:
    return {
        "control_id": "C-14",
        "control_objective_ref": "CO-4",
        "control_objective_text": "Accruals are recorded completely and accurately.",
        "test_steps": [
            {"test_step_id": "TS-4.2", "test_step_text": "Recalculate the accrual and agree to the GL."},
            {"test_step_id": "TS-9.9", "test_step_text": "A step that failed mid-run."},
        ],
    }


@pytest.fixture
def results() -> dict:
    conclusion = ConclusionOutput(
        test_step_id="TS-4.2",
        control_objective_ref="CO-4",
        conclusion="satisfied",
        narrative="Recalculated accrual ties to GL with no variance.",
        evidence_citations=[
            EvidenceCitation(
                evidence_id="ev_0001",
                source_file="Recon_Oct2026.xlsx",
                location="Sheet1!A1:B2",
                quote_or_summary="Recalculation ties to GL export with no variance",
                relevance="CY support for the accrual recalculation",
            )
        ],
        procedures_performed=["Recalculated the accrual", "Agreed the balance to the GL export"],
        ipe_completeness_accuracy_status="not_applicable",
        ipe_completeness_accuracy_evidence=[],
        exceptions=[],
        additional_support_requests=["Q4 support for the November accrual"],
        confidence="high",
        confidence_rationale="Single clear source, full sample coverage.",
        sample_coverage=SampleCoverage(
            total_required=1, total_found=1, missing=[], coverage_pct=100.0, complete=True
        ),
        model_metadata=ModelMetadata(
            model="claude-opus-5", prompt_version="v1", timestamp="2026-08-27T12:00:00+00:00", tool_call_count=3
        ),
    )
    failed = {
        "error": "test step 'TS-9.9' exceeded the 50,000-token budget after 4 turn(s)",
        "reason": "token_budget_exceeded",
        "tokens_used": 51_200,
        "turns_used": 4,
        "audit_log": [
            AuditLogEntry(
                turn=1,
                tool_name="search_cy_support",
                tool_use_id="t1",
                input={"query": "november accrual reconciliation", "top_k": 5},
                output={"results": []},
                is_error=False,
                timestamp="2026-08-27T12:01:00+00:00",
            )
        ],
    }
    return {"TS-4.2": {"conclusion": conclusion, "audit_log": []}, "TS-9.9": failed}


def test_excel_is_the_default_regardless_of_py_file_type(tmp_path: Path):
    # Excel is the deliverable default: it's the only output carrying the
    # full structure (per-step sheets, exhibits sheet, filterable summary).
    # Matching the PY file type meant a PDF precedent silently downgraded
    # this year's workpaper to the flat rendering.
    assert workpaper_path_for("PY_C14.xlsm", "C-14", tmp_path).suffix == ".xlsx"
    assert workpaper_path_for("PY_C14.pdf", "C-14", tmp_path).suffix == ".xlsx"
    # Control ids with filesystem-hostile characters can't break the filename.
    assert "/" not in workpaper_path_for("PY.pdf", "C/14 (AP)", tmp_path).name


def test_xlsx_workpaper_contents(spec: dict, results: dict, tmp_path: Path):
    path = build_workpaper(spec, results, "PY_Testing_C14.xlsx", tmp_path)

    wb = openpyxl.load_workbook(path)
    assert set(wb.sheetnames) == {"TS-4.2 Summary", "TS-9.9 Summary"}

    def sheet_text(name: str) -> str:
        return " ".join(
            str(cell.value) for row in wb[name].iter_rows() for cell in row if cell.value is not None
        )

    summary = sheet_text("TS-4.2 Summary")
    assert "DRAFT" in summary
    assert "C-14" in summary
    assert "Satisfied" in summary
    assert "INCOMPLETE" in summary  # the failed step is visible on the summary, not silently missing

    step = sheet_text("TS-4.2 Summary")
    assert "Recalculated accrual ties to GL with no variance." in step
    # Evidence is named in plain English, not by internal id -- "ev_0001"
    # tells a reviewer nothing about what they are looking at.
    assert "ev_0001" not in step
    assert "Sheet1" in step  # the cited sheet name
    assert "Recon_Oct2026.xlsx" in step
    assert "Q4 support for the November accrual" in step
    assert "claude-opus-5" in step

    failed = sheet_text("TS-9.9 Summary")
    assert "INCOMPLETE" in failed
    assert "token_budget_exceeded" in failed
    assert "november accrual reconciliation" in failed  # searches attempted before the abort are documented


def _make_support_pdf(path: Path) -> None:
    from reportlab.lib.pagesizes import letter as letter_size
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(str(path), pagesize=letter_size)
    c.drawString(72, 700, "Approved / Vendor Number 21022452 / Cost Center 0510")
    c.drawString(72, 680, "Payment Amount 11,678.47 dated 12/04/2025")
    c.showPage()
    c.save()


@pytest.fixture
def annotated_setup(spec: dict, results: dict, tmp_path: Path) -> tuple[dict, dict, Path]:
    # A real support file on disk whose deterministic re-extraction yields
    # ev_0001 for the page text -- the id the citation below points at.
    support = tmp_path / "support"
    support.mkdir()
    _make_support_pdf(support / "approval.pdf")
    spec = {**spec, "cy_support_files": ["approval.pdf"]}

    conclusion: ConclusionOutput = results["TS-4.2"]["conclusion"]
    annotated = conclusion.model_copy(
        update={
            "evidence_citations": [
                EvidenceCitation(
                    evidence_id="ev_0001",
                    source_file="approval.pdf",
                    location="approval.pdf p.1",
                    quote_or_summary="Approved / Vendor Number 21022452 / Cost Center 0510",
                    relevance="Approval evidence for the sampled payment",
                )
            ]
        }
    )
    results = {"TS-4.2": {"conclusion": annotated, "audit_log": []}}
    return spec, results, support


def test_xlsx_workpaper_embeds_annotated_exhibit(annotated_setup, tmp_path: Path):
    spec, results, support = annotated_setup
    path = build_workpaper(spec, results, "PY_Testing_C14.xlsx", tmp_path, support_dir=support)

    wb = openpyxl.load_workbook(path)
    ws = wb["Summary"]
    text = " ".join(str(c.value) for r in ws.iter_rows() for c in r if c.value is not None)
    assert "Tickmark" in text
    assert "approval.pdf p.1" in text

    # Exhibits live on their own sheet -- inline, full-page renders pushed
    # the conclusion ~110 rows down the step sheet.
    assert len(ws._images) == 0
    assert "see the 'Summary - Exhibits' sheet" in text

    ex_ws = wb["Summary - Exhibits"]
    # The page plus its marks as SEPARATE pictures: burned-in marks are
    # final, and a reviewer has to be able to drag a misplaced box onto the
    # right value rather than redo the exhibit by hand.
    assert len(ex_ws._images) >= 2
    offsets = [
        (im.anchor._from.colOff, im.anchor._from.rowOff)
        for im in ex_ws._images
        if hasattr(im.anchor, "_from")
    ]
    # At least one mark is pinned at a real pixel offset over the page.
    assert any(x or y for x, y in offsets), "overlays were not positioned"
    ex_text = " ".join(str(c.value) for r in ex_ws.iter_rows() for c in r if c.value is not None)
    assert "Evidence exhibits" in ex_text
    assert "could not be located" in ex_text  # explains a corner-letter exhibit
    # No legend here: it repeated the letter/quote/source already in the
    # Evidence Cited table, so a reviewer read the same rows twice and had
    # two places to keep in sync.
    assert "TICKMARK LEGEND" not in ex_text
    assert "Vendor Number 21022452" not in ex_text


def test_each_sampled_item_gets_its_own_sheet_and_restarts_tickmarks(
    spec: dict, results: dict, tmp_path: Path
):
    # A reviewer clearing selection 2 wants selection 2's evidence, not one
    # merged list where its first tickmark happens to be D.
    def cite(ev, sid, quote):
        return EvidenceCitation(
            evidence_id=ev,
            source_file="support.pdf",
            location="support.pdf p.1",
            quote_or_summary=quote,
            relevance="r",
            sample_id=sid,
        )

    conclusion = results["TS-4.2"]["conclusion"].model_copy(
        update={
            "evidence_citations": [
                cite("ev_0018", None, "Population extract, 30 rows"),  # step-wide
                cite("ev_0001", "1", "invoice 2859 approved"),
                cite("ev_0010", "1", "payment 881477"),
                cite("ev_0015", "2", "invoice 35713082 approved"),
            ]
        }
    )
    path = build_workpaper(
        spec, {"TS-4.2": {"conclusion": conclusion, "audit_log": []}}, "PY.xlsx", tmp_path
    )
    wb = openpyxl.load_workbook(path)

    assert wb.sheetnames == ["Summary", "1", "2"]

    def text(name):
        return " ".join(str(c.value) for r in wb[name].iter_rows() for c in r if c.value is not None)

    # Summary carries the narrative and procedures, not the evidence table.
    summary = text("Summary")
    assert "Recalculated accrual ties to GL with no variance." in summary
    assert "PROCEDURES PERFORMED" in summary
    assert "ev_0001" not in summary

    s1, s2 = text("1"), text("2")
    # Each item's own evidence, identified by its quote rather than an
    # internal id -- and sample 2's must not bleed onto sample 1's sheet.
    assert "invoice 2859 approved" in s1
    assert "payment 881477" in s1
    assert "invoice 35713082 approved" in s2
    assert "invoice 35713082 approved" not in s1
    assert "invoice 2859 approved" not in s2
    # Internal ids never surface in the workpaper.
    assert "ev_00" not in s1 and "ev_00" not in s2
    # Step-wide evidence (the population extract) belongs on the summary,
    # NOT on the first item. Riding along there gave sample 1 tickmarks its
    # own attributes never referenced -- a real workpaper had it carrying B
    # and C for the population and report parameters.
    assert "Population extract, 30 rows" not in s1
    assert "Population extract, 30 rows" in summary

    # Tickmarks restart per sheet: sample 2's single citation is A, not D.
    def tickmarks(name):
        ws = wb[name]
        col = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        return [v for v in col if v in ("A", "B", "C", "D")]

    # Two citations on sample 1 -- the third was step-wide and moved to the
    # summary, so no tickmark on this sheet goes unreferenced.
    assert tickmarks("1") == ["A", "B"]
    assert tickmarks("2") == ["A"]


def test_summary_matrix_is_one_row_per_sample_with_links_and_colors(
    spec: dict, results: dict, tmp_path: Path
):
    # A row per test step could only say "the step failed" -- never WHICH
    # selection failed, which is the first thing a reviewer asks.
    from agent.schemas import SampleResult

    conclusion = results["TS-4.2"]["conclusion"].model_copy(
        update={
            "conclusion": "not_satisfied",
            "sample_results": [
                SampleResult(sample_id="1", conclusion="satisfied"),
                SampleResult(sample_id="2", conclusion="not_satisfied", note="No approval evidence."),
            ],
            "evidence_citations": [
                EvidenceCitation(
                    evidence_id="ev_0001",
                    source_file="s.pdf",
                    location="s.pdf p.1",
                    quote_or_summary="q",
                    relevance="r",
                    sample_id=sid,
                )
                for sid in ("1", "2")
            ],
        }
    )
    path = build_workpaper(
        spec, {"TS-4.2": {"conclusion": conclusion, "audit_log": []}}, "PY.xlsx", tmp_path
    )
    ws = openpyxl.load_workbook(path)["Summary"]

    header = next(
        r for r in ws.iter_rows() if r[0].value == "Sample"
    )
    assert [c.value for c in header[:2]] == ["Sample", "Test step TS-4.2"]
    # No "Detail sheet" column: with several steps a sample has a sheet per
    # step, so one link per row could only reach the first of them.
    assert header[2].value is None

    rows = {r[0].value: r for r in ws.iter_rows() if r[0].value in ("1", "2")}
    # Each item's own verdict, not the step roll-up.
    assert rows["1"][1].value == "Satisfied"
    assert rows["2"][1].value == "Not satisfied"
    # Colour-coded so it reads at a glance.
    assert rows["1"][1].fill.start_color.rgb.endswith("C6EFCE")  # green
    assert rows["2"][1].fill.start_color.rgb.endswith("FFC7CE")  # red
    # The verdict cell IS the link -- it identifies exactly one
    # (sample, step) pair, which is exactly one sheet.
    assert "'1'!A1" in str(rows["1"][1].hyperlink.location or rows["1"][1].hyperlink.target)
    assert "'2'!A1" in str(rows["2"][1].hyperlink.location or rows["2"][1].hyperlink.target)


def test_attributes_table_shows_where_the_step_is_satisfied(spec: dict, results: dict, tmp_path: Path):
    # A reviewer signs off attribute by attribute. Before this they had to
    # read the whole narrative to find where any single attribute was met.
    from agent.schemas import AttributeResult

    cites = [
        EvidenceCitation(
            evidence_id=ev,
            source_file="s.pdf",
            location="s.pdf p.1",
            quote_or_summary="q",
            relevance="r",
            sample_id="1",
        )
        for ev in ("ev_0001", "ev_0010")
    ]
    conclusion = results["TS-4.2"]["conclusion"].model_copy(
        update={
            "evidence_citations": cites,
            "attribute_results": [
                AttributeResult(
                    attribute="Approval precedes payment",
                    sample_id="1",
                    result="satisfied",
                    value_observed="Approved 10/1/2025; paid 11/10/2025",
                    evidence_ids=["ev_0001", "ev_0010"],
                ),
                AttributeResult(
                    attribute="Approver has documented authority",
                    sample_id="1",
                    result="not_tested",
                    value_observed="No authority matrix provided",
                    evidence_ids=[],
                ),
            ],
        }
    )
    path = build_workpaper(
        spec, {"TS-4.2": {"conclusion": conclusion, "audit_log": []}}, "PY.xlsx", tmp_path
    )
    ws = openpyxl.load_workbook(path)["1"]
    rows = list(ws.iter_rows())

    header = next(r for r in rows if r[0].value == "Tickmark" and r[1].value == "Attribute")
    assert [c.value for c in header[:5]] == ["Tickmark", "Attribute", "Result", "Value observed", "Evidence"]

    by_attr = {r[1].value: r for r in rows if r[1].value in
               ("Approval precedes payment", "Approver has documented authority")}

    ok = by_attr["Approval precedes payment"]
    assert ok[2].value == "Satisfied"
    assert ok[2].fill.start_color.rgb.endswith("C6EFCE")  # green
    assert ok[3].value == "Approved 10/1/2025; paid 11/10/2025"
    # The chain that makes this useful: attribute -> tickmark -> the red
    # box on the exhibit page.
    assert ok[0].value == "A, B"

    untested = by_attr["Approver has documented authority"]
    assert untested[2].value == "Not tested"
    assert untested[2].fill.start_color.rgb.endswith("FFEB9C")  # amber
    assert not untested[0].value  # nothing cited, so no tickmark to point at


def test_summary_has_an_ipe_row_and_no_redundant_step_table(spec: dict, results: dict, tmp_path: Path):
    # IPE is a conclusion about the POPULATION, not about any sampled item,
    # so it sits at the same level as the selections rather than buried as
    # one attribute among many.
    from agent.schemas import SampleResult

    conclusion = results["TS-4.2"]["conclusion"].model_copy(
        update={
            "ipe_completeness_accuracy_status": "not_validated",
            "sample_results": [SampleResult(sample_id="1", conclusion="satisfied")],
            "evidence_citations": [
                EvidenceCitation(
                    evidence_id="ev_0001",
                    source_file="s.pdf",
                    location="p.1",
                    quote_or_summary="q",
                    relevance="r",
                    sample_id="1",
                )
            ],
        }
    )
    path = build_workpaper(
        spec, {"TS-4.2": {"conclusion": conclusion, "audit_log": []}}, "PY.xlsx", tmp_path
    )
    ws = openpyxl.load_workbook(path)["Summary"]
    rows = list(ws.iter_rows())

    ipe_row = next(r for r in rows if r[0].value == "IPE")
    assert ipe_row[1].value == "not_validated"
    assert ipe_row[1].fill.start_color.rgb.endswith("FFC7CE")  # red

    # On a single-step control the per-step table restated the matrix above
    # and the CONCLUSION section below -- every column of it was duplicated.
    assert not any(r[0].value == "Test step" and r[1].value == "Conclusion" for r in rows)


def test_no_sheet_carries_a_tickmark_its_attributes_never_use(spec: dict, results: dict, tmp_path: Path):
    # The defect class: step-wide evidence placed on a sampled item's sheet
    # gave it lettered rows nothing on that sheet referenced.
    from agent.schemas import AttributeResult

    conclusion = results["TS-4.2"]["conclusion"].model_copy(
        update={
            "evidence_citations": [
                EvidenceCitation(
                    evidence_id="ev_pop",
                    source_file="s.xlsx",
                    location="Population!A1",
                    quote_or_summary="population",
                    relevance="r",
                    sample_id=None,
                ),
                EvidenceCitation(
                    evidence_id="ev_inv",
                    source_file="s.pdf",
                    location="p.1",
                    quote_or_summary="invoice",
                    relevance="r",
                    sample_id="1",
                ),
            ],
            "attribute_results": [
                AttributeResult(
                    attribute="Invoice above threshold",
                    sample_id="1",
                    result="satisfied",
                    value_observed="$30,000",
                    evidence_ids=["ev_inv"],
                )
            ],
        }
    )
    path = build_workpaper(
        spec, {"TS-4.2": {"conclusion": conclusion, "audit_log": []}}, "PY.xlsx", tmp_path
    )
    ws = openpyxl.load_workbook(path)["1"]
    rows = list(ws.iter_rows())

    used, listed, mode = set(), set(), None
    for r in rows:
        first = r[0].value
        if first == "ATTRIBUTES TESTED":
            mode = "attr"
        elif first == "EVIDENCE CITED":
            mode = "ev"
        elif first and first != "Tickmark":
            if mode == "attr":
                used |= {t.strip() for t in str(first).split(",")}
            elif mode == "ev" and len(str(first)) == 1:
                listed.add(str(first))

    assert listed, "no evidence rows rendered"
    assert listed <= used, f"tickmarks listed but never referenced: {sorted(listed - used)}"


def test_multi_step_sample_sheets_name_their_step(spec: dict, results: dict, tmp_path: Path):
    # Every step has its own sample "1". Bare names collided into "1" and
    # "1(2)", which says nothing about which step the second belongs to --
    # and the single per-row link could only ever reach the first of them.
    from agent.schemas import SampleResult

    def step(step_id: str) -> dict:
        c = results["TS-4.2"]["conclusion"].model_copy(
            update={
                "test_step_id": step_id,
                "sample_results": [SampleResult(sample_id="1", conclusion="satisfied")],
                "evidence_citations": [
                    EvidenceCitation(
                        evidence_id="ev_1",
                        source_file="s.pdf",
                        location="p.1",
                        quote_or_summary="q",
                        relevance="r",
                        sample_id="1",
                    )
                ],
            }
        )
        return {"conclusion": c, "audit_log": []}

    path = build_workpaper(
        spec, {"TS-4.2": step("TS-4.2"), "TS-9.9": step("TS-9.9")}, "PY.xlsx", tmp_path
    )
    wb = openpyxl.load_workbook(path)

    assert "TS-4.2 1" in wb.sheetnames
    assert "TS-9.9 1" in wb.sheetnames
    assert not any(n.endswith("(2)") for n in wb.sheetnames), wb.sheetnames

    # Each verdict cell links to ITS OWN step's sheet for that sample.
    ws = wb["TS-4.2 Summary"]
    row = next(r for r in ws.iter_rows() if r[0].value == "1")
    assert "TS-4.2 1" in str(row[1].hyperlink.location or row[1].hyperlink.target)
    assert "TS-9.9 1" in str(row[2].hyperlink.location or row[2].hyperlink.target)


def test_untagged_citations_stay_on_one_sheet(spec: dict, results: dict, tmp_path: Path):
    # An older run (or a step with no sample) has no sample_id anywhere --
    # everything must stay on the summary sheet rather than vanishing.
    path = build_workpaper(
        spec, {"TS-4.2": results["TS-4.2"]}, "PY.xlsx", tmp_path
    )
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["Summary"]
    text = " ".join(str(c.value) for r in wb["Summary"].iter_rows() for c in r if c.value is not None)
    assert "Recon_Oct2026.xlsx" in text


def test_tickmark_boxes_are_padded_off_the_text():
    # Both the text-layer search and OCR return a box hugging the glyphs, so
    # drawing it as-is put the stroke ON the characters and made the boxed
    # value harder to read than the text around it.
    from agent.workpaper import _pad_rect

    rect = (100.0, 200.0, 160.0, 212.0)  # 12pt-tall line
    x0, top, x1, bottom = _pad_rect(rect, page_width=612.0, page_height=792.0)

    assert x0 < 100.0 and x1 > 160.0 and top < 200.0 and bottom > 212.0

    # Asymmetric: neighbouring lines sit close above/below, but there is
    # whitespace left and right. Equal padding wide enough to look right
    # horizontally struck through the following line on a real invoice.
    assert (100.0 - x0) > (200.0 - top)


def test_padded_box_never_leaves_the_page():
    from agent.workpaper import _pad_rect

    # A value flush against the page edge must not produce a negative or
    # overflowing rectangle.
    x0, top, x1, bottom = _pad_rect((0.0, 0.0, 612.0, 20.0), page_width=612.0, page_height=792.0)
    assert (x0, top) == (0.0, 0.0)
    assert x1 == 612.0 and bottom <= 792.0


def test_scanned_page_is_ocrd_once_not_once_per_tickmark(annotated_setup, tmp_path, monkeypatch):
    # OCR (~1.3s/page) dwarfs rendering (~0.06s). Calling it per mark made a
    # multi-citation page several times slower for identical output.
    from agent import wordboxes

    calls = {"n": 0}
    real = wordboxes.ocr_line_boxes

    def counting(img):
        calls["n"] += 1
        return real(img)

    monkeypatch.setattr(wordboxes, "ocr_line_boxes", counting)

    spec, results, support = annotated_setup
    conclusion = results["TS-4.2"]["conclusion"]
    # Three citations, all on the same page, none findable in the text layer
    # so every one falls through to the OCR path.
    cit = conclusion.evidence_citations[0]
    results = {
        "TS-4.2": {
            "conclusion": conclusion.model_copy(
                update={
                    "evidence_citations": [
                        cit.model_copy(update={"quote_or_summary": f"unfindable value {n}"}) for n in range(3)
                    ]
                }
            ),
            "audit_log": [],
        }
    }

    build_workpaper(spec, results, "PY.xlsx", tmp_path, support_dir=support)
    assert calls["n"] <= 1, f"OCR ran {calls['n']} times for one page"


def test_pdf_workpaper_embeds_annotated_exhibit(annotated_setup, tmp_path: Path):
    spec, results, support = annotated_setup
    path = build_workpaper(spec, results, "PY_Testing_C14.pdf", tmp_path, support_dir=support, fmt="pdf")

    with pdfplumber.open(path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
        image_count = sum(len(page.images) for page in pdf.pages)
    assert "Evidence exhibits" in text
    assert image_count >= 1


def test_workpaper_without_support_dir_is_text_only(annotated_setup, tmp_path: Path):
    # No support_dir (or a failed re-extraction) must degrade to the plain
    # citation table, never break the build.
    spec, results, _ = annotated_setup
    path = build_workpaper(spec, results, "PY_Testing_C14.xlsx", tmp_path)
    wb = openpyxl.load_workbook(path)
    assert len(wb["Summary"]._images) == 0


def test_excel_cited_evidence_gets_text_excerpt(spec: dict, results: dict, tmp_path: Path):
    # Excel-sourced citations can't be screenshotted -- the cited range's
    # extracted rows are excerpted into the workpaper instead.
    support = tmp_path / "support"
    support.mkdir()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Item", "Amount"])
    ws.append(["October accrual", "482110"])
    wb.save(support / "Recon_Oct2026.xlsx")
    spec = {**spec, "cy_support_files": ["Recon_Oct2026.xlsx"]}

    path = build_workpaper(spec, results, "PY_Testing_C14.xlsx", tmp_path, support_dir=support)
    out = openpyxl.load_workbook(path)
    text = " ".join(str(c.value) for r in out["TS-4.2 Summary - Exhibits"].iter_rows() for c in r if c.value is not None)
    assert "Exhibit A" in text
    assert "October accrual | 482110" in text


def test_step_sheet_puts_answers_before_supporting_detail(spec: dict, results: dict, tmp_path: Path):
    # A reviewer opening the sheet must see the verdict, coverage, IPE
    # status, exceptions and open requests before scrolling into the
    # narrative and evidence table -- inline exhibits previously pushed all
    # of that ~110 rows down the page.
    path = build_workpaper(spec, results, "PY_Testing_C14.xlsx", tmp_path)
    ws = openpyxl.load_workbook(path)["TS-4.2 Summary"]

    rows = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.isupper() and len(cell.value) > 3:
                rows.setdefault(cell.value, cell.row)

    assert rows["CONCLUSION"] < rows["EXCEPTIONS"] < rows["ADDITIONAL SUPPORT REQUESTED"]
    assert rows["ADDITIONAL SUPPORT REQUESTED"] < rows["DOCUMENTATION"]
    assert rows["DOCUMENTATION"] < rows["PROCEDURES PERFORMED"] < rows["EVIDENCE CITED"]
    assert rows["EVIDENCE CITED"] < rows["PREPARED BY"]


def test_empty_exceptions_says_none_not_silence(spec: dict, results: dict, tmp_path: Path):
    # "no exceptions" and "we didn't look" must not be indistinguishable in
    # a workpaper -- an empty section is a real audit assertion.
    path = build_workpaper(spec, results, "PY_Testing_C14.xlsx", tmp_path)
    ws = openpyxl.load_workbook(path)["TS-4.2 Summary"]
    text = " ".join(str(c.value) for r in ws.iter_rows() for c in r if c.value is not None)
    assert "None noted." in text


def test_summary_sheet_carries_ipe_and_open_request_counts(spec: dict, results: dict, tmp_path: Path):
    path = build_workpaper(spec, results, "PY_Testing_C14.xlsx", tmp_path)
    ws = openpyxl.load_workbook(path)["TS-4.2 Summary"]
    text = " ".join(str(c.value) for r in ws.iter_rows() for c in r if c.value is not None)
    assert "IPE status" in text
    assert "Open requests" in text
    assert "not_applicable" in text  # from the fixture's conclusion


def test_pdf_workpaper_contents(spec: dict, results: dict, tmp_path: Path):
    path = build_workpaper(spec, results, "PY_Testing_C14.pdf", tmp_path, fmt="pdf")

    assert path.suffix == ".pdf"
    with pdfplumber.open(path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    assert "DRAFT" in text
    assert "C-14" in text
    assert "Satisfied" in text
    assert "INCOMPLETE" in text
    assert "ev_0001" in text


def test_every_box_carries_its_tickmark_letter(annotated_setup, tmp_path: Path):
    # One citation can legitimately mark several values on a page (an
    # invoice number AND its amount). Labelling only the first left the
    # others as unattributed red rectangles, which reads as the tool having
    # boxed the wrong thing -- the first reaction a real reviewer had.
    from agent.workpaper import _render_pdf_exhibit
    from agent.extraction import extract

    _spec, _results, support = annotated_setup
    item = extract(support / "approval.pdf")[0]
    buf, _size, overlays = _render_pdf_exhibit(
        support / "approval.pdf",
        1,
        [("A", item, "Approved / Vendor Number 21022452 / Cost Center 0510")],
    )

    boxes = [o for o in overlays if o.name.startswith("box")]
    letters = [o for o in overlays if o.name.startswith("tickmark")]
    assert boxes, "nothing was located to box"
    assert len(letters) == len(boxes), "some boxes have no tickmark letter"
    # Each letter sits beside its own box, not stacked on the first one.
    assert len({(o.x, o.y) for o in letters}) == len(letters)


def test_source_workbook_is_attached_as_real_tabs(spec: dict, results: dict, tmp_path: Path):
    # A rendered excerpt of the population is a worse artefact than the
    # population itself -- a reviewer wants to filter it, tie it out, and
    # see the Parameters screenshot at full size.
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
    from io import BytesIO

    support = tmp_path / "support"
    support.mkdir()
    src = openpyxl.Workbook()
    sel = src.active
    sel.title = "Sample Selections"
    sel.append(["Sample #", "Invoice"])
    sel.append([1, "2859"])
    pop = src.create_sheet("Population")
    pop.append(["Invoice", "Amount"])
    for i in range(30):
        pop.append([f"INV-{i}", i * 100])
    params = src.create_sheet("Parameters")
    buf = BytesIO()
    PILImage.new("RGB", (80, 40), "white").save(buf, "PNG")
    buf.seek(0)
    params.add_image(XLImage(buf), "A1")
    src.save(support / "listing.xlsx")

    spec = {**spec, "cy_support_files": ["listing.xlsx"]}
    path = build_workpaper(
        spec,
        {"TS-4.2": results["TS-4.2"]},
        "PY.xlsx",
        tmp_path,
        support_dir=support,
        source_workbook="listing.xlsx",
    )
    wb = openpyxl.load_workbook(path)

    for name in ("Sample Selections", "Population", "Parameters"):
        assert name in wb.sheetnames, wb.sheetnames

    # Real data, not a picture of it.
    assert wb["Population"].max_row == 31
    assert wb["Population"]["A2"].value == "INV-0"
    # A tab whose whole content is a pasted screenshot keeps the screenshot.
    assert len(wb["Parameters"]._images) == 1


def test_source_tabs_do_not_collide_with_generated_sheets(spec: dict, results: dict, tmp_path: Path):
    # An uploaded tab named "Summary" must not overwrite the workpaper's own.
    support = tmp_path / "support"
    support.mkdir()
    src = openpyxl.Workbook()
    src.active.title = "Summary"
    src.active.append(["uploaded", "data"])
    src.save(support / "listing.xlsx")

    spec = {**spec, "cy_support_files": ["listing.xlsx"]}
    path = build_workpaper(
        spec,
        {"TS-4.2": results["TS-4.2"]},
        "PY.xlsx",
        tmp_path,
        support_dir=support,
        source_workbook="listing.xlsx",
    )
    wb = openpyxl.load_workbook(path)
    assert "Summary" in wb.sheetnames
    # The generated Summary still holds the conclusion, not the upload.
    text = " ".join(str(c.value) for r in wb["Summary"].iter_rows() for c in r if c.value is not None)
    assert "DRAFT" in text
    assert len([n for n in wb.sheetnames if n.startswith("Summary")]) >= 2


def test_evidence_is_labelled_in_plain_english():
    # "ev_0004" tells a reviewer nothing about what they are looking at.
    from agent.workpaper import _evidence_labels

    cits = [
        EvidenceCitation(
            evidence_id=ev,
            source_file=f,
            location=loc,
            quote_or_summary="q",
            relevance="r",
        )
        for ev, f, loc in [
            # A workbook: the preparer's own sheet name is already the right
            # word -- better than anything derived from the filename.
            ("ev_0018", "SS.156 Sample listing upload.xlsx", "Population!A1:T31"),
            ("ev_0022", "SS.156 Sample listing upload.xlsx", "Parameters!image1"),
            ("ev_0017", "SS.156 Sample listing upload.xlsx", "Sample Selections!A1:U3"),
            # A PDF: read the kind off the filename. This one names both
            # "invoice" and "approval" -- the document IS the approval.
            ("ev_0008", "SS.156 selection 1 - invoice 2859 approval.pdf", "p.1"),
            ("ev_0007", "SS.156 selection 1 invoice 2859.pdf", "p.1"),
            ("ev_0006", "Selection 1 Payment.pdf", "p.1"),
        ]
    ]
    labels = _evidence_labels(cits)

    assert labels["ev_0018"] == "Population"
    assert labels["ev_0022"] == "Parameters"
    assert labels["ev_0017"] == "Sample Selections"
    assert labels["ev_0008"] == "Approval p.1"
    assert labels["ev_0007"] == "Invoice p.1"
    assert labels["ev_0006"] == "Payment p.1"


def test_same_document_cited_twice_stays_distinguishable():
    from agent.workpaper import _evidence_labels

    cits = [
        EvidenceCitation(
            evidence_id=ev,
            source_file="approval.pdf",
            location="approval.pdf p.1",
            quote_or_summary="q",
            relevance="r",
        )
        for ev in ("ev_0001", "ev_0002")
    ]
    labels = _evidence_labels(cits)
    assert labels["ev_0001"] == "Approval p.1"
    assert labels["ev_0002"] == "Approval p.1 (2)"
