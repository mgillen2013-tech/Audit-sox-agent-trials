"""The joint: agent conclusions -> a workpaper that mirrors PY.

These are the tests that would have caught the two halves drifting apart.
Both existed and neither called the other for several rounds of work, so
the interesting assertions here are about the SEAM, not either side.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from agent.build_cy_workpaper import build_cy_workpaper
from agent.schemas import ConclusionOutput, ModelMetadata, SampleItem, SamplePopulationManifest


def _template(path: Path, *, sample: str = "Sample", population: str = "Population") -> Path:
    """A workbook shaped like a PY workpaper.

    The distinct number formats are not decoration. check_template refuses
    a template whose styles.xml defines fewer cell formats than the
    STYLE_* constants address, and a default openpyxl workbook has about
    two -- so a "minimal" fixture silently exercises the FALLBACK path
    while looking like it tests the mirroring one. Real Excel workpapers
    carry plenty; this makes the fixture resemble one.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sample
    ws["A1"] = "Invoice Number"
    ws["A2"] = "existing sample row"  # the row the builder replaces
    pop = wb.create_sheet(population)
    pop["A5"] = "Invoice Number"
    for r in range(6, 12):
        pop.cell(r, 1, f"population row {r}")
    for i in range(20):
        c = ws.cell(50 + i, 1, i)
        c.number_format = f"0.{'0' * (i % 8)}#{'#' * (i % 3)}"
    wb.save(path)
    return path


def _real_pdf(path: Path) -> Path:
    """A PDF that actually opens. b"%PDF-1.4" does not -- it is a valid file
    header and nothing else, and every renderer rejects it."""
    from reportlab.lib.pagesizes import letter as PAGE
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(str(path), pagesize=PAGE)
    c.setFont("Helvetica", 12)
    c.drawString(72, 700, "Invoice Number: 21022452")
    c.save()
    return path


def _spec() -> dict:
    return {
        "control_id": "C-14",
        "control_objective_ref": "CO-1",
        "control_objective_text": "Invoices are approved before payment.",
        "py_testing_file": "PY.xlsx",
        "cy_support_files": [],
        "test_steps": [{"test_step_id": "TS-1", "test_step_text": "Verify approval precedes payment."}],
    }


def _results(**overrides) -> dict:
    base = dict(
        test_step_id="TS-1",
        control_objective_ref="CO-1",
        conclusion="satisfied",
        narrative="Approval preceded payment.",
        evidence_citations=[
            {
                "evidence_id": "ev_0001",
                "source_file": "invoice.pdf",
                "location": "invoice.pdf p.1",
                "quote_or_summary": "Invoice 21022452",
                "relevance": "the invoice",
                "sample_id": "1",
            }
        ],
        procedures_performed=["inspection"],
        sample_results=[{"sample_id": "1", "conclusion": "satisfied"}],
        attribute_results=[
            {
                "attribute": "Invoice number agrees to the listing",
                "sample_id": "1",
                "result": "satisfied",
                "value_observed": "Invoice 21022452",
                "evidence_ids": ["ev_0001"],
            }
        ],
        ipe_completeness_accuracy_status="not_applicable",
        ipe_completeness_accuracy_evidence=[],
        exceptions=[],
        additional_support_requests=[],
        confidence="high",
        confidence_rationale="Clear.",
        sample_coverage=None,
        model_metadata=ModelMetadata(
            model="claude-opus-5", prompt_version="v1", timestamp="2026-09-03T00:00:00Z", tool_call_count=3
        ),
    )
    base.update(overrides)
    return {"TS-1": {"conclusion": ConclusionOutput(**base), "audit_log": []}}


def _manifests() -> dict:
    return {
        "TS-1": SamplePopulationManifest(
            test_step_id="TS-1",
            sample_size=1,
            samples=[
                SampleItem(
                    sample_id="1",
                    test_step_id="TS-1",
                    identifying_details="Invoice 21022452",
                    key_fields={"Invoice Number": "21022452"},
                )
            ],
        )
    }


def test_a_usable_py_template_is_mirrored(tmp_path: Path):
    _template(tmp_path / "PY.xlsx")
    _real_pdf(tmp_path / "invoice.pdf")
    outcome = build_cy_workpaper(
        _spec(), _results(), "PY.xlsx", tmp_path / "out",
        support_dir=tmp_path, sample_manifests=_manifests(),
    )
    assert outcome.mirrored_py
    wb = openpyxl.load_workbook(outcome.path)
    # The template's own tabs survive, with the Summary added in front.
    assert wb.sheetnames[0] == "Summary"
    assert {"Sample", "Population"} <= set(wb.sheetnames)
    wb.close()


def test_a_pdf_prior_year_falls_back_instead_of_failing(tmp_path: Path):
    # Plenty of PY workpapers are PDFs. A run that reached real conclusions
    # has already spent real money and real evidence review -- a
    # worse-looking file carrying them beats an error every time.
    (tmp_path / "PY.pdf").write_bytes(b"%PDF-1.4\n")
    outcome = build_cy_workpaper(
        _spec(), _results(), "PY.pdf", tmp_path / "out",
        support_dir=tmp_path, sample_manifests=_manifests(),
    )
    assert outcome.path.exists()
    assert not outcome.mirrored_py
    # And it must SAY so -- otherwise the only clue is that the file looks
    # nothing like last year's, which reads as a regression.
    assert any("could not mirror" in w for w in outcome.warnings)


def test_a_wrongly_shaped_workbook_falls_back_rather_than_writing_into_it(tmp_path: Path):
    # Named .xlsx but shaped for a different control. Writing into it would
    # put the sample rows in whatever tab happened to be second.
    _template(tmp_path / "PY.xlsx", sample="SomethingElse", population="Other")
    outcome = build_cy_workpaper(
        _spec(), _results(), "PY.xlsx", tmp_path / "out",
        support_dir=tmp_path, sample_manifests=_manifests(),
    )
    assert not outcome.mirrored_py
    assert outcome.path.exists()


def test_a_control_where_every_step_failed_still_produces_a_workpaper(tmp_path: Path):
    _template(tmp_path / "PY.xlsx")
    failed = {"TS-1": {"error": "token budget exceeded", "reason": "token_budget_exceeded",
                       "tokens_used": 1, "turns_used": 1, "audit_log": []}}
    outcome = build_cy_workpaper(
        _spec(), failed, "PY.xlsx", tmp_path / "out", support_dir=tmp_path,
    )
    assert outcome.path.exists()
    assert not outcome.mirrored_py


def test_the_population_is_never_emptied_by_the_joint(tmp_path: Path):
    # The template's population must survive when the caller has none of
    # its own -- deleting it removes the basis for the sample.
    _template(tmp_path / "PY.xlsx")
    _real_pdf(tmp_path / "invoice.pdf")
    outcome = build_cy_workpaper(
        _spec(), _results(), "PY.xlsx", tmp_path / "out",
        support_dir=tmp_path, sample_manifests=_manifests(),
    )
    wb = openpyxl.load_workbook(outcome.path)
    rows = [r for r in wb["Population"].iter_rows(min_row=6, values_only=True) if any(v is not None for v in r)]
    wb.close()
    assert len(rows) == 6


def test_a_corrupt_evidence_file_loses_one_exhibit_not_the_workpaper(tmp_path: Path):
    # Support arrives from scanners, mail clients and portals, and some of
    # it is broken. Losing one exhibit is recoverable; losing the workpaper
    # -- and with it every conclusion the run paid for -- is not.
    _template(tmp_path / "PY.xlsx")
    (tmp_path / "invoice.pdf").write_bytes(b"%PDF-1.4\n")  # a header and nothing else
    outcome = build_cy_workpaper(
        _spec(), _results(), "PY.xlsx", tmp_path / "out",
        support_dir=tmp_path, sample_manifests=_manifests(),
    )
    assert outcome.path.exists() and outcome.mirrored_py
    assert any("could not be read" in reason for *_x, reason in outcome.unplaced_callouts)
    assert "Wrote" in outcome.summary_line


# --------------------------------------------------------------------------
# One tab per sampled item
# --------------------------------------------------------------------------


def _two_sample_results():
    r = _results(
        conclusion="not_satisfied",
        sample_results=[
            {"sample_id": "1", "conclusion": "satisfied"},
            {"sample_id": "2", "conclusion": "not_satisfied", "note": "no approval"},
        ],
        attribute_results=[
            {"attribute": "Invoice agrees", "sample_id": "1", "result": "satisfied",
             "value_observed": "21022452", "evidence_ids": ["ev_0001"]},
            {"attribute": "Invoice agrees", "sample_id": "2", "result": "satisfied",
             "value_observed": "21022453", "evidence_ids": ["ev_0001"]},
        ],
    )
    return r


def _two_manifests():
    return {
        "TS-1": SamplePopulationManifest(
            test_step_id="TS-1", sample_size=2,
            samples=[
                SampleItem(sample_id="1", test_step_id="TS-1", identifying_details="a",
                           key_fields={"Invoice Number": "21022452"}),
                SampleItem(sample_id="2", test_step_id="TS-1", identifying_details="b",
                           key_fields={"Invoice Number": "21022453"}),
            ],
        )
    }


def test_each_sampled_item_gets_its_own_tab(tmp_path: Path):
    # The prior-year workpaper documents ONE selection per Sample tab, so
    # two selections mean two tabs shaped like that one -- not two rows
    # crammed onto it, which is what stacked their exhibits and produced
    # the large empty regions a reviewer flagged.
    _template(tmp_path / "PY.xlsx")
    _real_pdf(tmp_path / "invoice.pdf")
    outcome = build_cy_workpaper(
        _spec(), _two_sample_results(), "PY.xlsx", tmp_path / "out",
        support_dir=tmp_path, sample_manifests=_two_manifests(),
    )
    wb = openpyxl.load_workbook(outcome.path)
    assert "Sample 1" in wb.sheetnames and "Sample 2" in wb.sheetnames
    # Sibling tabs stay adjacent. Appending clones at the end scattered them
    # after Population and Parameters, which reads as a mistake in a
    # workpaper a reviewer pages through in order.
    names = wb.sheetnames
    assert names.index("Sample 2") == names.index("Sample 1") + 1
    wb.close()


def test_a_single_sample_keeps_the_templates_own_tab_name(tmp_path: Path):
    # The ordinary case must stay indistinguishable from last year's file.
    _template(tmp_path / "PY.xlsx")
    _real_pdf(tmp_path / "invoice.pdf")
    outcome = build_cy_workpaper(
        _spec(), _results(), "PY.xlsx", tmp_path / "out",
        support_dir=tmp_path, sample_manifests=_manifests(),
    )
    wb = openpyxl.load_workbook(outcome.path)
    assert "Sample" in wb.sheetnames
    assert not any(n.startswith("Sample ") for n in wb.sheetnames)
    wb.close()


def test_every_summary_link_points_at_a_tab_that_exists(tmp_path: Path):
    # Tab names and summary links are derived from the same sample_id for
    # exactly this reason. A first version named tabs after the check number
    # while the links used the id, so every link pointed at a tab that did
    # not exist -- and Excel does not complain, the click just does nothing.
    _template(tmp_path / "PY.xlsx")
    _real_pdf(tmp_path / "invoice.pdf")
    outcome = build_cy_workpaper(
        _spec(), _two_sample_results(), "PY.xlsx", tmp_path / "out",
        support_dir=tmp_path, sample_manifests=_two_manifests(),
    )
    wb = openpyxl.load_workbook(outcome.path)
    ws = wb["Summary"]
    targets = [
        ws.cell(r, 5).hyperlink.location.split("!")[0].strip("'")
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, 5).hyperlink
    ]
    assert targets
    assert all(t in wb.sheetnames for t in targets), (targets, wb.sheetnames)
    wb.close()
