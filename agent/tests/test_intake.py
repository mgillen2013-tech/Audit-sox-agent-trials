from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from agent.intake import (
    build_manifest_from_any_columns,
    build_manifests_from_rows,
    parse_sample_list,
    read_excel_rows,
)


def _write_sample_list(path: Path, rows: list[list]) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    header = [
        "test_step_id",
        "sample_id",
        "identifying_details",
        "population_description",
        "population_size",
        "selection_method",
        "po_number",
        "branch",
    ]
    ws.append(header)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


@pytest.fixture
def sample_list_xlsx(tmp_path: Path) -> Path:
    return _write_sample_list(
        tmp_path / "samples.xlsx",
        rows=[
            ["TS-4.2", "S01", "October accrual", "Monthly accruals FY2026", 12, "random", "", ""],
            ["TS-1.1", "S01", "PO 48213", "All POs > $5,000", 340, "random", "48213", "35210"],
            ["TS-1.1", "S02", "PO 48214", "All POs > $5,000", 340, "random", "48214", "35211"],
        ],
    )


def test_parses_one_manifest_per_test_step(sample_list_xlsx: Path):
    manifests = parse_sample_list(sample_list_xlsx)
    assert set(manifests) == {"TS-4.2", "TS-1.1"}


def test_sample_size_is_computed_not_read(sample_list_xlsx: Path):
    manifests = parse_sample_list(sample_list_xlsx)
    assert manifests["TS-1.1"].sample_size == 2
    assert manifests["TS-4.2"].sample_size == 1


def test_extra_columns_become_key_fields(sample_list_xlsx: Path):
    manifests = parse_sample_list(sample_list_xlsx)
    s01 = manifests["TS-1.1"].samples[0]
    assert s01.key_fields == {"po_number": "48213", "branch": "35210"}


def test_no_extra_columns_gives_none_key_fields(sample_list_xlsx: Path):
    manifests = parse_sample_list(sample_list_xlsx)
    accrual_sample = manifests["TS-4.2"].samples[0]
    assert accrual_sample.key_fields is None


def test_population_info_taken_from_first_row_per_step(sample_list_xlsx: Path):
    manifests = parse_sample_list(sample_list_xlsx)
    assert manifests["TS-1.1"].population_description == "All POs > $5,000"
    assert manifests["TS-1.1"].population_size == 340
    assert manifests["TS-1.1"].selection_method == "random"


def test_missing_required_column_raises(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["test_step_id", "sample_id"])  # missing identifying_details etc.
    ws.append(["TS-4.2", "S01"])
    path = tmp_path / "bad.xlsx"
    wb.save(path)

    with pytest.raises(ValueError, match="missing required column"):
        parse_sample_list(path)


def test_invalid_selection_method_raises(tmp_path: Path):
    path = _write_sample_list(
        tmp_path / "bad_method.xlsx",
        rows=[["TS-4.2", "S01", "October accrual", "Monthly", 12, "vibes", "", ""]],
    )
    with pytest.raises(ValueError, match="selection_method"):
        parse_sample_list(path)


def test_blank_rows_are_skipped(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        [
            "test_step_id",
            "sample_id",
            "identifying_details",
            "population_description",
            "population_size",
            "selection_method",
        ]
    )
    ws.append(["TS-4.2", "S01", "October accrual", "Monthly", 12, "random"])
    ws.append([None, None, None, None, None, None])
    path = tmp_path / "with_blank.xlsx"
    wb.save(path)

    manifests = parse_sample_list(path)
    assert manifests["TS-4.2"].sample_size == 1


# --------------------------------------------------------------------------
# build_manifests_from_rows directly -- what the Streamlit app calls
# --------------------------------------------------------------------------


def test_build_manifests_from_rows_matches_excel_path():
    rows = [
        {
            "test_step_id": "TS-1.1",
            "sample_id": "S01",
            "identifying_details": "PO 48213",
            "population_description": "All POs > $5,000",
            "selection_method": "random",
            "population_size": 340,
        },
        {
            "test_step_id": "TS-1.1",
            "sample_id": "S02",
            "identifying_details": "PO 48214",
            "population_description": "All POs > $5,000",
            "selection_method": "random",
            "population_size": 340,
        },
    ]
    manifests = build_manifests_from_rows(rows)
    assert manifests["TS-1.1"].sample_size == 2
    assert manifests["TS-1.1"].samples[0].sample_id == "S01"


def test_build_manifests_from_rows_rejects_missing_required_field():
    with pytest.raises(ValueError, match="is missing"):
        build_manifests_from_rows([{"test_step_id": "TS-1.1", "sample_id": "S01"}])


# --------------------------------------------------------------------------
# build_manifest_from_any_columns / read_excel_rows -- the arbitrary-column
# path, exercised against data shaped like a real E1 AP payment export
# (this is what actually broke the fixed-column parser in practice).
# --------------------------------------------------------------------------

_E1_STYLE_ROWS = [
    {
        "Sample Selection #": 1,
        "invoice number f0411.vinv": "IN-88213",
        "invoice payment amount f0414.paap": 12450.00,
        "business unit f0414.mcu": "35210",
        "check/ item f0413.docm": "48213",
    },
    {
        "Sample Selection #": 2,
        "invoice number f0411.vinv": "IN-88214",
        "invoice payment amount f0414.paap": 8300.00,
        "business unit f0414.mcu": "35211",
        "check/ item f0413.docm": "48214",
    },
]


def test_build_manifest_from_any_columns_uses_detected_id_column():
    manifest = build_manifest_from_any_columns(
        _E1_STYLE_ROWS,
        test_step_id="T1.1",
        population_description="All AP payments issued in the test period",
        selection_method="random",
        population_size=156,
    )
    assert manifest.sample_size == 2
    assert manifest.samples[0].sample_id == "1"
    assert manifest.samples[1].sample_id == "2"


def test_build_manifest_from_any_columns_builds_identifying_details_and_key_fields():
    manifest = build_manifest_from_any_columns(
        _E1_STYLE_ROWS, test_step_id="T1.1", population_description="AP payments", selection_method="random"
    )
    sample = manifest.samples[0]
    assert "invoice number f0411.vinv: IN-88213" in sample.identifying_details
    assert sample.key_fields["business unit f0414.mcu"] == "35210"


def test_build_manifest_from_any_columns_falls_back_to_row_number_without_id_column():
    rows = [{"invoice number": "IN-1"}, {"invoice number": "IN-2"}]
    manifest = build_manifest_from_any_columns(
        rows, test_step_id="T1.1", population_description="AP payments", selection_method="random"
    )
    assert manifest.samples[0].sample_id == "row_1"
    assert manifest.samples[1].sample_id == "row_2"


def test_build_manifest_from_any_columns_rejects_bad_selection_method():
    with pytest.raises(ValueError, match="selection_method"):
        build_manifest_from_any_columns(
            _E1_STYLE_ROWS, test_step_id="T1.1", population_description="AP payments", selection_method="vibes"
        )


def test_read_excel_rows_matches_real_export_shape(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Sample Selection #", "invoice number f0411.vinv", "business unit f0414.mcu"])
    ws.append([1, "IN-88213", "35210"])
    ws.append([2, "IN-88214", "35211"])
    path = tmp_path / "T1.1 SS.156 Population & Samples.xlsx"
    wb.save(path)

    rows = read_excel_rows(path)
    assert len(rows) == 2
    assert rows[0]["Sample Selection #"] == 1
    assert rows[0]["invoice number f0411.vinv"] == "IN-88213"


def test_build_manifest_from_any_columns_selection_method_optional():
    # The real workflow no longer collects selection_method -- must build a
    # valid manifest without it, not silently require a fake-good value.
    manifest = build_manifest_from_any_columns(
        _E1_STYLE_ROWS, test_step_id="T1.1", population_description="AP payments"
    )
    assert manifest.selection_method is None
    assert manifest.sample_size == 2


def test_read_excel_rows_targets_a_specific_tab(tmp_path: Path):
    # The real shape this is built for: one workbook, population on one
    # tab, the sample selections on another.
    wb = openpyxl.Workbook()
    pop_ws = wb.active
    pop_ws.title = "Population"
    pop_ws.append(["invoice number", "amount"])
    for i in range(1, 6):
        pop_ws.append([f"IN-{i}", i * 100])

    sample_ws = wb.create_sheet("Sample")
    sample_ws.append(["Sample Selection #", "invoice number"])
    sample_ws.append([1, "IN-2"])
    path = tmp_path / "T1.1 SS.156 Population & Samples.xlsx"
    wb.save(path)

    pop_rows = read_excel_rows(path, sheet_name="Population")
    sample_rows = read_excel_rows(path, sheet_name="Sample")

    assert len(pop_rows) == 5
    assert len(sample_rows) == 1
    assert sample_rows[0]["invoice number"] == "IN-2"

    # Default (no sheet_name) still reads the active sheet, unchanged
    # behavior for a single-tab file.
    assert read_excel_rows(path) == pop_rows
