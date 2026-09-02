"""Integration test: real files (synthetic, not hand-built EvidenceItems) ->
real extraction -> real sample-list parsing -> the tool loop, end to end,
against a fake client. This is the thing agent/example_run.py could never
prove, since it skipped extraction entirely.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from agent.run_control import iter_control_results, run_control
from agent.tests.conftest import FakeClient, fake_response as response, tool_use


@pytest.fixture
def control_dir(tmp_path: Path) -> tuple[Path, dict]:
    py_wb = openpyxl.Workbook()
    py_ws = py_wb.active
    py_ws.append(["Item", "PY Result"])
    py_ws.append(["Accrual", "Recalculated, agreed to GL, no variance"])
    py_wb.save(tmp_path / "PY_Testing_C14.xlsx")

    cy_wb = openpyxl.Workbook()
    cy_ws = cy_wb.active
    cy_ws.append(["Item", "Description"])
    cy_ws.append(["Accrual", "Recalculation ties to GL export with no variance"])
    cy_wb.save(tmp_path / "Recon_Oct2026.xlsx")

    sl_wb = openpyxl.Workbook()
    sl_ws = sl_wb.active
    sl_ws.append(
        [
            "test_step_id",
            "sample_id",
            "identifying_details",
            "population_description",
            "population_size",
            "selection_method",
        ]
    )
    sl_ws.append(["TS-4.2", "S01", "October accrual", "Monthly accruals FY2026", 12, "random"])
    sl_wb.save(tmp_path / "samples_C14.xlsx")

    spec = {
        "control_id": "C-14",
        "control_objective_ref": "CO-4",
        "control_objective_text": "Accruals are recorded completely and accurately.",
        "py_testing_file": "PY_Testing_C14.xlsx",
        "sample_list_file": "samples_C14.xlsx",
        "cy_support_files": ["Recon_Oct2026.xlsx"],
        "test_steps": [
            {
                "test_step_id": "TS-4.2",
                "test_step_text": "Recalculate the accrual and agree to the GL.",
                "py_conclusion_text": "Satisfied. No exceptions noted.",
            }
        ],
    }
    return tmp_path, spec


def _submit_conclusion_input(**overrides) -> dict:
    base = dict(
        test_step_id="TS-4.2",
        control_objective_ref="CO-4",
        conclusion="satisfied",
        narrative="Recalculated accrual ties to GL with no variance.",
        evidence_citations=[
            {
                "evidence_id": "ev_0001",
                "source_file": "Recon_Oct2026.xlsx",
                "location": "Sheet1!A1:B2",
                "quote_or_summary": "Recalculation ties to GL export with no variance",
                "relevance": "CY support for the accrual recalculation",
            }
        ],
        procedures_performed=["recalculation"],
        ipe_completeness_accuracy_status="not_applicable",
        ipe_completeness_accuracy_evidence=[],
        exceptions=[],
        additional_support_requests=[],
        confidence="high",
        confidence_rationale="Single clear source, full sample coverage.",
    )
    base.update(overrides)
    return base


def test_iter_control_results_yields_incrementally(control_dir: tuple[Path, dict]):
    # The Streamlit app relies on this being a real generator (results
    # available one at a time) rather than something that silently buffers
    # everything before yielding the first item.
    import inspect

    base_dir, spec = control_dir
    client = FakeClient(
        responses=[
            response([tool_use("t1", "search_cy_support", {"query": "accrual recalculation GL", "top_k": 5})]),
            response(
                [
                    tool_use(
                        "t2",
                        "check_sample_coverage",
                        {"required_sample_ids": [], "found_evidence_ids": ["S01"]},
                    )
                ]
            ),
            response([tool_use("t3", "submit_conclusion", _submit_conclusion_input())]),
        ]
    )

    gen = iter_control_results(spec, base_dir, client, model="claude-opus-5")
    assert inspect.isgenerator(gen)

    test_step_id, result = next(gen)
    assert test_step_id == "TS-4.2"
    assert result["conclusion"].conclusion == "satisfied"
    with pytest.raises(StopIteration):
        next(gen)


def test_run_control_wires_real_files_through_the_loop(control_dir: tuple[Path, dict]):
    base_dir, spec = control_dir
    client = FakeClient(
        responses=[
            response([tool_use("t1", "search_cy_support", {"query": "accrual recalculation GL", "top_k": 5})]),
            response(
                [
                    tool_use(
                        "t2",
                        "check_sample_coverage",
                        {"required_sample_ids": [], "found_evidence_ids": ["S01"]},
                    )
                ]
            ),
            response([tool_use("t3", "submit_conclusion", _submit_conclusion_input())]),
        ]
    )

    results = run_control(spec, base_dir, client, model="claude-opus-5")

    assert "TS-4.2" in results
    assert "error" not in results["TS-4.2"]
    conclusion = results["TS-4.2"]["conclusion"]
    assert conclusion.conclusion == "satisfied"
    assert conclusion.sample_coverage.complete is True
    assert conclusion.evidence_citations[0].evidence_id == "ev_0001"
    assert len(results["TS-4.2"]["audit_log"]) == 3

    # The manifest's sample_size (1 row) and population_size (12, from the
    # sample list fixture) must reach the model's first turn -- without
    # this the model has no way to tell a correct small sample from a
    # partial upload (the real bug this was built to fix). Content is in
    # block form by call time (agent/loop.py's cache-breakpoint pass
    # converts the bare string), so extract the text out of it.
    first_turn_content = client.calls[0]["messages"][0]["content"]
    first_turn_text = (
        first_turn_content
        if isinstance(first_turn_content, str)
        else " ".join(b.get("text", "") for b in first_turn_content if isinstance(b, dict))
    )
    assert "Sample: 1 item(s) selected for testing from a population of 12." in first_turn_text


def test_run_control_missing_sample_list_row_still_runs(control_dir: tuple[Path, dict]):
    # A test step with no matching rows in the sample list -> empty manifest
    # lookup -> check_sample_coverage should hit the no_sample_list_found path.
    base_dir, spec = control_dir
    spec["test_steps"].append(
        {
            "test_step_id": "TS-9.9",
            "test_step_text": "A step with no sample list rows.",
            "py_conclusion_text": "N/A",
        }
    )
    client = FakeClient(
        responses=[
            # TS-4.2 first (dict preserves insertion order, and run_control
            # iterates spec["test_steps"] in order)
            response([tool_use("t1", "search_cy_support", {"query": "accrual", "top_k": 5})]),
            response(
                [tool_use("t1b", "check_sample_coverage", {"required_sample_ids": [], "found_evidence_ids": ["S01"]})]
            ),
            response([tool_use("t2", "submit_conclusion", _submit_conclusion_input())]),
            # TS-9.9
            response(
                [
                    tool_use(
                        "t3",
                        "check_sample_coverage",
                        {"required_sample_ids": [], "found_evidence_ids": []},
                    )
                ]
            ),
            response(
                [
                    tool_use(
                        "t4",
                        "submit_conclusion",
                        _submit_conclusion_input(
                            test_step_id="TS-9.9", conclusion="insufficient_evidence", evidence_citations=[]
                        ),
                    )
                ]
            ),
        ]
    )

    results = run_control(spec, base_dir, client, model="claude-opus-5")

    assert results["TS-4.2"]["conclusion"].conclusion == "satisfied"
    assert results["TS-9.9"]["conclusion"].conclusion == "insufficient_evidence"


def test_run_ceiling_stops_later_steps_but_keeps_finished_ones(control_dir: tuple[Path, dict]):
    # The per-step cap resets every step, so a multi-step control could
    # spend N x that with nothing watching the total. The ceiling watches
    # the total -- and must not throw away work already paid for.
    from agent.tests.conftest import fake_usage

    base_dir, spec = control_dir
    spec["test_steps"].append(
        {"test_step_id": "TS-9.9", "test_step_text": "Second step.", "py_conclusion_text": "N/A"}
    )
    client = FakeClient(
        responses=[
            # TS-4.2 completes, spending 60K.
            response(
                [tool_use("t1", "search_cy_support", {"query": "accrual", "top_k": 5})],
                usage=fake_usage(input_tokens=60_000),
            ),
            response(
                [tool_use("t2", "check_sample_coverage", {"required_sample_ids": [], "found_evidence_ids": ["S01"]})]
            ),
            response([tool_use("t3", "submit_conclusion", _submit_conclusion_input())]),
            # TS-9.9 would run next, but the ceiling is already passed.
            response([tool_use("t4", "submit_conclusion", _submit_conclusion_input(test_step_id="TS-9.9"))]),
        ]
    )

    results = run_control(spec, base_dir, client, model="claude-opus-5", max_run_tokens=50_000)

    # Finished step keeps its conclusion -- that work is done and paid for.
    assert results["TS-4.2"]["conclusion"].conclusion == "satisfied"
    # Later step is skipped rather than started.
    assert results["TS-9.9"]["reason"] == "run_budget_exceeded"
    assert "skipped" in results["TS-9.9"]["error"]
    # And it really did not call the model for that step -- only TS-4.2's
    # three turns were spent.
    assert len(client.calls) == 3


def test_run_control_preserves_audit_log_on_incomplete_run(control_dir: tuple[Path, dict]):
    # This is the direct regression test for the real $65 failure: a step
    # that never reaches submit_conclusion must not come back as a bare
    # error string with everything already paid for thrown away.
    base_dir, spec = control_dir
    # run_test_step's default max_iterations is 15 -- iter_control_results
    # doesn't expose a smaller override, so this needs one response per
    # iteration, each a real (never-concluding) tool call.
    client = FakeClient(
        responses=[
            response([tool_use(f"t{i}", "search_cy_support", {"query": "accrual recalculation GL", "top_k": 5})])
            for i in range(15)
        ]
    )

    results = run_control(spec, base_dir, client, model="claude-opus-5")

    result = results["TS-4.2"]
    assert "error" in result
    assert result["reason"] == "max_iterations"
    assert len(result["audit_log"]) == 15
    assert all(entry.tool_name == "search_cy_support" for entry in result["audit_log"])
    assert "tokens_used" in result and "turns_used" in result


def test_the_ledger_captures_a_whole_control_run(control_dir: tuple[Path, dict]):
    # The ledger is threaded through iter_control_results into both OCR and
    # every test step. A ledger that only ever sees the tool loop would
    # under-report a scanned-support control by exactly the amount that
    # runs outside the per-step cap -- which is the spend hardest to
    # explain after the fact and the whole reason for the breakdown.
    from agent.costs import TokenLedger
    from agent.tests.conftest import fake_usage

    base_dir, spec = control_dir
    client = FakeClient(
        responses=[
            response(
                [tool_use("t1", "search_cy_support", {"query": "accrual recalculation GL", "top_k": 5})],
                usage=fake_usage(input_tokens=8_000, output_tokens=300),
            ),
            response(
                [tool_use("t2", "check_sample_coverage", {"required_sample_ids": [], "found_evidence_ids": ["S01"]})],
                usage=fake_usage(input_tokens=200, cache_read_input_tokens=8_200, output_tokens=150),
            ),
            response(
                [tool_use("t3", "submit_conclusion", _submit_conclusion_input())],
                usage=fake_usage(input_tokens=250, cache_read_input_tokens=8_600, output_tokens=1_400),
            ),
        ]
    )

    ledger = TokenLedger("claude-opus-5")
    run_control(spec, base_dir, client, model="claude-opus-5", ledger=ledger)

    assert [g.group for g in ledger.group_rows()] == ["TS-4.2"]
    assert len(ledger.records) == 3
    assert ledger.dollars > 0
    # The step's first prompt was itemised, so "why was turn 1 expensive"
    # is answerable without re-running anything.
    sections = {name for name, *_ in ledger.prompt_mix_rows("TS-4.2")}
    assert {"System prompt", "Tool schemas", "PY support excerpts"} <= sections
