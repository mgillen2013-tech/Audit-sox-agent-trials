"""Tests for the agent -> OOXML-builder translation.

The anchor-text choice is the part worth testing hardest: picking the
wrong token doesn't fail, it draws the red box around the wrong number,
which is worse than no box at all.
"""

from __future__ import annotations

from pathlib import Path

import pytest

from agent.render_bridge import anchor_candidates, build_workpaper_request, sample_items_from_conclusion
from agent.schemas import (
    ConclusionOutput,
    ModelMetadata,
    SampleItem,
    SamplePopulationManifest,
)


# --------------------------------------------------------------------------
# Anchor selection
# --------------------------------------------------------------------------


def test_amounts_are_anchored_whole():
    # "000" is a substring of every round amount on a page. A real run boxed
    # the wrong line items because a naive tokenizer split "$30,000.00" into
    # fragments, so the whole amount must survive as one token.
    anchors = anchor_candidates("Paid $30,000.00 on the invoice")
    assert "30,000.00" in anchors
    assert "000" not in anchors
    assert not any(a in {"30", "00"} for a in anchors)


def test_bare_years_are_dropped():
    # A year matches every date on the document, so it identifies nothing.
    assert "2025" not in anchor_candidates("Approved during 2025")


def test_dates_survive_as_whole_tokens():
    assert "11/10/2025" in anchor_candidates("paid 11/10/2025")


def test_anchors_keep_punctuation():
    # agent/wordboxes.py normalises ("30,000.00" -> "3000000") because it
    # matches normalised text. The OOXML builder does a RAW substring match
    # against the OCR token, so an anchor must look like what is printed.
    assert "30,000.00" in anchor_candidates("total was 30,000.00")


def test_anchors_never_contain_spaces():
    # find_anchor_box rejects a multi-word anchor outright -- word boxes are
    # single tokens.
    for anchor in anchor_candidates("Invoice 21022452 paid 11/10/2025 for $30,000.00"):
        assert " " not in anchor


def test_longest_anchor_comes_first():
    # The caller takes candidates[0], and a longer token is likelier to
    # identify exactly one place on the page.
    anchors = anchor_candidates("inv 21022452 amount 1,250.00")
    assert len(anchors[0]) >= len(anchors[-1])


def test_prose_with_no_distinctive_value_yields_nothing():
    # Better to write the legend entry and skip the box than to anchor on a
    # common word and box a random paragraph.
    assert anchor_candidates("Approved by the appropriate manager") == []


# --------------------------------------------------------------------------
# Translation
# --------------------------------------------------------------------------


def _manifest() -> SamplePopulationManifest:
    return SamplePopulationManifest(
        test_step_id="TS-1",
        sample_size=1,
        samples=[
            SampleItem(
                sample_id="1",
                test_step_id="TS-1",
                identifying_details="Invoice 21022452",
                key_fields={"invoice_number": "21022452", "vendor": "Premiere Onboard"},
            )
        ],
    )


def _conclusion(**overrides) -> ConclusionOutput:
    base = dict(
        test_step_id="TS-1",
        control_objective_ref="CO-1",
        conclusion="satisfied",
        narrative="Approval preceded payment for the selected item.",
        evidence_citations=[
            {
                "evidence_id": "ev_0001",
                "source_file": "invoice.pdf",
                "location": "invoice.pdf p.2",
                "quote_or_summary": "Invoice 21022452 for $30,000.00",
                "relevance": "the invoice",
                "sample_id": "1",
            }
        ],
        procedures_performed=["inspection"],
        sample_results=[{"sample_id": "1", "conclusion": "satisfied"}],
        attribute_results=[
            {
                "attribute": "Invoice amount agrees to payment",
                "sample_id": "1",
                "result": "satisfied",
                "value_observed": "Invoice total $30,000.00",
                "evidence_ids": ["ev_0001"],
            }
        ],
        ipe_completeness_accuracy_status="not_applicable",
        ipe_completeness_accuracy_evidence=[],
        exceptions=[],
        additional_support_requests=[],
        confidence="high",
        confidence_rationale="Native text.",
        sample_coverage=None,
        model_metadata=ModelMetadata(
            model="claude-opus-5", prompt_version="v1", timestamp="2026-09-03T00:00:00Z", tool_call_count=3
        ),
    )
    base.update(overrides)
    return ConclusionOutput(**base)


def test_attributes_become_lettered_tickmarks(tmp_path: Path):
    (tmp_path / "invoice.pdf").write_bytes(b"%PDF-1.4\n")
    items, warnings = sample_items_from_conclusion(_conclusion(), _manifest(), "C-14", tmp_path)

    assert len(items) == 1
    item = items[0]
    assert len(item.evidence_images) == 1
    tie_out = item.evidence_images[0].tie_outs[0]
    assert tie_out.letter == "A"
    # Anchored on the amount from value_observed, punctuation intact.
    assert tie_out.anchor_text == "30,000.00"
    assert not warnings


def test_every_letter_is_defined_in_the_narrative_legend(tmp_path: Path):
    # OoxmlSampleItem's own validator enforces this -- constructing the item
    # at all proves the legend lines are shaped the way it expects. This
    # asserts the content a reviewer actually reads.
    (tmp_path / "invoice.pdf").write_bytes(b"%PDF-1.4\n")
    items, _ = sample_items_from_conclusion(_conclusion(), _manifest(), "C-14", tmp_path)
    legend = " ".join(r.text for p in items[0].narrative for r in p.runs)
    assert "A - " in legend
    assert "Invoice amount agrees to payment" in legend


def test_citation_pages_are_converted_from_1_indexed_to_0_indexed(tmp_path: Path):
    # The agent writes "p.2" meaning the second page; EvidenceImage counts
    # from 0. Off by one here silently boxes a value on the wrong page.
    (tmp_path / "invoice.pdf").write_bytes(b"%PDF-1.4\n")
    items, _ = sample_items_from_conclusion(_conclusion(), _manifest(), "C-14", tmp_path)
    assert items[0].evidence_images[0].pdf_page == 1


def test_a_missing_support_file_warns_instead_of_crashing(tmp_path: Path):
    # Nothing written to tmp_path: the cited file does not exist.
    items, warnings = sample_items_from_conclusion(_conclusion(), _manifest(), "C-14", tmp_path)
    assert items and not items[0].evidence_images
    assert any("not present" in w for w in warnings)


def test_unanchorable_attribute_still_gets_a_legend_entry(tmp_path: Path):
    (tmp_path / "invoice.pdf").write_bytes(b"%PDF-1.4\n")
    conclusion = _conclusion(
        attribute_results=[
            {
                "attribute": "Approved by appropriate personnel",
                "sample_id": "1",
                "result": "satisfied",
                "value_observed": "Approved by the AP manager",  # no distinctive token
                "evidence_ids": ["ev_0001"],
            }
        ]
    )
    items, warnings = sample_items_from_conclusion(conclusion, _manifest(), "C-14", tmp_path)
    legend = " ".join(r.text for p in items[0].narrative for r in p.runs)
    assert "A - " in legend  # the reviewer still sees what was tested
    assert not items[0].evidence_images  # but no box is drawn in the wrong place
    assert any("no distinctive value" in w for w in warnings)


def test_a_failed_sample_is_never_marked_satisfied(tmp_path: Path):
    # The one failure mode worth being paranoid about: a workpaper that
    # silently reads "satisfied" when the item failed.
    (tmp_path / "invoice.pdf").write_bytes(b"%PDF-1.4\n")
    conclusion = _conclusion(
        conclusion="not_satisfied",
        sample_results=[{"sample_id": "1", "conclusion": "not_satisfied", "note": "no approval"}],
    )
    items, _ = sample_items_from_conclusion(conclusion, _manifest(), "C-14", tmp_path)
    assert items[0].conclusion_satisfied is False


def test_raw_data_uses_sample_key_fields_in_column_order(tmp_path: Path):
    (tmp_path / "invoice.pdf").write_bytes(b"%PDF-1.4\n")
    items, _ = sample_items_from_conclusion(_conclusion(), _manifest(), "C-14", tmp_path)
    assert items[0].raw_data.values == {"A": "21022452", "B": "Premiere Onboard"}


def test_build_workpaper_request_round_trips(tmp_path: Path):
    (tmp_path / "invoice.pdf").write_bytes(b"%PDF-1.4\n")
    request, warnings = build_workpaper_request(
        _conclusion(),
        _manifest(),
        control_id="C-14",
        template_path=tmp_path / "PY.xlsx",
        output_path=tmp_path / "out.xlsx",
        support_dir=tmp_path,
        population_rows=[{"invoice_number": "21022452", "vendor": "Premiere Onboard"}],
    )
    assert request.sample_items[0].control_id == "C-14"
    assert request.population_rows[0].values == {"A": "21022452", "B": "Premiere Onboard"}
    assert not warnings


# --------------------------------------------------------------------------
# End to end: a conclusion the agent could actually emit, through the bridge,
# into real native-shape XML positioned by real OCR. This is the only test
# that proves the two halves fit; everything above tests one side of the seam.
# --------------------------------------------------------------------------


def _invoice_pdf(path: Path) -> None:
    from reportlab.lib.pagesizes import letter as PAGE
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(str(path), pagesize=PAGE)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(72, 730, "INVOICE")
    c.setFont("Helvetica", 12)
    for y, line in [
        (690, "Invoice Number: 21022452"),
        (670, "Vendor: Premiere Onboard LLC"),
        (650, "Amount: 30,000.00"),
        (630, "Payment Date: 11/10/2025"),
    ]:
        c.drawString(72, y, line)
    c.save()


def test_a_conclusion_becomes_native_callouts_on_the_real_page(tmp_path: Path):
    import re

    from agent.ooxml import drawing_builder as db
    from agent.ooxml.ooxml_utils import ColumnLayout

    _invoice_pdf(tmp_path / "invoice.pdf")

    conclusion = _conclusion(
        attribute_results=[
            {
                "attribute": "Invoice amount agrees to payment",
                "sample_id": "1",
                "result": "satisfied",
                "value_observed": "Invoice total 30,000.00",
                "evidence_ids": ["ev_0001"],
            },
            {
                "attribute": "Payment date follows approval",
                "sample_id": "1",
                "result": "satisfied",
                "value_observed": "Paid 11/10/2025",
                "evidence_ids": ["ev_0001"],
            },
        ]
    )
    conclusion.evidence_citations[0].location = "invoice.pdf p.1"

    items, warnings = sample_items_from_conclusion(conclusion, _manifest(), "C-14", tmp_path)
    assert not warnings

    result = db.build_sample_drawing(items, ColumnLayout(widths_px={"A": 100, "B": 150}, default_px=64))

    # Every tickmark the bridge asked for was actually located on the page.
    assert result.unplaced_callouts == []
    # Native shapes, not pictures: two callouts plus the narrative textbox.
    assert len(re.findall(r"<xdr:sp[ >]", result.xml)) == 3
    assert len(re.findall(r"<xdr:pic[ >]", result.xml)) == 1
    assert sorted(set(re.findall(r"<a:t>([A-Z])</a:t>", result.xml))) == ["A", "B"]

    # The two boxes landed in DIFFERENT places -- a bridge that anchored both
    # attributes on the same token would place them identically and look
    # convincing while being useless.
    positions = re.findall(r'<xdr:pos x="(\d+)" y="(\d+)"/>', result.xml)
    callout_positions = set(positions[1:3])
    assert len(callout_positions) == 2


# --------------------------------------------------------------------------
# The summary tab. Its whole job is to be SHORT -- a summary a reviewer has
# to read is not a summary -- so most of these assert on what it leaves out.
# --------------------------------------------------------------------------


def test_summary_never_carries_the_narrative():
    # The narrative is multi-paragraph by design and belongs in the working
    # papers underneath. Pasting it into a summary cell is exactly the "too
    # much information" failure this tab exists to fix.
    from agent.render_bridge import summary_rows

    conclusion = _conclusion(
        narrative=(
            "IA performed a detailed inspection of the supporting documentation and "
            "noted that the invoice was routed for approval on the 26th, approved on "
            "the 1st, batched on the 7th and settled on the 10th, all of which is "
            "consistent with the control description and the prior year approach."
        )
    )
    rows = summary_rows([("1. Verify approval.", conclusion)])
    cells = " ".join(rows[0].attributes + rows[0].evidenced)
    assert "detailed inspection" not in cells
    assert all(len(c) <= 300 for c in rows[0].evidenced)


def test_ipe_always_gets_its_own_row():
    # A control can have every test step satisfied and still fail on IPE.
    # Folded into a step it is invisible to a reviewer scanning for it.
    from agent.render_bridge import summary_rows

    rows = summary_rows([("1. Verify approval.", _conclusion())])
    assert len(rows) == 2
    assert rows[-1].test_step.startswith("IPE")


def test_unvalidated_ipe_reads_as_an_exception_not_a_pass():
    from agent.render_bridge import summary_rows

    rows = summary_rows(
        [("1.", _conclusion(ipe_completeness_accuracy_status="not_validated"))]
    )
    assert rows[-1].verdict_colour == "red"
    assert "not validated" in rows[-1].result.lower()


def test_an_exception_names_which_sample_failed():
    # "Not satisfied" alone sends a reviewer hunting through every sample
    # sheet to find which one.
    from agent.render_bridge import summary_rows

    conclusion = _conclusion(
        conclusion="not_satisfied",
        sample_results=[
            {"sample_id": "1", "conclusion": "satisfied"},
            {"sample_id": "2", "conclusion": "not_satisfied"},
        ],
    )
    rows = summary_rows([("1.", conclusion)])
    assert "2" in rows[0].result
    assert rows[0].verdict_colour == "red"


def test_an_unrecognised_verdict_is_never_shown_as_a_pass():
    # Green on something we did not understand is the one failure mode that
    # actively misleads. Anything unknown must read as needing attention.
    from agent.ooxml.summary_sheet import SummaryRow

    assert SummaryRow(test_step="x", result="qualified, see memo").verdict_colour == "amber"
    assert SummaryRow(test_step="x", result="").verdict_colour == "amber"
    assert SummaryRow(test_step="x", result="Satisfied").verdict_colour == "green"


def test_every_summary_row_links_somewhere(tmp_path: Path):
    from agent.render_bridge import summary_rows

    for row in summary_rows([("1.", _conclusion())]):
        assert row.link_to and "!" in row.link_to


# --------------------------------------------------------------------------
# Raw-data-row callouts, and surviving a template that isn't this one.
# There are 172 controls; this package was built from exactly one of them.
# --------------------------------------------------------------------------


def test_the_extract_row_gets_boxed_too(tmp_path: Path):
    # Without this a reviewer sees a red box on an invoice and has to take
    # on faith that it is the row that was selected.
    (tmp_path / "invoice.pdf").write_bytes(b"%PDF-1.4\n")
    # The attribute has to observe a value that IS on the extract row --
    # here the invoice number, which is the first key field and so column A.
    # An amount the extract does not carry correctly boxes nothing.
    conclusion = _conclusion(
        attribute_results=[
            {
                "attribute": "Invoice number agrees to the sample listing",
                "sample_id": "1",
                "result": "satisfied",
                "value_observed": "Invoice 21022452",
                "evidence_ids": ["ev_0001"],
            }
        ]
    )
    items, _ = sample_items_from_conclusion(conclusion, _manifest(), "C-14", tmp_path)
    tie_outs = items[0].raw_data_tie_outs
    assert [t.letter for t in tie_outs] == ["A"]
    assert tie_outs[0].columns == ["A"]


def test_raw_columns_are_matched_on_value_not_column_name():
    # 172 controls means 172 extracts that name their columns differently.
    # Matching the value needs no per-control configuration at all.
    from agent.render_bridge import _raw_columns_for

    fields = {"Invoice Number F0411.VINV": "35713082", "Total Payment Amount": "32677.00"}
    assert _raw_columns_for(["35713082"], fields) == ["A"]
    assert _raw_columns_for(["32,677.00"], fields) == ["B"]  # formatting differs, still matches


def test_a_value_that_is_not_on_the_extract_row_boxes_nothing():
    # An attribute about an approver's identity has nothing to point at in
    # a payment extract. No box is the right answer, not a guess.
    from agent.render_bridge import _raw_columns_for

    assert _raw_columns_for(["Milosevic"], {"Payee Number": "27677628"}) == []


def test_raw_column_match_is_exact_not_substring():
    # "2859" is contained in a dozen unrelated numbers on a real extract
    # row, and a box over the wrong column is worse than no box.
    from agent.render_bridge import _raw_columns_for

    assert _raw_columns_for(["2859"], {"Document Number": "11928590"}) == []


def test_a_template_missing_a_required_tab_is_refused(tmp_path: Path):
    # Silently writing the sample rows into whatever tab happens to be
    # second is the failure this prevents. Across 172 controls a refusal is
    # a two-minute fix; a wrong workpaper is a finding.
    import openpyxl

    from agent.ooxml import ooxml_utils as ox

    wb = openpyxl.Workbook()
    wb.active.title = "NotTheSampleTab"
    src = tmp_path / "odd_template.xlsx"
    wb.save(src)

    work = str(tmp_path / "wd")
    ox.extract_template(str(src), work)
    with pytest.raises(ox.TemplateMismatch, match="no sheet named"):
        ox.check_template(work, required_sheets=["Sample", "Population"], max_style_index=0)


def test_a_template_without_the_expected_styles_is_refused(tmp_path: Path):
    # The STYLE_* constants were lifted from one specific styles.xml. On a
    # template with fewer formats they address past the end of cellXfs,
    # which makes Excel show a repair prompt naming nothing useful.
    import openpyxl

    from agent.ooxml import ooxml_utils as ox

    wb = openpyxl.Workbook()
    wb.active.title = "Sample"
    wb.create_sheet("Population")
    src = tmp_path / "plain.xlsx"
    wb.save(src)

    work = str(tmp_path / "wd2")
    ox.extract_template(str(src), work)
    with pytest.raises(ox.TemplateMismatch, match="cell formats"):
        ox.check_template(work, required_sheets=["Sample", "Population"], max_style_index=99)


def test_sheets_are_resolved_by_name_not_by_position(tmp_path: Path):
    # sheetN.xml numbering does not follow tab order -- in the real PY
    # template "IA Leadsheet" carries sheetId=4 while living in sheet1.xml.
    import openpyxl

    from agent.ooxml import ooxml_utils as ox

    wb = openpyxl.Workbook()
    wb.active.title = "First"
    wb.create_sheet("Sample")
    wb.create_sheet("Population")
    src = tmp_path / "ordered.xlsx"
    wb.save(src)

    work = str(tmp_path / "wd3")
    ox.extract_template(str(src), work)
    sheets = ox.resolve_sheets(work)
    assert set(sheets) == {"First", "Sample", "Population"}
    assert sheets["Sample"] != sheets["Population"]
