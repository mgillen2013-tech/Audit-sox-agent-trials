"""Extraction tests against synthetic files -- the "prove it against a real
file" check for the biggest unknown in the design (garbage extraction
poisons everything downstream).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from PIL import Image
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.pdfgen import canvas

from agent.extraction import extract_excel, extract_many, extract_pdf


# --------------------------------------------------------------------------
# Fixtures: build synthetic files rather than checking binaries into the repo
# --------------------------------------------------------------------------


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO Testing"

    header = ["Sample", "PO Number", "Branch", "Amount"]
    rows = [
        ["S01", "PO-48213", "35210", 12450],
        ["S02", "PO-48214", "35211", 8300],
        ["S03", "PO-48215", "35210", 15020],
    ]
    for r, values in enumerate([header, *rows], start=1):
        for c, value in enumerate(values, start=1):
            ws.cell(row=r, column=c, value=value)

    # Exception: preparer highlighted a cell and left a comment on it.
    flagged_cell = ws.cell(row=3, column=4)  # S02 amount
    flagged_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    flagged_cell.comment = Comment("Approval missing supervisor signature", "Preparer")

    path = tmp_path / "po_sample.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def native_text_pdf(tmp_path: Path) -> Path:
    path = tmp_path / "gl_export.pdf"
    doc = SimpleDocTemplate(str(path), pagesize=letter)
    data = [
        ["Account", "Description", "Amount"],
        ["6100", "Accrued Interest", "482110"],
        ["6200", "AP Trade", "119400"],
    ]
    table = Table(data)
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ]
        )
    )
    doc.build([table])
    return path


@pytest.fixture
def scanned_pdf(tmp_path: Path) -> Path:
    # A page with only a rasterized image and no text layer -- the case
    # extract_pdf should flag as needing OCR rather than silently dropping.
    img_path = tmp_path / "screenshot.png"
    Image.new("RGB", (400, 200), color="white").save(img_path)

    path = tmp_path / "screenshot_as_pdf.pdf"
    c = canvas.Canvas(str(path), pagesize=letter)
    c.drawImage(str(img_path), 50, 500, width=400, height=200)
    c.showPage()
    c.save()
    return path


# --------------------------------------------------------------------------
# Excel
# --------------------------------------------------------------------------


def test_excel_table_extracted_with_header_and_rows(sample_xlsx: Path):
    items = extract_excel(sample_xlsx)
    tables = [i for i in items if i.source_type == "excel_table"]

    assert len(tables) == 1
    table = tables[0].extracted_table
    assert table[0] == ["Sample", "PO Number", "Branch", "Amount"]
    assert table[1] == ["S01", "PO-48213", "35210", "12450"]
    assert tables[0].location.startswith("PO Testing!A1:")
    assert tables[0].extraction_confidence == 1.0


def test_excel_flagged_cell_captured_as_evidence(sample_xlsx: Path):
    items = extract_excel(sample_xlsx)
    cells = [i for i in items if i.source_type == "excel_cell"]

    assert len(cells) == 1
    cell = cells[0]
    assert cell.location == "PO Testing!D3"
    assert "FFFF00" in cell.extracted_text
    assert "Approval missing supervisor signature" in cell.extracted_text


def test_excel_all_ids_unique(sample_xlsx: Path):
    items = extract_excel(sample_xlsx)
    ids = [i.evidence_id for i in items]
    assert len(ids) == len(set(ids))


def test_excel_large_table_is_chunked_not_one_giant_blob(tmp_path: Path):
    # A real population export can be thousands of unbroken rows -- this is
    # what would otherwise become a single EvidenceItem that then gets
    # rendered in full into every prompt turn with no size cap of its own.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Invoice", "Amount"])
    for i in range(450):
        ws.append([f"INV-{i}", i])
    path = tmp_path / "big_population.xlsx"
    wb.save(path)

    items = extract_excel(path)
    tables = [i for i in items if i.source_type == "excel_table"]

    assert len(tables) == 3  # 450 data rows / 200 per chunk -> 3 chunks
    for t in tables:
        assert len(t.extracted_table) <= 201  # header + up to 200 data rows
        assert t.extracted_table[0] == ["Invoice", "Amount"]  # header repeated in every chunk

    # No row lost or duplicated across chunks.
    all_invoices = [row[0] for t in tables for row in t.extracted_table[1:]]
    assert all_invoices == [f"INV-{i}" for i in range(450)]


# --------------------------------------------------------------------------
# PDF
# --------------------------------------------------------------------------


def test_native_pdf_text_and_table_extracted(native_text_pdf: Path):
    items = extract_pdf(native_text_pdf)

    text_items = [i for i in items if i.source_type == "pdf_text"]
    table_items = [i for i in items if i.source_type == "pdf_table"]

    assert len(text_items) == 1
    assert "Accrued Interest" in text_items[0].extracted_text
    assert text_items[0].extraction_confidence == 1.0

    assert len(table_items) == 1
    table = table_items[0].extracted_table
    assert table[0] == ["Account", "Description", "Amount"]
    assert ["6100", "Accrued Interest", "482110"] in table


def test_scanned_pdf_flagged_not_dropped(scanned_pdf: Path):
    items = extract_pdf(scanned_pdf)

    assert len(items) == 1
    item = items[0]
    assert item.source_type == "image_ocr"
    assert item.extraction_confidence == 0.0
    assert item.extracted_text is None
    # The page still shows up with a location -- nothing silently vanished.
    assert "p.1" in item.location


# --------------------------------------------------------------------------
# Multi-file merge
# --------------------------------------------------------------------------


def test_extract_many_renumbers_ids_without_collision(sample_xlsx: Path, native_text_pdf: Path):
    # Each single-file extractor numbers ev_0001.. independently -- naively
    # concatenating extract(a) + extract(b) would collide. extract_many must not.
    solo_excel = extract_excel(sample_xlsx)
    solo_pdf = extract_pdf(native_text_pdf)
    assert solo_excel[0].evidence_id == solo_pdf[0].evidence_id == "ev_0001"  # confirms the collision exists

    merged = extract_many([sample_xlsx, native_text_pdf])

    assert len(merged) == len(solo_excel) + len(solo_pdf)
    ids = [item.evidence_id for item in merged]
    assert len(ids) == len(set(ids))

    # Original file/location/content is preserved -- only evidence_id changed.
    assert merged[0].source_file == "po_sample.xlsx"
    assert merged[len(solo_excel)].source_file == "gl_export.pdf"
