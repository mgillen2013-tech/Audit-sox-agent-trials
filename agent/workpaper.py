"""CY workpaper file generation -- the draft deliverable a reviewer opens.

Takes the structured results a control run already produced (ConclusionOutput
per test step, or the preserved failure info for a step that aborted) and
writes one workpaper file per control. Excel by default -- SOX testing lives
in workbooks, and only the xlsx output carries the full structure (a
filterable summary, one sheet per test step, an exhibits sheet per step).
PDF is available via fmt="pdf" as a flat read-only rendering, but is no
longer inferred from the PY file's type: matching PY meant a PDF precedent
silently downgraded this year's deliverable. This is a
CLEAN generated document with standard workpaper sections, not a cell-level
edit of the PY file -- per the design decision that a predictable layout
beats a fragile in-place edit of an arbitrary workbook.

Deliberately zero LLM calls: everything here is already in the conclusion
JSON. Generating the file costs nothing and is fully deterministic, so it
can run after every control run without touching the spending cap.

Every page/sheet is stamped DRAFT -- this output is a draft for human
review, never a finalized workpaper (same non-negotiable as the system
prompt).
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from agent.schemas import ConclusionOutput, EvidenceCitation, EvidenceItem

_DRAFT_BANNER = "DRAFT -- AI-prepared, pending human reviewer approval. Not a finalized workpaper."

_CONCLUSION_LABELS = {
    "satisfied": "Satisfied",
    "not_satisfied": "Not satisfied",
    "insufficient_evidence": "Insufficient evidence",
}


# ── Evidence exhibits (tickmark annotation) ────────────────────────────────
# Mirrors how a human workpaper points at evidence: the citation table gets a
# tickmark letter per citation (A, B, C...), and the cited source is embedded
# as an exhibit -- a rendered PDF page with red boxes drawn where the quoted
# text was found (tight boxes via text search; falls back to the extracted
# region's bbox; a scanned/screenshot page with no text layer becomes a
# full-page exhibit labeled with its letter). Excel-sourced citations get a
# text excerpt of the cited range instead of an image. All deterministic --
# zero LLM calls -- and every step degrades gracefully: if an exhibit can't
# be rendered, the citation row still stands, just without a picture.

_PDF_RENDER_DPI = 110
_MAX_EXHIBIT_WIDTH_PX = 640
_MAX_EXCERPT_ROWS = 12

_BBOX_RE = re.compile(r"p\.(\d+) \(bbox ([\d.\-]+),([\d.\-]+),([\d.\-]+),([\d.\-]+)\)")


def _tickmark_letters(n: int) -> list[str]:
    return [chr(65 + i) if i < 26 else str(i + 1) for i in range(n)]


def _clean_quote_for_search(quote: str) -> str | None:
    # The model's quote_or_summary may be a paraphrase or stitch pieces with
    # ellipses -- search for the longest literal-looking fragment.
    pieces = re.split(r"\.\.\.|…|\n", quote)
    best = max((re.sub(r"\s+", " ", p).strip() for p in pieces), key=len, default="")
    if len(best) < 12:
        return None
    return best[:60]


def _search_rects(page, quote: str) -> list[tuple[float, float, float, float]]:
    chunk = _clean_quote_for_search(quote)
    if not chunk:
        return []
    try:
        matches = page.search(chunk, regex=False, case=False)
    except Exception:  # noqa: BLE001 -- search is best-effort tightening only
        return []
    return [(m["x0"], m["top"], m["x1"], m["bottom"]) for m in matches[:3]]


def _item_rect(page, item: EvidenceItem) -> list[tuple[float, float, float, float]]:
    m = _BBOX_RE.search(item.location)
    if not m:
        return []
    x0, top, x1, bottom = (float(v) for v in m.groups()[1:])
    # A bbox covering (almost) the whole page isn't a pointer, it's the page
    # -- drawing a box around everything marks nothing.
    if (x1 - x0) * (bottom - top) >= 0.9 * page.width * page.height:
        return []
    return [(x0, top, x1, bottom)]


# Both the text-layer search and local OCR return a box hugging the glyphs
# exactly. Drawn as-is, the stroke lands ON the characters and makes the
# boxed value harder to read than the surrounding text -- the opposite of
# what a tickmark is for. Pad proportionally to the line height so small
# print and large headings both get sensible room, with a floor for tiny text.
# Asymmetric on purpose: there is usually whitespace to the left and right
# of a value, but the next line of text sits close above and below it. Equal
# padding wide enough to look right horizontally put the box's bottom edge
# through the following line ("Invoice no.: 2859" struck through "Terms: Net
# 45" on a real invoice), so vertical padding is roughly half the horizontal.
_BOX_PAD_X_FRACTION = 0.40
_BOX_PAD_Y_FRACTION = 0.18
_BOX_PAD_X_MIN_PT = 2.5
_BOX_PAD_Y_MIN_PT = 1.5


def _pad_rect(
    rect: tuple[float, float, float, float], page_width: float, page_height: float
) -> tuple[float, float, float, float]:
    x0, top, x1, bottom = rect
    height = bottom - top
    pad_x = max(_BOX_PAD_X_MIN_PT, height * _BOX_PAD_X_FRACTION)
    pad_y = max(_BOX_PAD_Y_MIN_PT, height * _BOX_PAD_Y_FRACTION)
    return (
        max(x0 - pad_x, 0.0),
        max(top - pad_y, 0.0),
        min(x1 + pad_x, page_width),
        min(bottom + pad_y, page_height),
    )


# Overlays are placed as SEPARATE pictures over a clean page render rather
# than burned into its pixels. Burned-in marks are final: if the agent puts
# a box slightly off, or the tickmark letter needs to sit elsewhere, a
# reviewer has no recourse short of redoing the exhibit by hand. As
# individual pictures, each box and each letter can be dragged, resized or
# deleted in Excel like any other shape -- which is what a preparer does to
# hand-drawn tickmarks anyway.
_LETTER_BADGE_PX = 24
_BOX_STROKE_PX = 3


@dataclass
class _Overlay:
    """One draggable mark, positioned in the final (post-downscale) pixel
    space of the page image it belongs to.
    """

    png: BytesIO
    x: int
    y: int
    w: int
    h: int
    name: str


def _box_overlay_png(w: int, h: int) -> BytesIO:
    from PIL import Image as PILImage, ImageDraw

    img = PILImage.new("RGBA", (max(w, 1), max(h, 1)), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)
    draw.rectangle([0, 0, max(w, 1) - 1, max(h, 1) - 1], outline=(255, 0, 0, 255), width=_BOX_STROKE_PX)
    buf = BytesIO()
    img.save(buf, "PNG")
    buf.seek(0)
    return buf


def _letter_overlay_png(letter: str) -> BytesIO:
    from PIL import Image as PILImage, ImageDraw, ImageFont

    size = _LETTER_BADGE_PX
    img = PILImage.new("RGBA", (size, size), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)
    draw.rectangle([0, 0, size - 1, size - 1], fill=(255, 0, 0, 255))
    try:
        font = ImageFont.load_default(size=18)
    except TypeError:  # older Pillow: no size kwarg
        font = ImageFont.load_default()
    draw.text((6, 2), letter, fill="white", font=font)
    buf = BytesIO()
    img.save(buf, "PNG")
    buf.seek(0)
    return buf


def _flatten(page_png: BytesIO, overlays: list[_Overlay]) -> BytesIO:
    """Burns the overlays onto the page -- for the PDF output only, where
    nothing can be movable anyway.
    """
    from PIL import Image as PILImage

    page_png.seek(0)
    page = PILImage.open(page_png).convert("RGBA")
    for ov in overlays:
        ov.png.seek(0)
        mark = PILImage.open(ov.png).convert("RGBA")
        page.alpha_composite(mark, (ov.x, ov.y))
    out = BytesIO()
    page.convert("RGB").save(out, "PNG")
    out.seek(0)
    return out


def _place_overlay(ws, overlay: _Overlay, anchor_row: int) -> None:
    """Pins one overlay over the page image using a pixel offset from the
    same cell the page is anchored to, so it lands exactly on the value it
    marks -- and stays independently movable.
    """
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU

    img = XLImage(overlay.png)
    img.anchor = OneCellAnchor(
        _from=AnchorMarker(
            col=0,
            row=anchor_row - 1,  # AnchorMarker rows are 0-based
            colOff=pixels_to_EMU(overlay.x),
            rowOff=pixels_to_EMU(overlay.y),
        ),
        ext=XDRPositiveSize2D(pixels_to_EMU(overlay.w), pixels_to_EMU(overlay.h)),
    )
    ws.add_image(img)


def _render_pdf_exhibit(
    pdf_path: Path, page_num: int, marks: list[tuple[str, EvidenceItem, str]]
) -> tuple[BytesIO, tuple[int, int], list[_Overlay]]:
    """marks: (tickmark letter, evidence item, quote to locate).

    Returns the CLEAN page render, its pixel size, and the marks as
    separately-positioned overlays. Nothing is drawn onto the page itself:
    the caller places each overlay as its own picture so a reviewer can
    move or delete a misplaced box without touching the page image.
    """
    import pdfplumber

    from agent.wordboxes import find_text_boxes, ocr_line_boxes

    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[page_num - 1]
        pim = page.to_image(resolution=_PDF_RENDER_DPI)
        scale = _PDF_RENDER_DPI / 72.0
        # OCR of this page, read at most ONCE and reused by every mark on
        # it. OCR (~1.3s) dwarfs rendering (~0.06s), so doing it per mark
        # made a 3-citation page three times slower for identical output.
        ocr_lines: list | None = None

        # (letter, [padded rects in PDF points])
        located: list[tuple[str, list[tuple[float, float, float, float]]]] = []
        for letter, item, quote in marks:
            rects = _search_rects(page, quote) or _item_rect(page, item)

            if not rects:
                # No text layer to search (a scanned page / E1 screenshot).
                # Local OCR measures where the quoted values actually sit;
                # without this the letter just parks in the corner pointing
                # at nothing. Pixel boxes convert back to PDF points so
                # everything is positioned through the same path below.
                if ocr_lines is None:
                    ocr_lines = ocr_line_boxes(pim.original.convert("RGB"))
                rects = [
                    (x0 / scale, y0 / scale, x1 / scale, y1 / scale)
                    for x0, y0, x1, y1 in find_text_boxes(ocr_lines, quote)
                ]

            located.append((letter, [_pad_rect(r, page.width, page.height) for r in rects]))

        pil = pim.original.convert("RGB")

    ratio = 1.0
    if pil.width > _MAX_EXHIBIT_WIDTH_PX:
        ratio = _MAX_EXHIBIT_WIDTH_PX / pil.width
        pil = pil.resize((_MAX_EXHIBIT_WIDTH_PX, int(pil.height * ratio)))

    # PDF points -> rendered pixels -> final (downscaled) pixels.
    px = scale * ratio
    overlays: list[_Overlay] = []
    unanchored = 0
    for letter, rects in located:
        for i, (x0, top, x1, bottom) in enumerate(rects):
            bx, by = int(x0 * px), int(top * px)
            bw, bh = max(int((x1 - x0) * px), 1), max(int((bottom - top) * px), 1)
            overlays.append(_Overlay(_box_overlay_png(bw, bh), bx, by, bw, bh, f"box {letter}{i or ''}"))

        if rects:
            x0, top = rects[0][0], rects[0][1]
            lx = max(int(x0 * px) - _LETTER_BADGE_PX - 4, 0)
            ly = max(int(top * px) - 4, 0)
        else:
            # Nothing locatable on the page (an unreadable scan): the whole
            # page is the exhibit, so stack the letters in the corner.
            lx, ly = 6 + unanchored * (_LETTER_BADGE_PX + 6), 6
            unanchored += 1
        overlays.append(
            _Overlay(
                _letter_overlay_png(letter), lx, ly, _LETTER_BADGE_PX, _LETTER_BADGE_PX, f"tickmark {letter}"
            )
        )

    buf = BytesIO()
    pil.save(buf, "PNG")
    buf.seek(0)
    return buf, (pil.width, pil.height), overlays


def _excerpt_lines(item: EvidenceItem) -> list[str]:
    if item.extracted_table:
        rows = item.extracted_table[:_MAX_EXCERPT_ROWS]
        lines = [" | ".join(row) for row in rows]
        if len(item.extracted_table) > _MAX_EXCERPT_ROWS:
            lines.append(f"... ({len(item.extracted_table) - _MAX_EXCERPT_ROWS} more row(s) in the source)")
        return lines
    if item.extracted_text:
        return item.extracted_text.splitlines()[:_MAX_EXCERPT_ROWS]
    return []


def _build_step_exhibits(
    citations: list[EvidenceCitation],
    evidence_map: dict[str, EvidenceItem],
    support_dir: Path,
) -> tuple[list[str], list[tuple[str, BytesIO, tuple[int, int], list[_Overlay]]], list[tuple[str, list[str]]]]:
    """Returns (tickmark letter per citation, PDF exhibit images as
    (caption, png bytes, (w,h)), Excel/text excerpts as (caption, lines)).
    """
    letters = _tickmark_letters(len(citations))

    pdf_groups: dict[tuple[str, int], list[tuple[str, EvidenceItem, str]]] = {}
    excerpts: list[tuple[str, list[str]]] = []
    for letter, cit in zip(letters, citations):
        item = evidence_map.get(cit.evidence_id)
        if item is None:
            continue
        m = _BBOX_RE.search(item.location)
        if item.source_type in ("pdf_text", "pdf_table", "image_ocr") and m:
            key = (item.source_file, int(m.group(1)))
            pdf_groups.setdefault(key, []).append((letter, item, cit.quote_or_summary))
        else:
            lines = _excerpt_lines(item)
            if lines:
                excerpts.append((f"Exhibit {letter} — {item.location}", lines))

    images: list[tuple[str, BytesIO, tuple[int, int], list[_Overlay]]] = []
    for (source_file, page_num), marks in pdf_groups.items():
        try:
            buf, size, overlays = _render_pdf_exhibit(support_dir / source_file, page_num, marks)
        except Exception:  # noqa: BLE001 -- an unrenderable page must not sink the workpaper
            continue
        mark_list = ", ".join(letter for letter, _, _ in marks)
        images.append((f"Exhibit ({mark_list}) — {source_file} p.{page_num}", buf, size, overlays))

    return letters, images, excerpts


def workpaper_path_for(py_testing_filename: str, control_id: str, out_dir: Path) -> Path:
    """Excel by default -- SOX testing lives in workbooks, and only the xlsx
    output carries the full structure (per-step sheets, an exhibits sheet,
    a filterable summary). PDF is produced only when explicitly asked for
    via `fmt`, not inferred from the PY file's type: matching PY meant a
    PDF precedent silently downgraded this year's deliverable to the
    flat-page rendering.
    """
    safe_control = re.sub(r"[^\w.-]+", "_", control_id) or "control"
    return out_dir / f"{safe_control}_CY_Testing_DRAFT.xlsx"


def build_workpaper(
    spec: dict[str, Any],
    results: dict[str, dict],
    py_testing_filename: str,
    out_dir: Path,
    support_dir: Path | None = None,
    fmt: str = "xlsx",
) -> Path:
    """Writes the control's CY workpaper and returns the written path.

    fmt: "xlsx" (default) or "pdf". Excel is the default deliverable and the
    only one carrying the full structure; PDF is a flat read-only rendering.

    spec: the same control spec dict run_control uses (control_id,
    control_objective_ref, control_objective_text, test_steps).
    results: {test_step_id: {"conclusion", "audit_log"} | {"error", ...}} --
    exactly what iter_control_results yielded. Failed steps are documented
    in the workpaper as incomplete rather than silently omitted: a reviewer
    needs to see that a step has no conclusion yet, not a file that looks
    finished with a step missing.

    support_dir: the directory still holding spec["cy_support_files"] (the
    run's staging dir). When given, citations get tickmark letters and the
    cited sources are embedded as annotated exhibits. Extraction is
    deterministic, so re-extracting here reproduces the exact evidence_ids
    the run used -- same files, same order, same ids. When omitted (or if
    re-extraction fails) the workpaper is text-only, never broken.
    """
    out_path = workpaper_path_for(py_testing_filename, spec["control_id"], out_dir)
    if fmt == "pdf":
        out_path = out_path.with_suffix(".pdf")
    step_texts = {s["test_step_id"]: s.get("test_step_text", "") for s in spec.get("test_steps", [])}

    evidence_map: dict[str, EvidenceItem] = {}
    if support_dir is not None:
        try:
            from agent.extraction import extract_many

            items = extract_many([support_dir / f for f in spec.get("cy_support_files", [])])
            evidence_map = {item.evidence_id: item for item in items}
        except Exception:  # noqa: BLE001 -- exhibits are an enhancement, not a precondition
            evidence_map = {}

    if out_path.suffix == ".pdf":
        _build_pdf(spec, results, step_texts, out_path, evidence_map, support_dir)
    else:
        _build_xlsx(spec, results, step_texts, out_path, evidence_map, support_dir)
    return out_path


# --------------------------------------------------------------------------
# Excel
# --------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")


def _sheet_title(test_step_id: str) -> str:
    # Excel forbids : \ / ? * [ ] in sheet names and caps them at 31 chars.
    return re.sub(r"[:\\/?*\[\]]", "_", test_step_id)[:31] or "step"


def _sample_sheet_name(test_step_id: str, sample_id: str, single_step: bool) -> str:
    """Sample sheets are named for the item alone on a single-step control
    (tabs "1", "2"). On a multi-step control they must carry the step too:
    every step has its own sample 1, and bare names collided into "1" and
    "1(2)", which says nothing about which step it belongs to.
    """
    return sample_id if single_step else f"{test_step_id} {sample_id}"


def _exhibit_sheet_title(test_step_id: str) -> str:
    # Same sanitizing, but leave room for the " - Exhibits" suffix within
    # Excel's 31-char sheet-name limit.
    base = re.sub(r"[:\\/?*\[\]]", "_", test_step_id)[:20] or "step"
    return f"{base} - Exhibits"


_SECTION_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)
_CONCLUSION_COLORS = {
    "satisfied": "006100",
    "not_satisfied": "9C0006",
    "insufficient_evidence": "9C5700",
}


def _build_xlsx(
    spec: dict[str, Any],
    results: dict[str, dict],
    step_texts: dict[str, str],
    out_path: Path,
    evidence_map: dict[str, EvidenceItem],
    support_dir: Path | None,
) -> None:
    wb = openpyxl.Workbook()
    # Exhibit PNG buffers must stay alive until wb.save() -- openpyxl reads
    # image data at save time, not at add_image time.
    _live_buffers: list[BytesIO] = []
    wb.remove(wb.active)  # sheets are created explicitly below

    single_step = len(results) == 1
    first_sheet = True

    for test_step_id, result in results.items():
        citations = [] if "error" in result else list(result["conclusion"].evidence_citations)
        step_level, groups = _sample_groups(citations)
        per_sample = bool(groups)

        # Sheet 1 for this step: everything a reviewer reads top-to-bottom
        # -- control header, conclusion, exceptions, open requests,
        # documentation, procedures. The old layout split this across a
        # Summary tab and a step tab, so the header lived somewhere you had
        # to click away from to read anything.
        summary_name = "Summary" if single_step else f"{test_step_id} Summary"
        summary_ws = wb.create_sheet(_unique_sheet_name(wb, summary_name))
        letters, images, excerpts = _write_step_sheet(
            summary_ws,
            test_step_id,
            step_texts.get(test_step_id, ""),
            result,
            evidence_map,
            support_dir,
            # Step-wide evidence (population extract, IPE parameters) belongs
            # here, alongside the IPE attribute it supports. When nothing is
            # sample-tagged there is nowhere else for evidence to go, so it
            # all stays here rather than vanishing.
            citations=step_level if per_sample else citations,
            spec=spec if first_sheet else None,
            results=results if first_sheet else None,
            include_step_detail=True,
        )
        first_sheet = False
        if images or excerpts:
            ex_ws = wb.create_sheet(_unique_sheet_name(wb, f"{summary_ws.title} - Exhibits"))
            _write_exhibit_sheet(ex_ws, summary_ws.title, images, excerpts, _live_buffers)

        # Then one sheet per sampled item: just that item's evidence, with
        # tickmarks restarting at A, plus its own exhibits.
        if per_sample:
            for sample_id, group in groups:
                name = _unique_sheet_name(wb, _sample_sheet_name(test_step_id, sample_id, single_step))
                ws = wb.create_sheet(name)
                letters, images, excerpts = _write_step_sheet(
                    ws,
                    test_step_id,
                    step_texts.get(test_step_id, ""),
                    result,
                    evidence_map,
                    support_dir,
                    citations=group,
                    sample_id=sample_id,
                    include_step_detail=False,
                )
                if images or excerpts:
                    ex_ws = wb.create_sheet(_unique_sheet_name(wb, f"{name} - Exhibits"))
                    _write_exhibit_sheet(ex_ws, name, images, excerpts, _live_buffers)

    wb.save(out_path)


def _sample_groups(
    citations: list[EvidenceCitation],
) -> tuple[list[EvidenceCitation], list[tuple[str, list[EvidenceCitation]]]]:
    """Splits a step's citations into step-wide evidence and one group per
    sampled item, so each selection gets its own sheet, its own exhibits,
    and tickmarks that restart at A -- a reviewer clearing selection 2
    wants selection 2's evidence, not one merged list where its first
    tickmark happens to be F.

    Step-wide citations (the population extract, IPE parameters, a policy)
    stay on the summary sheet. They used to ride along on the FIRST item's
    sheet, which gave that item tickmarks its own attributes never
    referenced -- a real workpaper had sample 1 carrying B and C for the
    population and report parameters while its attributes cited only A and
    D-G. Evidence about the population is not evidence about selection 1.

    Returns ([], []) when nothing is sample-tagged, so the caller keeps
    everything on one sheet as before.
    """
    by_sample: dict[str, list[EvidenceCitation]] = {}
    for c in citations:
        if c.sample_id:
            by_sample.setdefault(c.sample_id, []).append(c)

    if not by_sample:
        return [], []

    step_level = [c for c in citations if not c.sample_id]
    return step_level, list(by_sample.items())


def _unique_sheet_name(wb, desired: str) -> str:
    """Excel sheet names must be unique and <=31 chars. Two test steps can
    each have a sample called "1", so fall back to a suffix rather than
    letting openpyxl silently rename or collide.
    """
    base = _sheet_title(desired)
    if base not in wb.sheetnames:
        return base
    for n in range(2, 100):
        candidate = f"{base[:28]}({n})"
        if candidate not in wb.sheetnames:
            return candidate
    return base[:28] + "(x)"


# Conditional fills matching how a reviewer scans a workpaper: green =
# clear, amber = look at it, red = deal with it.
_FILL_GOOD = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_WARN = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_BAD = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FONT_GOOD = Font(bold=True, color="006100")
_FONT_WARN = Font(bold=True, color="9C5700")
_FONT_BAD = Font(bold=True, color="9C0006")

_VERDICT_STYLE = {
    "satisfied": (_FILL_GOOD, _FONT_GOOD),
    "not_satisfied": (_FILL_BAD, _FONT_BAD),
    "insufficient_evidence": (_FILL_WARN, _FONT_WARN),
}
_CONFIDENCE_STYLE = {
    "high": (_FILL_GOOD, _FONT_GOOD),
    "medium": (_FILL_WARN, _FONT_WARN),
    "low": (_FILL_BAD, _FONT_BAD),
}
_IPE_STYLE = {
    "validated": (_FILL_GOOD, _FONT_GOOD),
    "not_validated": (_FILL_BAD, _FONT_BAD),
    "not_applicable": (None, None),
}


def _styled(ws, row: int, col: int, value: Any, style: tuple | None) -> None:
    cell = ws.cell(row, col, value)
    if style and style[0]:
        cell.fill = style[0]
    if style and style[1]:
        cell.font = style[1]


def _count_style(n: int) -> tuple:
    return (_FILL_GOOD, _FONT_GOOD) if n == 0 else (_FILL_WARN, _FONT_WARN)


def _link_cell_to_sheet(ws, row: int, col: int, sheet_name: str) -> None:
    """Makes an existing cell a clickable jump to a tab, keeping whatever
    text and colour it already carries -- the verdict cell stays a verdict,
    it just also navigates.
    """
    cell = ws.cell(row, col)
    cell.hyperlink = f"#'{sheet_name}'!A1"
    cell.font = Font(
        bold=cell.font.bold, color=cell.font.color, underline="single"
    )


def _sample_ids_for(result: dict) -> list[str]:
    """Sampled items in stable order: the model's own per-item results if it
    gave them, else whatever the citations were tagged with.
    """
    if "error" in result:
        return []
    conclusion: ConclusionOutput = result["conclusion"]
    if conclusion.sample_results:
        return [r.sample_id for r in conclusion.sample_results]
    seen: list[str] = []
    for c in conclusion.evidence_citations:
        if c.sample_id and c.sample_id not in seen:
            seen.append(c.sample_id)
    return seen


def _write_control_summary(ws, row: int, results: dict[str, dict]) -> int:
    """One row per SAMPLED ITEM, one column per test step.

    A row per test step could only ever say "the step failed" -- never
    which selection failed, which is the first thing a reviewer asks. This
    is the matrix an auditor actually reads: down the samples, across the
    steps, with each cell the verdict for that item on that step. Falls
    back to a row per step when nothing is sample-tagged.
    """
    step_ids = list(results)
    all_samples: list[str] = []
    for result in results.values():
        for sid in _sample_ids_for(result):
            if sid not in all_samples:
                all_samples.append(sid)

    if not all_samples:
        return _write_step_summary_rows(ws, row, results)

    # No separate "Detail sheet" column: with several steps a sample has a
    # sheet PER STEP, so one link per row could only ever reach the first
    # of them. The verdict cell is the link instead -- it already
    # identifies exactly one (sample, step) pair, which is exactly one
    # sheet, and it is what a reviewer would click anyway.
    headers = ["Sample"] + [f"Test step {s}" for s in step_ids]
    for col, name in enumerate(headers, 1):
        c = ws.cell(row, col, name)
        c.font = Font(bold=True)
        c.fill = _HEADER_FILL
    header_row = row
    row += 1

    for sid in all_samples:
        ws.cell(row, 1, sid).font = Font(bold=True)
        for i, step_id in enumerate(step_ids):
            result = results[step_id]
            if "error" in result:
                _styled(ws, row, 2 + i, "INCOMPLETE", (_FILL_BAD, _FONT_BAD))
                continue
            conclusion: ConclusionOutput = result["conclusion"]
            per_item = {r.sample_id: r for r in conclusion.sample_results}
            if sid in per_item:
                verdict = per_item[sid].conclusion
            elif sid in _sample_ids_for(result):
                # Not broken out by the model -- the step's own verdict is
                # the best available statement for this item.
                verdict = conclusion.conclusion
            else:
                ws.cell(row, 2 + i, "n/a")
                continue
            _styled(
                ws, row, 2 + i, _CONCLUSION_LABELS.get(verdict, verdict), _VERDICT_STYLE.get(verdict)
            )
            _link_cell_to_sheet(ws, row, 2 + i, _sample_sheet_name(step_id, sid, len(step_ids) == 1))
        row += 1

    # IPE gets its own row rather than being buried as one attribute among
    # many: it is a conclusion about the POPULATION, not about any sampled
    # item, and a reviewer scanning this table needs to see it at the same
    # level as the selections.
    ws.cell(row, 1, "IPE").font = Font(bold=True)
    for i, step_id in enumerate(step_ids):
        result = results[step_id]
        if "error" in result:
            _styled(ws, row, 2 + i, "INCOMPLETE", (_FILL_BAD, _FONT_BAD))
            continue
        status = result["conclusion"].ipe_completeness_accuracy_status
        _styled(ws, row, 2 + i, status, _IPE_STYLE.get(status))
    row += 1

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{row - 1}"
    row += 1

    # The per-step table below is only worth its space on a multi-step
    # control. With one step every column of it is already stated by the
    # matrix above or the CONCLUSION section below -- it was pure
    # duplication, and its "Test step" and "Conclusion" columns in
    # particular restated the matrix.
    if len(results) > 1:
        row = _write_step_summary_rows(ws, row, results)
    return row


def _write_step_summary_rows(ws, row: int, results: dict[str, dict]) -> int:
    """Per-step overview, for a control with more than one step. On a
    single-step control this is skipped: the matrix above and the
    CONCLUSION section below already carry every column of it.
    """
    headers = ["Test step", "Conclusion", "Confidence", "Sample coverage", "IPE status", "Exceptions", "Open requests"]
    for col, name in enumerate(headers, 1):
        c = ws.cell(row, col, name)
        c.font = Font(bold=True)
        c.fill = _HEADER_FILL
    row += 1
    for test_step_id, result in results.items():
        ws.cell(row, 1, test_step_id)
        if "error" in result:
            _styled(ws, row, 2, "INCOMPLETE -- run did not finish", (_FILL_BAD, _FONT_BAD))
        else:
            c: ConclusionOutput = result["conclusion"]
            _styled(
                ws,
                row,
                2,
                _CONCLUSION_LABELS.get(c.conclusion, c.conclusion),
                _VERDICT_STYLE.get(c.conclusion),
            )
            _styled(ws, row, 3, c.confidence, _CONFIDENCE_STYLE.get(c.confidence))
            if c.sample_coverage:
                sc = c.sample_coverage
                _styled(
                    ws,
                    row,
                    4,
                    f"{sc.total_found}/{sc.total_required}",
                    (_FILL_GOOD, _FONT_GOOD) if sc.complete else (_FILL_BAD, _FONT_BAD),
                )
            _styled(
                ws,
                row,
                5,
                c.ipe_completeness_accuracy_status,
                _IPE_STYLE.get(c.ipe_completeness_accuracy_status),
            )
            _styled(ws, row, 6, len(c.exceptions), _count_style(len(c.exceptions)))
            _styled(
                ws,
                row,
                7,
                len(c.additional_support_requests),
                _count_style(len(c.additional_support_requests)),
            )
        row += 1
    return row + 1


_ATTRIBUTE_STYLE = {
    "satisfied": (_FILL_GOOD, _FONT_GOOD),
    "not_satisfied": (_FILL_BAD, _FONT_BAD),
    "not_tested": (_FILL_WARN, _FONT_WARN),
}
_ATTRIBUTE_LABELS = {
    "satisfied": "Satisfied",
    "not_satisfied": "Not satisfied",
    "not_tested": "Not tested",
}


def _write_attribute_table(
    ws,
    row: int,
    attributes: list[Any],
    citations: list[EvidenceCitation],
    letters: list[str],
) -> int:
    """Attribute-by-attribute: what the step requires, what was observed,
    which tickmark proves it, and whether it holds.

    The tickmark column is the point of this table -- it chains an
    attribute to a cited evidence row and, through that letter, to the red
    box drawn on the exhibit page. Without it a reviewer has to read the
    whole narrative to find where any one attribute was satisfied.
    """
    if not attributes:
        return row

    # One evidence_id can be cited several times on a sheet -- three
    # different values quoted off the same invoice page get tickmarks A, B
    # and C. Keying by id alone silently kept only the last of them and
    # labelled every attribute "C". Collect them all instead: an attribute
    # citing that page legitimately points at each marked spot on it.
    letters_by_evidence: dict[str, list[str]] = {}
    for i, cit in enumerate(citations):
        if i < len(letters):
            letters_by_evidence.setdefault(cit.evidence_id, []).append(letters[i])

    for col in range(1, 7):
        ws.cell(row, col).fill = _SECTION_FILL
    ws.cell(row, 1, "ATTRIBUTES TESTED").font = _SECTION_FONT
    row += 2

    headers = ["Tickmark", "Attribute", "Result", "Value observed", "Evidence"]
    for col, name in enumerate(headers, 1):
        c = ws.cell(row, col, name)
        c.font = Font(bold=True)
        c.fill = _HEADER_FILL
    row += 1

    for attr in attributes:
        marks: list[str] = []
        for e in attr.evidence_ids:
            for letter in letters_by_evidence.get(e, []):
                if letter not in marks:
                    marks.append(letter)
        tick = ws.cell(row, 1, ", ".join(marks))
        tick.font = Font(bold=True, color="AA0000")
        ws.cell(row, 2, attr.attribute).alignment = Alignment(wrap_text=True, vertical="top")
        _styled(ws, row, 3, _ATTRIBUTE_LABELS.get(attr.result, attr.result), _ATTRIBUTE_STYLE.get(attr.result))
        ws.cell(row, 4, attr.value_observed).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 5, ", ".join(attr.evidence_ids)).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    return row + 1


def _write_evidence_rows(ws, row: int, citations: list[EvidenceCitation], letters: list[str]) -> int:
    """The evidence table itself. Returns the next free row."""
    headers = ["Tickmark", "Evidence ID", "Source file", "Location", "Quote / summary", "Relevance"]
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row, col, name)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
    row += 1
    for i, cit in enumerate(citations):
        letter = letters[i] if i < len(letters) else ""
        for col, value in enumerate(
            [letter, cit.evidence_id, cit.source_file, cit.location, cit.quote_or_summary, cit.relevance], 1
        ):
            c = ws.cell(row, col, value)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 1:
                c.font = Font(bold=True, color="AA0000")
        row += 1
    return row + 1


def _write_evidence_table(
    ws,
    row: int,
    citations: list[EvidenceCitation],
    letters: list[str],
    sample_id: str | None,
    sheet_title: str,
    attributes: list[Any] | None = None,
) -> None:
    """A per-sample sheet: which item this is, then that item's evidence
    with its own A/B/C tickmarks. The DRAFT banner and test-step header are
    already on the sheet -- the caller writes them for every sheet before
    branching here.
    """
    ws.cell(row, 1, "Sample").font = Font(bold=True)
    ws.cell(row, 2, str(sample_id or "")).font = Font(bold=True)
    row += 2

    # Attributes first: a reviewer signs off attribute by attribute, and
    # the evidence table below is what backs each one up.
    row = _write_attribute_table(ws, row, attributes or [], citations, letters)

    for col in range(1, 7):
        ws.cell(row, col).fill = _SECTION_FILL
    ws.cell(row, 1, "EVIDENCE CITED").font = _SECTION_FONT
    row += 2

    if not citations:
        ws.cell(row, 1, "No evidence was cited for this sampled item.").font = Font(italic=True)
        return

    row = _write_evidence_rows(ws, row, citations, letters)
    ws.cell(row, 1, f"Exhibits for these citations: see the '{sheet_title} - Exhibits' sheet.").font = Font(
        italic=True
    )


def _write_exhibit_sheet(
    ws,
    test_step_id: str,
    images: list[tuple[str, BytesIO, tuple[int, int], list[_Overlay]]],
    excerpts: list[tuple[str, list[str]]],
    live_buffers: list[BytesIO],
) -> None:
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    for col in "CDEF":
        ws.column_dimensions[col].width = 30

    from openpyxl.drawing.image import Image as XLImage

    row = 1
    ws.cell(row, 1, f"Evidence exhibits — test step {test_step_id}").font = Font(bold=True, size=12)
    row += 2

    # No tickmark legend here: it repeated the letter, quote and source
    # that the Evidence Cited table already carries, so a reviewer read the
    # same rows twice and had two places to keep in sync.
    ws.cell(row, 1, "Red boxes mark the cited value on the page. A letter in the corner means the page").font = Font(
        italic=True
    )
    row += 1
    ws.cell(row, 1, "is the exhibit but the specific value could not be located on it.").font = Font(italic=True)
    row += 2

    for caption, buf, (w, h), overlays in images:
        ws.cell(row, 1, caption).font = Font(bold=True)
        row += 1
        page_img = XLImage(buf)
        page_img.width, page_img.height = w, h
        ws.add_image(page_img, f"A{row}")
        live_buffers.append(buf)

        # Each box and each tickmark letter is its own picture, pinned to
        # the same cell as the page and offset into place. In Excel they
        # are ordinary shapes: a reviewer can drag a misplaced box onto the
        # right value, resize it, or delete it -- none of which is possible
        # once marks are burned into the page pixels.
        for ov in overlays:
            _place_overlay(ws, ov, anchor_row=row)
            live_buffers.append(ov.png)

        row += int(h / 19) + 3  # default row height ~19px at 96dpi

    for caption, lines in excerpts:
        ws.cell(row, 1, caption).font = Font(bold=True)
        row += 1
        for line in lines:
            ws.cell(row, 1, line).alignment = Alignment(vertical="top")
            row += 1
        row += 1


def _write_step_sheet(
    ws,
    test_step_id: str,
    test_step_text: str,
    result: dict,
    evidence_map: dict[str, EvidenceItem],
    support_dir: Path | None,
    citations: list[EvidenceCitation] | None = None,
    sample_id: str | None = None,
    spec: dict[str, Any] | None = None,
    results: dict[str, dict] | None = None,
    include_step_detail: bool = True,
) -> tuple[list[str], list[tuple[str, BytesIO, tuple[int, int], list[_Overlay]]], list[tuple[str, list[str]]]]:
    """Writes one sheet, ordered so a reviewer reads answers first
    (conclusion, coverage, IPE, exceptions, open requests) before the
    supporting detail. Returns the exhibit material for the caller to place
    on a separate sheet -- inline images used to push the conclusion ~110
    rows down the page.

    include_step_detail: True for the step's summary sheet (control header,
    conclusion, exceptions, open requests, documentation, procedures);
    False for a per-sample sheet, which carries only that item's evidence.
    spec/results: passed only for the workbook's first sheet, which carries
    the control header and the across-steps summary table.
    """
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100
    for col in "CDEF":
        ws.column_dimensions[col].width = 34

    row = 1

    def section(title: str) -> None:
        nonlocal row
        for col in range(1, 7):
            ws.cell(row, col).fill = _SECTION_FILL
        c = ws.cell(row, 1, title)
        c.font = _SECTION_FONT
        row += 2

    def put(label: str, value: str, *, bold: bool = False, color: str | None = None) -> None:
        nonlocal row
        ws.cell(row, 1, label).font = Font(bold=True)
        c = ws.cell(row, 2, value)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if bold or color:
            c.font = Font(bold=bold, color=color) if color else Font(bold=True)
        row += 1

    def put_block(text: str) -> None:
        nonlocal row
        c = ws.cell(row, 2, text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        row += 2

    def put_list(values: list[str], *, empty: str | None = None) -> None:
        nonlocal row
        if not values:
            if empty:
                ws.cell(row, 2, empty).font = Font(italic=True)
                row += 2
            return
        for v in values:
            c = ws.cell(row, 2, f"• {v}")
            c.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        row += 1

    ws.cell(row, 1, _DRAFT_BANNER).font = Font(bold=True, color="9C0006")
    row += 2

    ipe_shown_in_matrix = False
    # Control header, on the workbook's first sheet only -- it used to sit
    # on a separate Summary tab a reviewer had to click away from.
    if spec is not None:
        put("Control ID", str(spec.get("control_id", "")))
        put("Control objective ref", str(spec.get("control_objective_ref", "")))
        put("Control objective", str(spec.get("control_objective_text", "")))
        row += 1
        # Shown even on a single-step control: with a per-sample matrix
        # this is no longer a restatement of the sheet below it -- it is
        # the only place that says how each individual selection came out.
        if results is not None:
            row = _write_control_summary(ws, row, results)
            # The matrix carries an IPE row whenever it renders, so the
            # CONCLUSION block below must not restate it.
            ipe_shown_in_matrix = any(_sample_ids_for(r) for r in results.values())

    put("Test step", test_step_id)
    put("Test step text", test_step_text)
    row += 1

    if "error" in result:
        section("RESULT — INCOMPLETE")
        put("Status", "INCOMPLETE -- run did not finish", bold=True, color="9C0006")
        put("Error", str(result["error"]))
        if "reason" in result:
            put("Abort reason", str(result["reason"]))
            put("Cost-weighted tokens used", f"{result.get('tokens_used', 0):,}")
        audit_log = result.get("audit_log") or []
        if audit_log:
            put("Tool calls before abort", str(len(audit_log)))
            row += 1
            section("SEARCHES ATTEMPTED BEFORE ABORT")
            put_list(
                [
                    str(e.input.get("query", ""))
                    for e in audit_log
                    if e.tool_name == "search_cy_support" and e.input.get("query")
                ]
            )
        return [], [], []

    conclusion: ConclusionOutput = result["conclusion"]
    sheet_citations = conclusion.evidence_citations if citations is None else citations

    # Tickmarks are lettered per SHEET, so each sampled item starts at A
    # instead of continuing a running sequence from earlier items. Lettered
    # unconditionally: they label the rows of the evidence table, so a
    # workpaper built without exhibits (no support files to hand) still
    # gets A/B/C rather than a blank column.
    letters: list[str] = _tickmark_letters(len(sheet_citations))
    exhibit_images: list[tuple[str, BytesIO, tuple[int, int], list[_Overlay]]] = []
    excerpts: list[tuple[str, list[str]]] = []
    if sheet_citations and evidence_map and support_dir is not None:
        letters, exhibit_images, excerpts = _build_step_exhibits(
            sheet_citations, evidence_map, support_dir
        )

    if not include_step_detail:
        # Per-sample sheet: this item's evidence only. The conclusion,
        # narrative and procedures are step-wide and live on the summary
        # sheet -- repeating them per item is noise, not documentation.
        _write_evidence_table(
            ws,
            row,
            sheet_citations,
            letters,
            sample_id,
            ws.title,
            attributes=[a for a in conclusion.attribute_results if a.sample_id == sample_id],
        )
        return letters, exhibit_images, excerpts

    # ── Answers first ──────────────────────────────────────────────────
    section("CONCLUSION")
    put(
        "Conclusion",
        _CONCLUSION_LABELS.get(conclusion.conclusion, conclusion.conclusion),
        bold=True,
        color=_CONCLUSION_COLORS.get(conclusion.conclusion),
    )
    put("Confidence", conclusion.confidence)
    put("Confidence rationale", conclusion.confidence_rationale)
    if conclusion.sample_coverage:
        sc = conclusion.sample_coverage
        put(
            "Sample coverage",
            f"{sc.total_found} of {sc.total_required} ({sc.coverage_pct}%)"
            + (f"; missing: {', '.join(sc.missing)}" if sc.missing else ""),
        )
    if not ipe_shown_in_matrix:
        put("IPE status", conclusion.ipe_completeness_accuracy_status)
    row += 1

    # Attributes sit here, straight after the conclusion, for the same
    # reason they lead each sampled item's sheet: they are the structured
    # "where is this satisfied" view. They used to render after
    # DOCUMENTATION and PROCEDURES, so the summary contradicted the sample
    # sheets on where a reviewer should look first.
    summary_attributes = [
        a
        for a in conclusion.attribute_results
        if a.sample_id is None or (sheet_citations and citations is None)
    ]
    if summary_attributes:
        row = _write_attribute_table(ws, row, summary_attributes, sheet_citations, letters)

    section("EXCEPTIONS")
    put_list(conclusion.exceptions, empty="None noted.")

    section("ADDITIONAL SUPPORT REQUESTED")
    put_list(conclusion.additional_support_requests, empty="None — testing is complete on the evidence provided.")

    # ── Supporting detail ──────────────────────────────────────────────
    section("DOCUMENTATION")
    put_block(conclusion.narrative)

    section("PROCEDURES PERFORMED")
    put_list(conclusion.procedures_performed)

    if conclusion.ipe_completeness_accuracy_evidence:
        section("IPE COMPLETENESS & ACCURACY EVIDENCE")
        put_list(conclusion.ipe_completeness_accuracy_evidence)

    if sheet_citations:
        section("EVIDENCE CITED")
        row = _write_evidence_rows(ws, row, sheet_citations, letters)
        if exhibit_images or excerpts:
            ws.cell(row, 1, f"Exhibits for these citations: see the '{ws.title} - Exhibits' sheet.").font = Font(
                italic=True
            )
            row += 2

    section("PREPARED BY")
    md = conclusion.model_metadata
    put("Prepared by", f"CY testing agent ({md.model}, prompt {md.prompt_version})")
    put("Timestamp", md.timestamp)
    put("Tool calls", str(md.tool_call_count))

    return letters, exhibit_images, excerpts


# --------------------------------------------------------------------------
# PDF
# --------------------------------------------------------------------------


def _build_pdf(
    spec: dict[str, Any],
    results: dict[str, dict],
    step_texts: dict[str, str],
    out_path: Path,
    evidence_map: dict[str, EvidenceItem],
    support_dir: Path | None,
) -> None:
    # Imported here, not module-level: reportlab is only needed when PY was a
    # PDF, and keeping the import local means an Excel-only user without it
    # installed can still generate xlsx workpapers.
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    body = styles["BodyText"]
    h1, h2 = styles["Heading1"], styles["Heading2"]
    draft = Paragraph(f"<b>{_DRAFT_BANNER}</b>", body)

    story: list[Any] = [draft, Spacer(1, 8)]
    story.append(Paragraph(f"Control {spec['control_id']} -- CY Testing", h1))
    story.append(
        Paragraph(
            f"<b>Objective ({spec.get('control_objective_ref', '')}):</b> "
            f"{spec.get('control_objective_text', '')}",
            body,
        )
    )
    story.append(Spacer(1, 12))

    def bullet_list(label: str, values: list[str]) -> None:
        if not values:
            return
        story.append(Paragraph(f"<b>{label}:</b>", body))
        for v in values:
            story.append(Paragraph(f"• {v}", body))
        story.append(Spacer(1, 6))

    for test_step_id, result in results.items():
        story.append(Paragraph(f"Test step {test_step_id}", h2))
        step_text = step_texts.get(test_step_id, "")
        if step_text:
            story.append(Paragraph(f"<b>Test step:</b> {step_text}", body))

        if "error" in result:
            story.append(Paragraph("<b>Status: INCOMPLETE -- run did not finish</b>", body))
            story.append(Paragraph(f"Error: {result['error']}", body))
            if "reason" in result:
                story.append(
                    Paragraph(
                        f"Abort reason: {result['reason']} ({result.get('tokens_used', 0):,} tokens used)", body
                    )
                )
            story.append(Spacer(1, 12))
            continue

        conclusion: ConclusionOutput = result["conclusion"]
        story.append(
            Paragraph(
                f"<b>Conclusion:</b> {_CONCLUSION_LABELS.get(conclusion.conclusion, conclusion.conclusion)} "
                f"(confidence: {conclusion.confidence})",
                body,
            )
        )
        story.append(Paragraph(f"<b>Documentation:</b> {conclusion.narrative}", body))
        story.append(Spacer(1, 6))
        bullet_list("Procedures performed", conclusion.procedures_performed)

        letters: list[str] = []
        exhibit_images: list[tuple[str, BytesIO, tuple[int, int], list[_Overlay]]] = []
        excerpts: list[tuple[str, list[str]]] = []
        if conclusion.evidence_citations and evidence_map and support_dir is not None:
            letters, exhibit_images, excerpts = _build_step_exhibits(
                conclusion.evidence_citations, evidence_map, support_dir
            )

        if conclusion.evidence_citations:
            rows = [["Tickmark", "Evidence ID", "Source file", "Location", "Quote / summary"]] + [
                [
                    Paragraph(f"<b><font color='red'>{letters[i] if i < len(letters) else ''}</font></b>", body),
                    Paragraph(cit.evidence_id, body),
                    Paragraph(cit.source_file, body),
                    Paragraph(cit.location, body),
                    Paragraph(cit.quote_or_summary, body),
                ]
                for i, cit in enumerate(conclusion.evidence_citations)
            ]
            table = Table(rows, colWidths=[0.7 * inch, 0.8 * inch, 1.4 * inch, 1.6 * inch, 2.5 * inch])
            table.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            story.append(table)
            story.append(Spacer(1, 6))

        if exhibit_images or excerpts:
            from reportlab.platypus import Image as RLImage

            story.append(Paragraph("<b>Evidence exhibits:</b>", body))
            for caption, buf, (w, h), overlays in exhibit_images:
                story.append(Paragraph(f"<i>{caption}</i>", body))
                display_w = min(6.5 * inch, w)
                # A PDF is a flat rendering by definition -- nothing in it
                # is movable -- so the marks are composited back onto the
                # page here. The xlsx output keeps them as separate,
                # draggable pictures instead.
                story.append(RLImage(_flatten(buf, overlays), width=display_w, height=h * (display_w / w)))
                story.append(Spacer(1, 8))
            for caption, lines in excerpts:
                story.append(Paragraph(f"<i>{caption}</i>", body))
                for line in lines:
                    story.append(Paragraph(line, body))
                story.append(Spacer(1, 8))

        if conclusion.sample_coverage:
            sc = conclusion.sample_coverage
            story.append(
                Paragraph(
                    f"<b>Sample coverage:</b> {sc.total_found} of {sc.total_required} ({sc.coverage_pct}%)"
                    + (f"; missing: {', '.join(sc.missing)}" if sc.missing else ""),
                    body,
                )
            )
        story.append(Paragraph(f"<b>IPE status:</b> {conclusion.ipe_completeness_accuracy_status}", body))
        bullet_list("IPE C&A evidence", conclusion.ipe_completeness_accuracy_evidence)
        bullet_list("Exceptions", conclusion.exceptions)
        bullet_list("Additional support requested", conclusion.additional_support_requests)

        md = conclusion.model_metadata
        story.append(
            Paragraph(
                f"<i>Prepared by CY testing agent ({md.model}, prompt {md.prompt_version}) -- "
                f"{md.timestamp}; {md.tool_call_count} tool calls.</i>",
                body,
            )
        )
        story.append(Spacer(1, 16))

    SimpleDocTemplate(str(out_path), pagesize=letter).build(story)
