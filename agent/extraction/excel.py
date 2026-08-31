"""Excel extraction: openpyxl -> EvidenceItem list.

Per the design doc, each worksheet produces two kinds of EvidenceItem:

1. ``excel_table`` -- one per detected table block (contiguous non-empty
   rows, first row treated as the header). This is what search_cy_support
   mostly returns.
2. ``excel_cell`` -- one per cell that carries a fill color or a comment.
   Preparers color-code exceptions and leave comments; that's real signal a
   plain table dump would discard, so it gets its own EvidenceItem rather
   than being silently flattened into the table text. This is deliberately
   NOT one EvidenceItem per cell -- only annotated cells, to keep volume
   sane on large sheets.

extraction_confidence is always 1.0 here: this is structured data read
directly from the file, not inferred from an image.
"""

from __future__ import annotations

import itertools
from pathlib import Path
from typing import Iterator

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from agent.schemas import EvidenceItem


def extract_excel(path: str | Path) -> list[EvidenceItem]:
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    filename = path.name
    counter = itertools.count(1)

    items: list[EvidenceItem] = []
    for ws in wb.worksheets:
        items.extend(_extract_tables(ws, filename, counter))
        items.extend(_extract_annotated_cells(ws, filename, counter))
        items.extend(_extract_images(ws, filename, counter))
    return items


def _extract_images(
    ws: Worksheet, filename: str, counter: Iterator[int]
) -> list[EvidenceItem]:
    """Pasted screenshots living on a worksheet.

    These were invisible before: extraction read cells only, so a tab whose
    entire content is a pasted image came through as an EMPTY sheet and
    never entered the evidence pool at all. That is not a rare shape -- a
    real IPE "Parameters" tab was exactly this: a screenshot of the report
    query showing the filters and "Selected rows: 57,039". The agent could
    not compare that record count against the 30-row population extract
    because it never saw the tab existed.

    Emitted as image_ocr placeholders, the same shape a scanned PDF page
    produces, so the existing OCR pass picks them up and transcribes them.
    """
    items: list[EvidenceItem] = []
    for idx, image in enumerate(getattr(ws, "_images", []), start=1):
        items.append(
            EvidenceItem(
                evidence_id=f"ev_{next(counter):04d}",
                source_file=filename,
                source_type="image_ocr",
                location=f"{ws.title}!image{idx}",
                extracted_text=None,
                extracted_table=None,
                extraction_confidence=0.0,
                preview_ref=f"{filename}!{ws.title}!image{idx}",
            )
        )
    return items


def _row_is_empty(row) -> bool:
    return all(cell.value in (None, "") for cell in row)


_MAX_DATA_ROWS_PER_TABLE = 200
# A contiguous non-empty block larger than this becomes several EvidenceItems
# instead of one. Without this, a real export with thousands of unbroken
# rows (a full population listing, no blank-row breaks) becomes a single
# giant table -- and since a human-written PY testing workpaper's excerpts
# get rendered into every prompt turn with no truncation of their own (see
# agent/loop.py's PY-excerpt budget), one such item is enough to blow a
# single test step's token usage up by two orders of magnitude. This caught
# exactly that on a real run.


def _extract_tables(
    ws: Worksheet, filename: str, counter: Iterator[int]
) -> list[EvidenceItem]:
    items: list[EvidenceItem] = []
    block: list[tuple[int, tuple]] = []  # (row_idx, cells) accumulated for current block

    def to_str_row(cells, max_col: int) -> list[str]:
        return [("" if c.value is None else str(c.value)) for c in cells] + [""] * (max_col - len(cells))

    def flush_block():
        if not block:
            return
        header_row_idx, header_cells = block[0]
        data = block[1:]
        max_col = max(len(cells) for _, cells in block)
        header_str = to_str_row(header_cells, max_col)
        col_letter = get_column_letter(max_col)

        if len(data) <= _MAX_DATA_ROWS_PER_TABLE:
            end_row = block[-1][0]
            table = [header_str] + [to_str_row(cells, max_col) for _, cells in data]
            location = f"{ws.title}!A{header_row_idx}:{col_letter}{end_row}"
            _emit(table, location)
            return

        for i in range(0, len(data), _MAX_DATA_ROWS_PER_TABLE):
            piece = data[i : i + _MAX_DATA_ROWS_PER_TABLE]
            chunk_start, chunk_end = piece[0][0], piece[-1][0]
            table = [header_str] + [to_str_row(cells, max_col) for _, cells in piece]
            location = f"{ws.title}!A{header_row_idx}+A{chunk_start}:{col_letter}{chunk_end}"
            _emit(table, location)

    def _emit(table: list[list[str]], location: str) -> None:
        items.append(
            EvidenceItem(
                evidence_id=f"ev_{next(counter):04d}",
                source_file=filename,
                source_type="excel_table",
                location=location,
                extracted_text=None,
                extracted_table=table,
                extraction_confidence=1.0,
                preview_ref=f"{filename}!{location}",
            )
        )

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        # Trim trailing fully-empty cells so max_col reflects real content.
        cells = list(row)
        while cells and cells[-1].value in (None, ""):
            cells.pop()

        if not cells:
            flush_block()
            block = []
            continue

        block.append((row_idx, tuple(cells)))

    flush_block()
    return items


# Real preparer convention seen in a live run: the ENTIRE row of a sampled
# population item gets highlighted (every column, same color), not one cell.
# The naive per-cell logic below turned one such row into ~20 separate
# EvidenceItems like "value: 'PK'; fill_color: FFFFFF00" -- individually
# meaningless (no column label, no relation to the other 19 fragments of the
# same row) and already redundant with the row's own full content, which
# _extract_tables already captured as real text. On a real run this flooded
# the evidence pool (57 fragments from 3 highlighted rows) and the model
# burned its whole tool-call budget trying to chase them down as if they
# were distinct pieces of evidence instead of realizing they're one marked
# row. A row where most of its populated cells share the SAME highlight
# color collapses to a single "this row is marked" EvidenceItem instead.
_WHOLE_ROW_FRACTION = 0.6
_WHOLE_ROW_MIN_CELLS = 3


def _extract_annotated_cells(
    ws: Worksheet, filename: str, counter: Iterator[int]
) -> list[EvidenceItem]:
    items: list[EvidenceItem] = []

    for row in ws.iter_rows():
        non_empty = [c for c in row if c.value is not None]
        if not non_empty:
            continue

        flagged: list[tuple] = []  # (cell, fill_color, comment_text)
        for cell in non_empty:
            comment_text = cell.comment.text.strip() if cell.comment else None
            fill_color = _fill_color(cell)
            if comment_text is not None or fill_color is not None:
                flagged.append((cell, fill_color, comment_text))

        if not flagged:
            continue

        # Commented cells are always individually meaningful (a preparer
        # wrote something specific there) -- never absorbed into a row
        # summary, whole-row-colored or not.
        commented = [f for f in flagged if f[2] is not None]
        colorable = [f for f in flagged if f[2] is None and f[1] is not None]

        colors = {f[1] for f in colorable}
        is_whole_row = (
            len(colors) == 1
            and len(colorable) >= _WHOLE_ROW_MIN_CELLS
            and len(colorable) / len(non_empty) >= _WHOLE_ROW_FRACTION
        )

        if is_whole_row:
            color = next(iter(colors))
            first_cell, last_cell = colorable[0][0], colorable[-1][0]
            items.append(
                EvidenceItem(
                    evidence_id=f"ev_{next(counter):04d}",
                    source_file=filename,
                    source_type="excel_cell",
                    location=f"{ws.title}!{first_cell.coordinate}:{last_cell.coordinate}",
                    extracted_text=(
                        f"Row {first_cell.row}: entire row highlighted (fill_color: {color}) across "
                        f"{len(colorable)} of {len(non_empty)} populated columns -- likely marks a "
                        "selected/sampled item within this population. See the table extraction for "
                        "this row's full field values."
                    ),
                    extracted_table=None,
                    extraction_confidence=1.0,
                    preview_ref=f"{filename}!{ws.title}!row{first_cell.row}",
                )
            )
        else:
            commented = flagged  # no whole-row collapse -- every flagged cell stands alone as before

        for cell, fill_color, comment_text in commented:
            parts = [f"value: {cell.value!r}"]
            if fill_color:
                parts.append(f"fill_color: {fill_color}")
            if comment_text:
                parts.append(f"comment: {comment_text}")

            items.append(
                EvidenceItem(
                    evidence_id=f"ev_{next(counter):04d}",
                    source_file=filename,
                    source_type="excel_cell",
                    location=f"{ws.title}!{cell.coordinate}",
                    extracted_text="; ".join(parts),
                    extracted_table=None,
                    extraction_confidence=1.0,
                    preview_ref=f"{filename}!{ws.title}!{cell.coordinate}",
                )
            )
    return items


def _fill_color(cell) -> str | None:
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    rgb = getattr(fg, "rgb", None)
    if isinstance(rgb, str) and rgb not in ("00000000",):
        return rgb
    return None
