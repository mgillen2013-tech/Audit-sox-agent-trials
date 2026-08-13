"""Parses the CY sample list -- the intake input from design doc section 1
that previously had no code behind it (SamplePopulationManifest existed as
a schema, but nothing built it from an actual file).

Two entry points, sharing one core:
- parse_sample_list(path): reads an uploaded Excel file.
- build_manifests_from_rows(rows): builds the same result from plain
  row dicts -- used by the Streamlit app, which lets a user type/edit
  samples in a web table instead of building an Excel file by hand.

Required fields per row (column names in the Excel case; dict keys in the
row-dict case), matched case-insensitively for the Excel path:
    test_step_id          -- which test step this sample belongs to
    sample_id              -- "S01".."S25", unique within a test step
    identifying_details    -- free text describing the item
    population_description -- same value repeated on every row for a given
                               test_step_id, e.g. "All POs > $5,000 issued
                               Oct 2025-Sep 2026"
    selection_method        -- one of: random, haphazard, judgmental, all_items

Optional field:
    population_size         -- leave blank if unknown

Any other field (e.g. po_number, branch, date, amount) is captured per row
as SampleItem.key_fields, so search_cy_support can be pointed at a specific
field instead of a fuzzy description match.

sample_size is NOT a field -- it's computed as the row count per
test_step_id, so it can never drift out of sync with the actual sample list
the way a manually-typed count could.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from agent.schemas import SampleItem, SamplePopulationManifest

_REQUIRED_COLUMNS = {
    "test_step_id",
    "sample_id",
    "identifying_details",
    "population_description",
    "selection_method",
}
_KNOWN_COLUMNS = _REQUIRED_COLUMNS | {"population_size"}
_SELECTION_METHODS = ("random", "haphazard", "judgmental", "all_items")


def _clean(value: Any) -> str | None:
    if value is None or value == "":
        return None
    return str(value).strip()


def build_manifests_from_rows(rows: list[dict[str, Any]]) -> dict[str, SamplePopulationManifest]:
    """rows: each a dict with at least the required keys above (any casing
    of key is fine as long as it's consistent -- callers building rows
    themselves, like the Streamlit app, should use the exact lowercase
    field names). Extra keys become SampleItem.key_fields.
    """
    grouped: dict[str, list[SampleItem]] = {}
    population_info: dict[str, dict[str, Any]] = {}

    for i, row in enumerate(rows):
        test_step_id = _clean(row.get("test_step_id"))
        sample_id = _clean(row.get("sample_id"))
        identifying_details = _clean(row.get("identifying_details"))
        if not test_step_id or not sample_id or not identifying_details:
            raise ValueError(f"row {i + 1} is missing test_step_id, sample_id, or identifying_details")

        key_fields = {k: _clean(v) for k, v in row.items() if k not in _KNOWN_COLUMNS}
        key_fields = {k: v for k, v in key_fields.items() if v is not None} or None

        grouped.setdefault(test_step_id, []).append(
            SampleItem(
                sample_id=sample_id,
                test_step_id=test_step_id,
                identifying_details=identifying_details,
                key_fields=key_fields,
            )
        )

        if test_step_id not in population_info:
            selection_method = _clean(row.get("selection_method"))
            if selection_method not in _SELECTION_METHODS:
                raise ValueError(
                    f"test_step_id {test_step_id!r} has selection_method {selection_method!r}, "
                    f"must be one of {'/'.join(_SELECTION_METHODS)}"
                )
            pop_size_raw = row.get("population_size")
            population_info[test_step_id] = {
                "population_description": _clean(row.get("population_description")) or "",
                "population_size": int(pop_size_raw) if pop_size_raw not in (None, "") else None,
                "selection_method": selection_method,
            }

    manifests: dict[str, SamplePopulationManifest] = {}
    for test_step_id, samples in grouped.items():
        info = population_info[test_step_id]
        manifests[test_step_id] = SamplePopulationManifest(
            test_step_id=test_step_id,
            population_description=info["population_description"],
            population_size=info["population_size"],
            sample_size=len(samples),
            selection_method=info["selection_method"],
            samples=samples,
        )
    return manifests


def parse_sample_list(path: str | Path) -> dict[str, SamplePopulationManifest]:
    """Returns {test_step_id: SamplePopulationManifest}, reading from an
    uploaded Excel file (one sheet, header row + one row per sample).
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    all_rows = list(ws.iter_rows())
    if not all_rows:
        raise ValueError(f"{path.name}: sheet is empty")

    header_row = all_rows[0]
    col_index: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell.value is None:
            continue
        col_index[str(cell.value).strip().lower()] = idx

    missing = _REQUIRED_COLUMNS - set(col_index)
    if missing:
        raise ValueError(
            f"{path.name}: missing required column(s) {sorted(missing)}. "
            f"Found columns: {sorted(col_index)}"
        )

    rows: list[dict[str, Any]] = []
    for row in all_rows[1:]:
        if all(c.value in (None, "") for c in row):
            continue
        rows.append({name: (row[idx].value if idx < len(row) else None) for name, idx in col_index.items()})

    try:
        return build_manifests_from_rows(rows)
    except ValueError as exc:
        raise ValueError(f"{path.name}: {exc}") from exc


# --------------------------------------------------------------------------
# Arbitrary-column path -- for real sample/population exports that don't
# (and shouldn't have to) match the clean column names above. A real E1
# AP payment export, for example, has columns like "invoice number
# f0411.vinv" and "check/ item f0413.docm" -- no test_step_id, sample_id,
# identifying_details, population_description, or selection_method concept
# anywhere in it, because those are audit judgments, not data the source
# system tracks. Forcing a rename before upload is exactly the kind of
# friction this app exists to remove.
# --------------------------------------------------------------------------

_SAMPLE_ID_CANDIDATES = {
    "sample selection #",
    "sample #",
    "sample no",
    "sample no.",
    "sample id",
    "sampleid",
    "sample_id",
    "sample number",
    "id",
    "item",
    "item #",
    "item no",
    "line",
    "line #",
    "line no",
    "#",
}


def read_excel_rows(path: str | Path) -> list[dict[str, Any]]:
    """Reads a plain Excel sheet into row dicts keyed by header text, with
    no required columns -- feeds build_manifest_from_any_columns().
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    all_rows = list(ws.iter_rows())
    if not all_rows:
        return []

    header_row = all_rows[0]
    headers = [
        str(cell.value).strip() if cell.value is not None else f"column_{idx + 1}"
        for idx, cell in enumerate(header_row)
    ]

    rows: list[dict[str, Any]] = []
    for row in all_rows[1:]:
        if all(c.value in (None, "") for c in row):
            continue
        rows.append({headers[idx]: (row[idx].value if idx < len(row) else None) for idx in range(len(headers))})
    return rows


def build_manifest_from_any_columns(
    rows: list[dict[str, Any]],
    test_step_id: str,
    population_description: str,
    selection_method: str,
    population_size: int | None = None,
) -> SamplePopulationManifest:
    """One file = one test step's sample here (matches how a real
    population/sample export naturally exists -- one export per test, not
    one file spanning several). sample_id is auto-detected from a
    likely-looking id column (e.g. "Sample Selection #"); if none is found,
    falls back to the row's position. identifying_details and key_fields
    are built from every column present, since in an export like this every
    column genuinely is an identifying transaction field.
    """
    if selection_method not in _SELECTION_METHODS:
        raise ValueError(
            f"selection_method {selection_method!r} must be one of {'/'.join(_SELECTION_METHODS)}"
        )

    id_col = None
    if rows:
        for col in rows[0]:
            normalized = _clean(col)
            if normalized and normalized.lower() in _SAMPLE_ID_CANDIDATES:
                id_col = col
                break

    samples: list[SampleItem] = []
    for i, row in enumerate(rows, start=1):
        cleaned = {k: _clean(v) for k, v in row.items()}
        cleaned = {k: v for k, v in cleaned.items() if v is not None}
        if not cleaned:
            continue

        sample_id = cleaned.get(id_col) if id_col else None
        if not sample_id:
            sample_id = f"row_{i}"

        identifying_details = "; ".join(f"{k}: {v}" for k, v in cleaned.items())

        samples.append(
            SampleItem(
                sample_id=str(sample_id),
                test_step_id=test_step_id,
                identifying_details=identifying_details,
                key_fields=cleaned,
            )
        )

    return SamplePopulationManifest(
        test_step_id=test_step_id,
        population_description=population_description,
        population_size=population_size,
        sample_size=len(samples),
        selection_method=selection_method,
        samples=samples,
    )
